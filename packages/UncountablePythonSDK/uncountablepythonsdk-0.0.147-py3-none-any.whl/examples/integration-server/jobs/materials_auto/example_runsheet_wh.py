from io import BytesIO

from openpyxl import Workbook, load_workbook
from uncountable.core.file_upload import DataFileUpload, FileUpload
from uncountable.integration.job import JobArguments, RunsheetWebhookJob, register_job
from uncountable.types import (
    download_file_t,
    entity_t,
    export_default_runsheet_t,
    identifier_t,
    webhook_job_t,
)
from uncountable.types.client_base import APIRequest

from pkgs.serialization_util import serialize_for_storage

RUNSHEET_REF_NAME = "recipe_export_runsheet"
RUNSHEET_REF_NAME_2 = "recipe_export_runsheet_2"


@register_job
class StandardRunsheetGenerator(RunsheetWebhookJob):
    def build_runsheet(
        self,
        *,
        args: JobArguments,
        payload: webhook_job_t.RunsheetWebhookPayload,
    ) -> FileUpload:
        args.logger.log_info("Exporting default runsheets")

        entity_identifiers: list[identifier_t.IdentifierKey] = [
            identifier_t.IdentifierKeyId(id=entity.id) for entity in payload.entities
        ]

        combined_wb = Workbook()
        combined_sheet = combined_wb.active or combined_wb.create_sheet(
            title="Combined Runsheet"
        )
        combined_sheet.title = "Combined Runsheet"

        for ref_name in [RUNSHEET_REF_NAME, RUNSHEET_REF_NAME_2]:
            api_request = APIRequest(
                method=export_default_runsheet_t.ENDPOINT_METHOD,
                endpoint=export_default_runsheet_t.ENDPOINT_PATH,
                args=export_default_runsheet_t.Arguments(
                    entities=entity_identifiers,
                    runsheet_key=identifier_t.IdentifierKeyRefName(ref_name=ref_name),
                    entity_type=payload.entities[0].type
                    if payload.entities
                    else entity_t.EntityType.RECIPE,
                ),
            )

            response = args.client.do_request(
                api_request=api_request,
                return_type=export_default_runsheet_t.Data,
            )

            print(
                "--------------------------------"
                f"Runsheet metadata: {serialize_for_storage(response.default_runsheet_metadata.columns)}, "
                "--------------------------------"
            )

            file_query = download_file_t.FileDownloadQueryTextDocumentId(
                text_document_id=response.text_document_id,
            )

            downloaded_files = args.client.download_files(file_query=file_query)
            file_data = downloaded_files[0].data.read()

            wb = load_workbook(filename=BytesIO(file_data))
            for sheet_name in wb.sheetnames:
                for row in wb[sheet_name].iter_rows(values_only=True):
                    combined_sheet.append(row)

        output = BytesIO()
        combined_wb.save(output)
        output.seek(0)

        return DataFileUpload(
            data=output,
            name="combined_runsheet.xlsx",
        )
