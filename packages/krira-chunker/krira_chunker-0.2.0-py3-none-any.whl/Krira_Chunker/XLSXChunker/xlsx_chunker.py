"""
XLSX chunker for Krira_Chunker.
"""

import os
from typing import Generator, Dict, Any, List, Tuple, Iterator

from ..config import ChunkConfig
from ..core import FastChunker, LOGGER
from ..exceptions import DependencyNotInstalledError, FileSizeLimitError


def _xlsx_row_to_text(headers: List[str], row: Tuple[Any, ...]) -> str:
    """Convert row to text representation."""
    parts = []
    for h, v in zip(headers, row):
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        parts.append(f"{h}: {s}")
    return " | ".join(parts)


class XLSXChunker:
    """
    Class-based XLSX chunker with lazy openpyxl loading.
    
    Uses streaming read mode for memory efficiency.
    Each row is treated as an atomic unit.
    
    Example:
        >>> cfg = ChunkConfig(max_chars=2000)
        >>> chunker = XLSXChunker(cfg)
        >>> for chunk in chunker.chunk_file("data.xlsx"):
        ...     print(chunk["text"][:100])
    """
    
    def __init__(self, cfg: ChunkConfig = None):
        """
        Initialize XLSX chunker.
        
        Args:
            cfg: Chunk configuration. Uses defaults if None.
        """
        self.cfg = cfg or ChunkConfig()
        self._chunker = None
    
    @property
    def chunker(self) -> FastChunker:
        """Lazy-load FastChunker."""
        if self._chunker is None:
            self._chunker = FastChunker(self.cfg)
        return self._chunker
    
    def _get_openpyxl(self):
        """Lazy import openpyxl."""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise DependencyNotInstalledError("openpyxl", "xlsx", "Excel processing")
    
    def _get_polars(self):
        """Lazy import polars (fallback)."""
        try:
            import polars as pl
            return pl
        except ImportError:
            return None
    
    def chunk_file(self, file_path: str) -> Iterator[Dict[str, Any]]:
        """
        Chunk an XLSX file.
        
        Each row is treated as an atomic unit that will not be split.
        
        Args:
            file_path: Path to XLSX file.
            
        Yields:
            Chunk dictionaries with row and sheet metadata.
            
        Raises:
            DependencyNotInstalledError: If openpyxl not installed.
            FileSizeLimitError: If file exceeds size limit.
        """
        cfg = self.cfg
        
        # Security: check file size
        if os.path.exists(file_path):
            size = os.path.getsize(file_path)
            if size > cfg.security_max_file_bytes:
                raise FileSizeLimitError(file_path, size, cfg.security_max_file_bytes)
        
        base_meta = {
            "source": os.path.basename(file_path),
            "source_path": os.path.abspath(file_path),
            "source_type": "xlsx",
        }
        
        # Try openpyxl first (streaming)
        try:
            openpyxl = self._get_openpyxl()
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            chunk_index = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                meta_sheet = dict(base_meta)
                meta_sheet["sheet"] = sheet_name
                
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue
                
                headers = []
                for i, h in enumerate(header_row):
                    hs = str(h).strip() if h is not None and str(h).strip() else f"col_{i+1}"
                    headers.append(hs)
                
                batch_units: List[str] = []
                batch_row_ids: List[int] = []
                prev_tail_units: List[str] = []
                prev_tail_ids: List[int] = []
                
                row_i = 0
                for row in rows_iter:
                    row_i += 1
                    row_text = _xlsx_row_to_text(headers, row)
                    if row_text and len(row_text) >= cfg.min_chars:
                        batch_units.append(row_text)
                        batch_row_ids.append(row_i)
                    
                    if len(batch_units) >= cfg.xlsx_batch_rows:
                        # Add overlap from previous batch
                        units = (prev_tail_units + batch_units) if prev_tail_units else batch_units
                        ids = (prev_tail_ids + batch_row_ids) if prev_tail_ids else batch_row_ids
                        
                        for ch in self.chunker.chunk_units(
                            units=units,
                            base_meta=meta_sheet,
                            joiner="\n",
                            locator=f"xlsx|sheet={sheet_name}",
                            range_key="row",
                            range_vals=ids,
                            start_chunk_index=chunk_index,
                        ):
                            chunk_index = ch["metadata"]["chunk_index"] + 1
                            yield ch
                        
                        # Compute tail overlap for next batch
                        k = min(cfg.xlsx_batch_overlap_rows, len(batch_units))
                        prev_tail_units = batch_units[-k:]
                        prev_tail_ids = batch_row_ids[-k:]
                        
                        batch_units, batch_row_ids = [], []
                
                # Flush remainder
                if batch_units:
                    units = (prev_tail_units + batch_units) if prev_tail_units else batch_units
                    ids = (prev_tail_ids + batch_row_ids) if prev_tail_ids else batch_row_ids
                    
                    for ch in self.chunker.chunk_units(
                        units=units,
                        base_meta=meta_sheet,
                        joiner="\n",
                        locator=f"xlsx|sheet={sheet_name}",
                        range_key="row",
                        range_vals=ids,
                        start_chunk_index=chunk_index,
                    ):
                        chunk_index = ch["metadata"]["chunk_index"] + 1
                        yield ch
            
            try:
                wb.close()
            except Exception:
                pass
            return
            
        except DependencyNotInstalledError:
            raise
        except Exception as e:
            LOGGER.warning("OpenPyXL streaming failed, trying Polars fallback: %s", e)
            
            # Fallback: try Polars
            pl = self._get_polars()
            if pl is None:
                raise DependencyNotInstalledError("openpyxl", "xlsx", "Excel processing")
            
            try:
                df = pl.read_excel(file_path)
                cols = df.columns
                units = []
                row_ids = []
                for rid, row in enumerate(df.iter_rows(), start=1):
                    row_dict = dict(zip(cols, row))
                    s = " | ".join([
                        f"{k}: {v}" for k, v in row_dict.items() 
                        if v is not None and str(v).strip() != ""
                    ])
                    if s and len(s) >= cfg.min_chars:
                        units.append(s)
                        row_ids.append(rid)
                
                chunk_index = 0
                for ch in self.chunker.chunk_units(
                    units=units,
                    base_meta=base_meta,
                    joiner="\n",
                    locator="xlsx",
                    range_key="row",
                    range_vals=row_ids,
                    start_chunk_index=chunk_index,
                ):
                    yield ch
                return
            except Exception as e2:
                LOGGER.error("XLSX Error (both methods failed): %s", e2)
                raise DependencyNotInstalledError("openpyxl", "xlsx", "Excel processing")


# Backward compatibility function
def iter_chunks_from_xlsx(
    file_path: str,
    cfg: ChunkConfig = None
) -> Generator[Dict[str, Any], None, None]:
    """
    Iterate over chunks from an XLSX file.
    
    Args:
        file_path: Path to XLSX file.
        cfg: Chunk configuration.
        
    Yields:
        Chunk dictionaries.
    """
    chunker = XLSXChunker(cfg)
    yield from chunker.chunk_file(file_path)
