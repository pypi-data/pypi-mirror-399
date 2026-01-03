from __future__ import annotations

from typing import Sequence, TypeVar, TYPE_CHECKING

if TYPE_CHECKING:
    from typing import Any, Callable, Literal, Union
    from openpyxl import Workbook, _ZipFileFileProtocol
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    from openpyxl.formatting import Rule
    from openpyxl.styles import Alignment, Border, Color, PatternFill, Font
    from openpyxl.worksheet.filters import FilterColumn, Filters, Top10, CustomFilter, DynamicFilter, ColorFilter, IconFilter
    import datetime as dt

Column = TypeVar("Column", int, str, tuple[str,...])
Row = TypeVar("Row", bound=int)
Range = TypeVar("Range", bound=str)
Ranges = TypeVar("Ranges", list, Column, Row, Range)

MinCol = TypeVar("MinCol", bound=int)
MinRow = TypeVar("MinRow", bound=int)
MaxCol = TypeVar("MaxCol", bound=int)
MaxRow = TypeVar("MaxRow", bound=int)

Width = TypeVar("Width", float, str)
Height = TypeVar("Height", float, str)
Multiple = TypeVar("Multiple", bound=str)

Node = TypeVar("Node", bound=tuple[int,int])
TopLeft = TypeVar("TopLeft", bound=Node)
TopRight = TypeVar("TopRight", bound=Node)
BottomLeft = TypeVar("BottomLeft", bound=Node)
BottomRight = TypeVar("BottomRight", bound=Node)

State = TypeVar("State", bound=dict)

SINGLE_WIDTH: float = 8.43
SINGLE_HEIGHT: float = 15.0

ALIGN_CENTER = {"horizontal": "center", "vertical": "center"}


class StyleConfig(dict):
    def __init__(
            self,
            alignment: dict | None = None,
            border: dict | None = None,
            fill: dict | None = None,
            font: dict | None = None,
            number_format: str | None = None,
            hyperlink: str | None = None,
        ):
        return super().__init__(
            alignment=alignment, border=border, fill=fill, font=font,
            number_format=number_format, hyperlink=hyperlink)


class RuleConfig(dict):
    def __init__(
            self,
            operator: Literal[
                "endsWith", "containsText", "beginsWith", "lessThan", "notBetween", "lessThanOrEqual",
                "notEqual", "notContains", "between", "equal", "greaterThanOrEqual", "greaterThan", "formula"],
            formula: Sequence | None = None,
            stop_if_true: bool | None = None,
            border: dict | None = None,
            fill: dict | None = None,
            font: dict | None = None,
        ):
        return super().__init__(
            operator=operator, formula=formula, stop_if_true=stop_if_true,
            border=border, fill=fill, font=font)


class ConditionalConfig(dict):
    def __init__(
            self,
            ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
            range_type: Literal["column","row","range","auto"],
            rule: RuleConfig,
        ):
        return super().__init__(ranges=ranges, range_type=range_type, rule=rule)


class MergeConfig(dict):
    def __init__(
            self,
            ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
            range_type: Literal["column","row","range","auto"],
            mode: Literal["all","blank","same_value"],
            styles: StyleConfig | None = {"alignment": ALIGN_CENTER},
        ):
        return super().__init__(ranges=ranges, range_type=range_type, mode=mode, styles=styles)


class FilterConfig(dict):
    def __init__(
            self,
            filter_type: Literal["value","top10","custom","dynamic","color","icon","blank","notBlank"],
            **config
        ):
        return super().__init__(filter_type=filter_type, **config)


class ColumnFilters(dict):
    def __init__(
            self,
            range: Range | Literal[":all:"],
            filters: Sequence[tuple[Column, Sequence[FilterConfig]]] = list(),
            button: Literal["always","hidden","auto"] = "always",
        ):
        return super().__init__(range=range, filters=filters, button=button)


def filter_warnings():
    import warnings
    warnings.filterwarnings("ignore", module="openpyxl.*")


###################################################################
###################### Convert Excel to JSON ######################
###################################################################

def to_unique_headers(headers: list[str]) -> list[str]:
    unique = list()
    for header in headers:
        header_str, suffix = str(header), 1
        while header_str in unique:
            header_str = f"{header}_{suffix}"
            suffix += 1
        unique.append(header_str)
    return unique


def csv2json(
        io: _ZipFileFileProtocol,
        header: int = 0,
        delimiter: str = ",",
        lineterminator: str = "\r\n",
        encoding: str | None = "utf-8",
    ) -> list[dict]:
    import os
    if isinstance(io, str) and os.path.exists(io):
        with open(io, 'r', encoding=encoding) as file:
            csv2json(file, header)

    import csv
    if isinstance(io, bytes):
        from io import BytesIO, TextIOWrapper
        io = TextIOWrapper(BytesIO(io), encoding=encoding)
    rows = list(csv.reader(io, delimiter=delimiter, lineterminator=lineterminator))
    header_row = to_unique_headers(rows[header])
    return [dict(zip(header_row, row)) for row in rows[(header+1):]]


def excel2json(
        io: _ZipFileFileProtocol,
        sheet_name: str | None = None,
        header: int = 1,
        warnings: bool = True
    ) -> list[dict]:
    from openpyxl import load_workbook
    from io import BytesIO
    if not warnings:
        filter_warnings()

    wb = load_workbook(BytesIO(io) if isinstance(io, bytes) else io)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = to_unique_headers([cell.value for cell in next(ws.iter_rows(min_row=header, max_row=header))])
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=(header+1), values_only=True)]


###################################################################
################### Convert CSV or JSON to Excel ##################
###################################################################

def csv2excel(
        obj: Sequence[Sequence[Any]] | dict[str,Sequence[Sequence[Any]]],
        sheet_name: str = "Sheet1",
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        column_styles: dict[Column, StyleConfig] = dict(),
        row_styles: dict[Row, StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit:",":fit_header:",":fit_values:"] | None = ":fit:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range, StyleConfig]] = list(),
        column_filters: ColumnFilters = dict(),
        filter_mode: Literal["openpyxl","xml"] | None = None,
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
        zoom_scale: int | None = None,
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_styles=column_styles, row_styles=row_styles, column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, merge_cells=merge_cells, range_styles=range_styles,
        column_filters=column_filters, filter_action={"openpyxl": "hide", "xml": "list"}.get(filter_mode),
        hyperlink=hyperlink, truncate=truncate, wrap_text=wrap_text, freeze_panes=freeze_panes, zoom_scale=zoom_scale)

    states = list()
    for index, (name, rows) in enumerate(obj.items(), start=1):
        ws, state = _rows2sheet(wb, rows, index, name, header_rows, header_styles, **kwargs)
        states.append(state)

    if filter_mode == "xml":
        filtered_rows = {index: state["filtered_rows"] for index, state in enumerate(states, start=1) if state["filtered_rows"]}
        return apply_filters(wb, filtered_rows)
    else:
        return wb


def json2excel(
        obj: Sequence[dict] | dict[str,Sequence[dict]],
        sheet_name: str = "Sheet1",
        header: Literal["first","all"] | None = "first",
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        column_styles: dict[Column, StyleConfig] = dict(),
        row_styles: dict[Row, StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit:",":fit_header:",":fit_values:"] | None = ":fit:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range, StyleConfig]] = list(),
        column_filters: ColumnFilters = dict(),
        filter_mode: Literal["openpyxl","xml"] | None = None,
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
        zoom_scale: int | None = None,
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_styles=column_styles, row_styles=row_styles, column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, merge_cells=merge_cells, range_styles=range_styles,
        column_filters=column_filters, filter_action={"openpyxl": "hide", "xml": "list"}.get(filter_mode),
        hyperlink=hyperlink, truncate=truncate, wrap_text=wrap_text, freeze_panes=freeze_panes, zoom_scale=zoom_scale)

    def _get_all_keys(rows: Sequence[dict]) -> list:
        keys = list()
        for row in rows:
            for key in row.keys():
                if key not in keys:
                    keys.append(key)
        return keys

    def _get_json_keys(rows: Sequence[dict], how: Literal["first","all"]) -> list:
        if not rows:
            return list()
        elif how == "first":
            return list(rows[0].keys())
        else:
            return _get_all_keys(rows)

    def _keys_to_rows(keys: list) -> list[list]:
        key_depths = {len(key) for key in keys if isinstance(key, tuple)}
        if key_depths:
            def get_value(key: Any, depth: int):
                if isinstance(key, tuple):
                    return key[depth] if depth < len(key) else None
                else:
                    return key if depth == 0 else None
            return [[get_value(key, depth) for key in keys] for depth in range(max(key_depths))]
        else:
            return [keys]

    states = list()
    for index, (name, rows) in enumerate(obj.items(), start=1):
        keys = _get_json_keys(rows, how=(header or "first"))
        values = [[row.get(key, None) for key in keys] for row in rows]
        headers = _keys_to_rows(keys) if header else list()
        header_rows = list(range(1,len(headers)+1)) if headers else list()
        ws, state = _rows2sheet(wb, (headers + values), index, name, header_rows, header_styles, **kwargs)
        states.append(state)

    if filter_mode == "xml":
        filtered_rows = {index: state["filtered_rows"] for index, state in enumerate(states, start=1) if state["filtered_rows"]}
        return apply_filters(wb, filtered_rows)
    else:
        return wb


def _rows2sheet(
        wb: Workbook,
        rows: Sequence[Sequence[Any]],
        sheet_index: int,
        sheet_name: str = "Sheet1",
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        **kwargs
    ) -> tuple[Worksheet, State]:
    if sheet_index == 1:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    if not rows:
        return

    for row in rows:
        ws.append(row)

    if not isinstance(header_styles, dict):
        header_styles = _yellow_headers() if header_styles == "yellow" else dict()

    state = style_sheet(ws, header_rows, header_styles, **kwargs)
    return ws, state


def save_excel_to_tempfile(wb: Workbook) -> str:
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
        tmp_path = tmp_file.name
        wb.save(tmp_path)
        return tmp_path


def _yellow_headers() -> StyleConfig:
    return {
        "align": ALIGN_CENTER,
        "fill": {"color": "#FFFF00", "fill_type": "solid"},
        "font": {"color": "#000000", "bold": True},
    }


###################################################################
######################### Style Worksheet #########################
###################################################################

def style_sheet(
        ws: Worksheet,
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig = dict(),
        column_styles: dict[Column, StyleConfig] = dict(),
        row_styles: dict[Row, StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit:",":fit_header:",":fit_values:"] | None = ":fit:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range, StyleConfig]] = list(),
        column_filters: ColumnFilters = dict(),
        filter_action: Literal["hide","list"] | None = None,
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: Range | None = "A2",
        zoom_scale: int | None = None,
    ) -> State:
    min_col, max_col = 'A', get_column_letter(ws.max_column)
    min_row, max_row = ((max(header_rows) + 1) if header_rows else 1), ws.max_row
    size = dict(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)

    headers = ([tuple(ws.cell(row=row_idx, column=col_idx).value for row_idx in header_rows)
        for col_idx in range(1, ws.max_column+1)] if header_rows else list())

    if truncate:
        row_height = SINGLE_HEIGHT if row_height is None else row_height
        wrap_text = True

    # STYLE CELLS BY COLUMN

    column_styles = _init_column_styles(column_styles, headers) if column_styles else dict()
    column_width = _init_column_width(column_width, headers) if column_width is not None else dict()
    fit_width = {col_idx: mode for col_idx, mode in column_width.items() if isinstance(mode, str)}

    for col_idx, column in enumerate(ws.columns, start=1):
        fit_mode = fit_width.get(col_idx)
        max_width = SINGLE_WIDTH

        for row_idx, cell in enumerate(column, start=1):
            text = str(x) if (x := cell.value) is not None else str()

            if hyperlink and text.startswith("https://"):
                cell.hyperlink = text
                cell.font = _font(color="#0000FF", underline="single")

            if wrap_text:
                cell.alignment = _alignment(wrap_text=True)

            if row_idx in header_rows:
                if header_styles:
                    style_cell(cell, **header_styles)
                if fit_mode in (":fit:",":fit_header:"):
                    max_width = max(max_width, get_cell_width(text))
            elif col_idx in column_styles:
                style_cell(cell, **column_styles[col_idx])
            elif row_idx in row_styles:
                style_cell(cell, **row_styles[row_idx])

            if (fit_mode in (":fit:",":fit_values:")) and (row_idx not in header_rows):
                max_width = max(max_width, get_cell_width(text))

        # CHANGE COLUMN WIDTH

        width = min(max_width + 2., 25.) if fit_mode else column_width.get(col_idx)
        if isinstance(width, float):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            column_width["width"] = width

    # CHANGE ROW HEIGHT

    row_height = _init_row_height(row_height) if row_height is not None else dict()

    if isinstance(row_height, dict):
        for row_idx, height in row_height.items():
            ws.row_dimensions[row_idx].height = height
    elif row_height is not None:
        for row_idx in range(1, max_row+1):
            ws.row_dimensions[row_idx].height = SINGLE_HEIGHT

    # CONDITIONAL FORMATTING

    def _to_excel_range(range_string: str) -> str:
        return to_valid_excel_range(ws, range_string)

    for config in conditional_formatting:
        ranges = get_ranges(config["ranges"], (config.get("range_type") or "auto"), **size, headers=headers)
        if ranges:
            range_string = ' '.join(map(_to_excel_range, ranges))
            ws.conditional_formatting.add(range_string, _conditional_rule(**config["rule"]))

    # MERGE CELLS

    for config in merge_cells:
        styles = config.get("styles")
        for range_string in get_ranges(config["ranges"], config.get("range_type", "auto"), **size, headers=headers):
            for merge_range in find_merge_ranges(ws, _to_excel_range(range_string), (config.get("mode") or "all")):
                ws.merge_cells(merge_range)
                if styles:
                    col_start, row_start, _, _ = range_boundaries(_to_excel_range(merge_range))
                    style_cell(ws.cell(row_start, col_start), **styles)

    # STYLE CELLS BY RANGE

    for range_string, styles in range_styles:
        col_start, row_start, col_end, row_end = range_boundaries(_to_excel_range(range_string))
        for row in ws.iter_rows(row_start, row_end, col_start, col_end):
            for cell in row:
                style_cell(cell, **styles)

    # FILTER BY COLUMN

    filtered_rows = set()

    if column_filters:
        from openpyxl.worksheet.filters import AutoFilter
        range_string = f"A1:{max_col}{max_row}" if (ref := column_filters["range"]) == ":all:" else _to_excel_range(ref)
        _, min_row, _, max_row = range_boundaries(range_string)
        row_range = [row_idx for row_idx in range(min_row, max_row+1) if row_idx not in header_rows]

        filters = list()
        for column, configs in (column_filters.get("filters") or list()):
            if (col_idx := get_column_index(column, headers)) is not None:
                filters.append(_filter_column(col_idx, configs, column_filters.get("button", "always")))
                if filter_action:
                    values = [(row_idx, ws.cell(row_idx, col_idx).value) for row_idx in row_range if row_idx not in filtered_rows]
                    filtered_rows = filtered_rows.union(filter_values(values, configs))
        ws.auto_filter = AutoFilter(ref=range_string, filterColumn=filters)

        if (filter_action == "hide") and filtered_rows:
            for row_idx in filtered_rows:
                ws.row_dimensions[row_idx].hidden = True

    # EXTRA SETTINGS

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    if zoom_scale:
        ws.sheet_view.zoomScale = zoom_scale

    return dict(size, column_styles=column_styles, column_width=column_width, row_height=row_height, filtered_rows=filtered_rows)


def style_cell(
        cell: Cell,
        alignment: dict | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        number_format: str | None = None,
        hyperlink: str | None = None,
        **kwargs
    ):
    if alignment:
        cell.alignment = _alignment(**alignment)
    if border:
        cell.border = _border(**border)
    if fill:
        cell.fill = _fill(**fill)
    if font:
        cell.font = _font(**font)
    if number_format is not None:
        cell.number_format = number_format
    if hyperlink is not None:
        cell.hyperlink = hyperlink


def get_cell_width(value: str) -> float:
    try:
        # 한글: 1.8배, 공백: 1.2배, 영문/숫자: 1배
        return sum(1.8 if ord(c) > 12799 else 1.2 if c.isspace() else 1. for c in value)
    except:
        return 0.


###################################################################
########################### Column utils ##########################
###################################################################

def get_column_index(column: Column, headers: list[tuple[str,...]] = list()) -> int | None:
    if isinstance(column, int):
        return column
    elif isinstance(column, str):
        for index, header in enumerate(headers, start=1):
            if (len(header) == 1) and (column == header[0]):
                return index
    elif isinstance(column, tuple):
        try:
            return headers.index(column) + 1
        except:
            pass
    return None


def get_column_letter(column: Column, headers: list[tuple[str,...]] = list()) -> str | None:
    from openpyxl.utils import get_column_letter as get_letter
    if isinstance(column, int):
        return get_letter(column)
    elif isinstance(col_idx := get_column_index(column, headers), int):
        return get_letter(col_idx)
    else:
        return None


def colstr(col_idx: int) -> str:
    from openpyxl.utils import get_column_letter as get_letter
    return get_letter(col_idx)


###################################################################
########################### Range utils ###########################
###################################################################

def get_ranges(
        ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
        range_type: Literal["column","row","range","auto"],
        min_col: str,
        max_col: str,
        min_row: int,
        max_row: int,
        headers: list[tuple[str,...]] = list(),
    ) -> list[str]:

    def _auto_detect(value: Column | Row | Range) -> Literal["column","row","range","auto"]:
        if isinstance(value, int):
            return "row"
        elif isinstance(value, str):
            return "range" if is_range_string(value) else "column"
        elif isinstance(value, tuple):
            return "column"
        else:
            return "auto"

    def _make_range_string(value: Column | Row | Range, range_type: Literal["column","row","range","auto"]) -> str:
        if range_type == "auto":
            range_type = _auto_detect(value)

        if (range_type == "column") and (column := get_column_letter(value, headers)):
            return f"{column}{min_row}:{column}{max_row}"
        elif (range_type == "row") and isinstance(value, int):
            return f"{min_col}{value}:{max_col}{value}"
        elif (range_type == "range") and isinstance(value, str) and is_range_string(value):
            return value
        else:
            raise ValueError(f"Invalid Excel range format: {value}")

    if isinstance(ranges, list):
        return [range_string for value in ranges if (range_string := _make_range_string(value, range_type))]
    else:
        return [_make_range_string(ranges, range_type)]


def is_range_string(value: str) -> bool:
    import re
    if ':' in value:
        min_col, min_row, max_col, max_row = map(bool, split_range_string(value))
        if min_col and max_col:
            return min_row or (not max_row)
        else:
            return (not min_col) and (not max_col) and (min_row and max_row)
    else:
        col, row = r"\$?[A-Z]+", r"\$?[1-9][0-9]*"
        return re.match(r"^{}{}$".format(col, row), value)


def to_valid_excel_range(ws: Worksheet, value: str) -> str:
    if ':' not in value:
        return value

    def absolute(x: str) -> Literal['$','']:
        return '$' if x.startswith('$') else ''

    min_col, min_row, max_col, max_row = split_range_string(value)
    if not min_col:
        min_col = ((absolute(min_row) or absolute(max_col)) + 'A')
    if not min_row:
        min_row = ((absolute(min_col) or absolute(max_row)) + '1')
    if not max_col:
        max_col = ((absolute(max_row) or absolute(min_col)) + get_column_letter(ws.max_column))
    if not max_row:
        max_row = ((absolute(max_col) or absolute(min_row)) + str(ws.max_row))
    return f"{min_col}{min_row}:{max_col}{max_row}"


def split_range_string(value: str) -> tuple[MinCol,MinRow,MaxCol,MaxRow]:
    import re
    col, row = r"\$?[A-Z]+", r"\$?[1-9][0-9]*"
    cell = r"^({})?({})?$".format(col, row)
    top_left, bottom_right = value.split(':')
    min_col, min_row = match.groups() if (match := re.search(cell, top_left)) else (None, None)
    max_col, max_row = match.groups() if (match := re.search(cell, bottom_right)) else (None, None)
    return (min_col or str()), (min_row or str()), (max_col or str()), (max_row or str())


###################################################################
########################### Style config ##########################
###################################################################

def _init_column_styles(
        column_styles: dict[Column, StyleConfig],
        headers: list[tuple[str,...]] = list(),
    ) -> dict[int,StyleConfig]:
    return {col_idx: styles for column, styles in column_styles.items()
        if ((col_idx := get_column_index(column, headers)) is not None)}


def _init_column_width(
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit:",":fit_header:",":fit_values:"],
        headers: list[tuple[str,...]] = list(),
    ) -> dict[int, Union[float,Literal[":fit:",":fit_header:",":fit_values:"]]]:

    def _set_width(value: Width) -> float | Literal[":fit:",":fit_header:",":fit_values:"]:
        if isinstance(value, str):
            if value in (":fit:",":fit_header:",":fit_values:"):
                return value
            elif value.endswith('x'):
                value = SINGLE_WIDTH * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(column_width, dict):
        return {col_idx: width for column, value in column_width.items()
                if ((col_idx := get_column_index(column, headers)) is not None)
                    and ((width := _set_width(value)) is not None)}
    else:
        value = _set_width(column_width)
        if value is not None:
            return {col_idx: value for column in headers
                    if (col_idx := get_column_index(column, headers)) is not None}
        else:
            return dict()


def _init_row_height(
        row_height: dict[Row,Height] | float | Multiple | Literal["single"],
    ) -> dict[int,float] | float | None:

    def _set_height(value: Width) -> float:
        if isinstance(value, str):
            if value == "single":
                return SINGLE_HEIGHT
            elif value.endswith('x'):
                value = SINGLE_HEIGHT * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(row_height, dict):
        return {row_idx: height for row_idx, value in row_height.items()
                if (height := _set_height(value)) is not None}
    else:
        return _set_height(row_height)


###################################################################
########################### Style object ##########################
###################################################################

def _alignment(**kwargs) -> Alignment:
    from openpyxl.styles import Alignment
    return Alignment(**kwargs)


def _border(**kwargs: dict) -> Border:
    from openpyxl.styles import Border, Side
    def side(color: str | None = None, **kwargs) -> Side:
        return Side(color=(_color(color) if color is not None else None), **kwargs)
    return Border(**{property: side(**config) for property, config in kwargs.items()})


def _fill(color: str | None = None, **kwargs) -> PatternFill:
    from openpyxl.styles import PatternFill
    for property, value in kwargs.items():
        if property in {"fgColor","bgColor","start_color","end_color"}:
            kwargs[property] = _color(value)
    if color is not None:
        color = _color(color)
        kwargs.update(start_color=color, end_color=color)
    return PatternFill(**kwargs)


def _font(color: str | None = None, **kwargs) -> Font:
    from openpyxl.styles import Font
    return Font(color=(_color(color) if color is not None else None), **kwargs)


def _color(rgb: Any, alpha: str = "FF") -> Color:
    from openpyxl.styles import Color
    if isinstance(rgb, str):
        return Color((alpha + rgb[1:]) if rgb.startswith('#') else rgb)
    elif isinstance(rgb, dict):
        return Color(**rgb)
    elif isinstance(rgb, Color):
        return rgb
    else:
        return None


###################################################################
########################## Filter object ##########################
###################################################################

def _filter_column(
        col_idx: int,
        configs: Sequence[FilterConfig],
        button: Literal["always","hidden","auto"] = "always",
    ) -> FilterColumn:
    from openpyxl.worksheet.filters import FilterColumn
    kwargs, custom_filters, blank = dict(), list(), None

    for config in configs:
        filter_type = config["filter_type"]
        if filter_type == "value":
            kwargs["filters"] = _value_filter(**config)
        elif filter_type == "top10":
            kwargs["top10"] = _top10_filter(**config)
        elif filter_type == "custom":
            custom_filters.append(_custom_filter(**config))
        elif filter_type == "dynamic":
            kwargs["dynamicFilter"] = _dynamic_filter(**config)
        elif filter_type == "color":
            kwargs["colorFilter"] = _color_filter(**config)
        elif filter_type == "icon":
            kwargs["iconFilter"] = _icon_filter(**config)
        elif filter_type == "blank":
            blank = True
        elif filter_type == "notBlank":
            blank = False
        else:
            pass

    if custom_filters:
        from openpyxl.worksheet.filters import CustomFilters
        kwargs["customFilters"] = CustomFilters(customFilter=custom_filters)

    filter_column = FilterColumn(colId=(col_idx-1), **_filter_button_options(button), **kwargs)
    return _add_blank_filter(filter_column, blank) if isinstance(blank, bool) else filter_column



def _value_filter(
        blank: bool | None = None,
        calendar_type: str | None = None,
        values: Sequence[str] | None = None,
        dates: Sequence[dict] | None = None,
        **kwargs
    ) -> Filters:
    from openpyxl.worksheet.filters import Filters, DateGroupItem
    return Filters(
        **(dict(blank=blank) if isinstance(blank, bool) else dict()),
        **(dict(calendarType=calendar_type) if isinstance(calendar_type, str) else dict()),
        **(dict(filter=values) if isinstance(values, Sequence) else dict()),
        **(dict(dateGroupItem=[DateGroupItem(**item) for item in dates]) if isinstance(dates, Sequence) else dict()),
    )


def _top10_filter(
        value: float,
        top: bool | None = None,
        percent: bool | None = None,
        filter_value: float | None = None,
        **kwargs
    ) -> Top10:
    from openpyxl.worksheet.filters import Top10
    return Top10(
        **(dict(top=top) if isinstance(top, bool) else dict()),
        **(dict(percent=percent) if isinstance(percent, bool) else dict()),
        val = value,
        **(dict(filterVal=filter_value) if filter_value is not None else dict()),
    )


def _custom_filter(
        operator: Literal["equal","lessThan","lessThanOrEqual","notEqual","greaterThanOrEqual","greaterThan"],
        value: str,
        **kwargs
    ) -> CustomFilter:
    from openpyxl.worksheet.filters import CustomFilter
    return CustomFilter(operator=operator, val=value)


def _dynamic_filter(
        type: Literal[
            "null", "aboveAverage", "belowAverage", "tomorrow",
            "today", "yesterday", "nextWeek", "thisWeek", "lastWeek", "nextMonth",
            "thisMonth", "lastMonth", "nextQuarter", "thisQuarter", "lastQuarter",
            "nextYear", "thisYear", "lastYear", "yearToDate", "Q1", "Q2", "Q3", "Q4",
            "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"],
        value: float | None = None,
        datetime: dt.datetime | None = None,
        max_value: float | None = None,
        max_datetime: dt.datetime | None = None,
        **kwargs
    ) -> DynamicFilter:
    from openpyxl.worksheet.filters import DynamicFilter
    return DynamicFilter(type=type,
        **(dict(value=value) if value is not None else dict()),
        **(dict(datetime=datetime) if datetime is not None else dict()),
        **(dict(max_value=max_value) if max_value is not None else dict()),
        **(dict(max_datetime=max_datetime) if max_datetime is not None else dict()),
    )


def _color_filter(
        dfx_id: int | None = None,
        color_type: Literal["cell","font"] = "cell",
        **kwargs
    ) -> ColorFilter:
    from openpyxl.worksheet.filters import ColorFilter
    return ColorFilter(
        **(dict(dxfId=dfx_id) if isinstance(dfx_id, int) else dict()),
        **(dict(cellColor=(color_type == "cell")) if color_type in {"cell","font"} else dict()),
    )


def _icon_filter(
        icon_set: Literal[
            "3Arrows", "3ArrowsGray", "3Flags",
            "3TrafficLights1", "3TrafficLights2", "3Signs", "3Symbols", "3Symbols2",
            "4Arrows", "4ArrowsGray", "4RedToBlack", "4Rating", "4TrafficLights",
            "5Arrows", "5ArrowsGray", "5Rating", "5Quarters"],
        icon_id: int | None = None,
        **kwargs
    ) -> IconFilter:
    from openpyxl.worksheet.filters import IconFilter
    return IconFilter(
        iconSet = icon_set,
        **(dict(iconId=icon_id) if isinstance(icon_id, int) else dict()),
    )


def _add_blank_filter(filter_column: FilterColumn, blank: bool) -> FilterColumn:
    if blank:
        from openpyxl.worksheet.filters import Filters, BlankFilter
        filter_column.filters = Filters(blank=True, filter=[BlankFilter()])
    else:
        filter_column.filters.blank = False
    return filter_column


def _filter_button_options(how: Literal["always","hidden","auto"] = "always") -> dict:
    if how == "hidden":
        return dict(hiddenButton=True)
    elif how == "auto":
        return dict(showButton=False)
    else:
        return dict() # default = dict(hiddenButton=False, showButton=True)


###################################################################
########################## Filter values ##########################
###################################################################

def filter_values(
        values: Sequence[tuple[Row,Any]],
        configs: Sequence[FilterConfig],
    ) -> set[Row]:
    conditions, global_config = _init_conditions(configs)
    filtered_rows = set()

    for row_idx, value in values:
        if row_idx not in filtered_rows:
            for condition in conditions:
                if not condition(value):
                    filtered_rows.add(row_idx)
                    break

    if global_config:
        try:
            return filtered_rows.union(_apply_global_filter(values, **global_config))
        except:
            return filtered_rows
    else:
        return filtered_rows


def _init_conditions(configs: Sequence[FilterConfig]) -> tuple[Callable[[Any],bool], dict]:
    conditions, global_config = list(), dict()
    for config in configs:
        filter_type = config["filter_type"]
        if filter_type == "value":
            conditions.append(lambda cell_value: _apply_value_filter(cell_value, **config))
        elif filter_type == "top10":
            global_config.update(config)
        elif filter_type == "custom":
            conditions.append(lambda cell_value: _apply_custom_filter(cell_value, **config))
        elif filter_type == "dynamic":
            if config.get("type") in {"aboveAverage","belowAverage"}:
                global_config["average"] = config["type"][:-7]
            else:
                conditions.append(lambda cell_value: _apply_dynamic_filter(cell_value, **config))
        elif filter_type == "color":
            pass # conditions.append(lambda cell: _apply_color_filter(cell, **config))
        elif filter_type == "icon":
            pass # conditions.append(lambda cell: _apply_icon_filter(cell, **config))
        elif filter_type == "blank":
            conditions.append(lambda cell_value: cell_value is None)
        elif filter_type == "notBlank":
            conditions.append(lambda cell_value: cell_value is not None)
        else:
            pass
    return conditions, global_config


def _apply_value_filter(
        cell_value: Any,
        blank: bool | None = None,
        values: Sequence[str] | None = None,
        dates: Sequence[dict] | None = None,
        **kwargs
    ) -> bool:
    if isinstance(blank, bool):
        if ((cell_value is None) if blank else (cell_value is not None)):
            return True

    if str(cell_value) in (values or list()):
        return True

    import datetime as dt
    for date in (dates or list()):
        if isinstance(date, dict) and isinstance(cell_value, dt.date):
            group = date["dateTimeGrouping"]
            if getattr(cell_value, group) == date.get(group):
                return True

    return False


def _apply_custom_filter(
        cell_value: Any,
        operator: Literal["equal","lessThan","lessThanOrEqual","notEqual","greaterThanOrEqual","greaterThan"],
        value: str,
        **kwargs
    ) -> bool:
    try:
        if operator == "equal":
            return cell_value == value
        elif operator == "lessThan":
            return cell_value < value
        elif operator == "lessThanOrEqual":
            return cell_value <= value
        elif operator == "notEqual":
            return cell_value != value
        elif operator == "greaterThanOrEqual":
            return cell_value >= value
        elif operator == "greaterThan":
            return cell_value > value
        else:
            return False
    except:
        return False


def _apply_dynamic_filter(
        cell_value: Any,
        type: Literal[
            "null", "tomorrow", "today", "yesterday", "nextWeek", "thisWeek", "lastWeek", "nextMonth",
            "thisMonth", "lastMonth", "nextQuarter", "thisQuarter", "lastQuarter",
            "nextYear", "thisYear", "lastYear", "yearToDate", "Q1", "Q2", "Q3", "Q4",
            "M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"],
        tz: str = "local",
        **kwargs
    ) -> bool:
    import pendulum
    import datetime as dt

    if type == "null":
        return cell_value is None

    # Convert cell value to pendulum date
    if isinstance(cell_value, dt.datetime):
        cell_date = pendulum.instance(cell_value, tz=tz).date()
    elif isinstance(cell_value, dt.date):
        cell_date = cell_value
    else:
        return False
    today = pendulum.today(tz).date()

    # Day-based filters
    if type in {"tomorrow", "today", "yesterday"}:
        delta = {"tomorrow": 1, "today": 0, "yesterday": -1}[type]
        return cell_date == today.add(days=delta)

    # Week-based filters
    elif type in {"nextWeek", "thisWeek", "lastWeek"}:
        start_of_week = today.start_of("week")
        if type == "thisWeek":
            return start_of_week <= cell_date < start_of_week.add(weeks=1)
        elif type == "nextWeek":
            return start_of_week.add(weeks=1) <= cell_date < start_of_week.add(weeks=2)
        elif type == "lastWeek":
            return start_of_week.subtract(weeks=1) <= cell_date < start_of_week

    # Month-based filters
    elif type in {"nextMonth", "thisMonth", "lastMonth"}:
        if type == "thisMonth":
            return (cell_date.year == today.year) and (cell_date.month == today.month)
        elif type == "nextMonth":
            next_month = today.add(months=1)
            return (cell_date.year == next_month.year) and (cell_date.month == next_month.month)
        elif type == "lastMonth":
            last_month = today.subtract(months=1)
            return (cell_date.year == last_month.year) and (cell_date.month == last_month.month)

    # Quarter-based filters
    elif type in {"nextQuarter", "thisQuarter", "lastQuarter"}:
        cell_quarter = (cell_date.month - 1) // 3 + 1
        today_quarter = (today.month - 1) // 3 + 1

        if type == "thisQuarter":
            return (cell_date.year == today.year) and (cell_quarter == today_quarter)
        elif type == "nextQuarter":
            next_quarter_date = today.add(months=3)
            next_quarter = (next_quarter_date.month - 1) // 3 + 1
            return (cell_date.year == next_quarter_date.year) and (cell_quarter == next_quarter)
        elif type == "lastQuarter":
            last_quarter_date = today.subtract(months=3)
            last_quarter = (last_quarter_date.month - 1) // 3 + 1
            return (cell_date.year == last_quarter_date.year) and (cell_quarter == last_quarter)
    
    # Year-based filters
    elif type in {"nextYear", "thisYear", "lastYear"}:
        if type == "thisYear":
            return cell_date.year == today.year
        elif type == "nextYear":
            return cell_date.year == (today.year + 1)
        elif type == "lastYear":
            return cell_date.year == (today.year - 1)

    # Year to date
    elif type == "yearToDate":
        return (cell_date.year == today.year) and (cell_date <= today)

    # Specific quarter filters
    elif type in {"Q1", "Q2", "Q3", "Q4"}:
        quarter = int(type[1])
        return ((cell_date.month - 1) // 3 + 1) == quarter

    # Specific month filters
    elif type in {"M1", "M2", "M3", "M4", "M5", "M6", "M7", "M8", "M9", "M10", "M11", "M12"}:
        month = int(type[1:])
        return cell_date.month == month

    return False


def _apply_global_filter(
        cell_values: Sequence[tuple[Row,Any]],
        value: float | None = None,
        top: bool | None = None,
        percent: bool | None = None,
        average: Literal["above","below"] | None = None,
        **kwargs
    ) -> set[Row]:
    filtered_rows = {row_idx for row_idx, value in cell_values if value is None}
    values = sorted([(row_idx, value) for row_idx, value in cell_values if value is not None], key=(lambda x: x[1]))

    if average:
        avg_value = sum(map(lambda x: x[1], values)) / len(values)
        rows = list()
        for row_idx, value in values:
            if ((value >= avg_value) if average == "above" else (value <= avg_value)):
                rows.append(row_idx)
            else:
                filtered_rows.add(row_idx)
    else:
        rows = list(map(lambda x: x[0], values))

    if value is not None:
        if percent:
            n = max(round_half_up(len(rows) * (value / 100)), 1)
            return filtered_rows.union(rows[:n] if top else rows[(n*-1):])
        else:
            return filtered_rows.union(rows[:value] if top else rows[(value*-1):])
    else:
        return filtered_rows


def round_half_up(value: float) -> int:
    from decimal import Decimal, ROUND_HALF_UP
    return int(Decimal(value).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


###################################################################
########################## Apply filters ##########################
###################################################################

def apply_filters(wb: Workbook, filtered_rows: dict[Row,set[Row]]) -> Workbook:
    import xml.etree.ElementTree as ET
    import os

    for index, rows in filtered_rows.items():
        ws = wb.worksheets[index-1]
        if not (ws.auto_filter and ws.auto_filter.filterColumn):
            continue

        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        ET.register_namespace('', ns["main"])

        input_path = save_excel_to_tempfile(wb)
        try:
            wb = _apply_filter(input_path, index, rows, ns)
        finally:
            os.remove(input_path)

    return wb


def _apply_filter(input_path: str, sheet_index: int, filtered_rows: set[Row], ns: dict) -> Workbook:
    from openpyxl import load_workbook
    from tempfile import NamedTemporaryFile
    import xml.etree.ElementTree as ET
    import zipfile
    import os

    with zipfile.ZipFile(input_path, 'r') as zip_in:
        sheet_xml = f"xl/worksheets/sheet{sheet_index}.xml"
        with zip_in.open(sheet_xml) as zip_file:
            tree = ET.parse(zip_file)
            root = tree.getroot()

        sheet_data = root.find(".//main:sheetData", ns)
        if sheet_data is not None:
            for row in sheet_data.findall(".//main:row", ns):
                row_idx = int(row.get('r', 0))
                if row_idx in filtered_rows:
                    row.set("hidden", '1')
                elif row.get("hidden"):
                    del row.attrib["hidden"]

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as file:
            output_path = file.name

        try:
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for zip_info in zip_in.infolist():
                    if zip_info.filename == sheet_xml:
                        xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                        zip_out.writestr(zip_info, xml_bytes)
                    else:
                        zip_out.writestr(zip_info, zip_in.read(zip_info.filename))
            return load_workbook(output_path)
        finally:
            os.remove(output_path)


###################################################################
###################### Conditional formatting #####################
###################################################################

def _conditional_rule(
        operator: Literal[
            "endsWith", "containsText", "beginsWith", "lessThan", "notBetween", "lessThanOrEqual",
            "notEqual", "notContains", "between", "equal", "greaterThanOrEqual", "greaterThan", "formula"],
        formula: Sequence | None = None,
        stop_if_true: bool | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        **kwargs
    ) -> Rule:
    if operator == "formula":
        return _formula_rule(formula, stop_if_true, border, fill, font)

    from openpyxl.formatting.rule import CellIsRule
    styles = dict()
    if border:
        styles["border"] = _border(**border)
    if fill:
        styles["fill"] = _fill(**fill)
    if font:
        styles["font"] = _font(**font)
    return CellIsRule(operator=operator, formula=formula, stopIfTrue=stop_if_true, **styles)


def _formula_rule(
        formula: Sequence | None = None,
        stop_if_true: bool | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        **kwargs
    ) -> Rule:
    from openpyxl.formatting.rule import FormulaRule
    styles = dict()
    if border:
        styles["border"] = _border(**border)
    if fill:
        styles["fill"] = _fill(**fill)
    if font:
        styles["font"] = _font(**font)
    return FormulaRule(formula=formula, stopIfTrue=stop_if_true, **styles)


###################################################################
########################### Merge cells ###########################
###################################################################

def find_merge_ranges(
        ws: Worksheet,
        range_string: Range,
        mode: Literal["all","blank","same_value"] = "all",
        priority: Literal["by_row","by_col"] = "by_row",
    ) -> list[Range]:
    if mode == "all":
        return [range_string]
    merge_ranges = list()

    from collections import deque
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    num_rows, num_cols = (max_row - min_row + 1), (max_col - min_col + 1)
    adapter = {"by_row": "width", "by_col": "height"}

    rows = [list(row) for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True)]
    visited = [[False for _ in range(num_cols)] for _ in range(num_rows)]

    def _bfs(row_seq: int, col_seq: int) -> list[tuple[int,int]]:
        queue, cells = deque(), [(row_seq, col_seq)]
        queue.append((row_seq, col_seq))
        visited[row_seq][col_seq] = True
        value = rows[row_seq][col_seq]

        while queue:
            r, c = queue.popleft()
            for nr, nc in [(r, c+1), (r+1, c)]: # [Right, Down]
                if (0 <= nr < num_rows) and (0 <= nc < num_cols) and (not visited[nr][nc]):
                    if (rows[nr][nc] == value) if mode == "same_value" else (rows[nr][nc] is None):
                        visited[nr][nc] = True
                        queue.append((nr, nc))
                        cells.append((nr, nc))
        return cells

    for row_seq in range(num_rows):
        for col_seq in range(num_cols):
            if (not visited[row_seq][col_seq]) and (rows[row_seq][col_seq] is not None):
                cells = _bfs(row_seq, col_seq)
                if len(cells) < 2:
                    continue

                best_corners = get_largest_rectangle(cells, top_left=cells[0], priority=adapter[priority])
                (row_start, col_start) = best_corners[0][0] # Top-Left
                (row_end, col_end) = best_corners[1][1] # Bottom-Right

                for r, c in cells:
                    if not ((row_start <= r <= row_end) and (col_start <= c <= col_end)):
                        visited[r][c] = False

                if (row_start != row_end) or (col_start != col_end):
                    merge_ranges.append(f"{colstr(col_start)}{row_start+min_row}:{colstr(col_end)}{row_end+min_col}")
                else:
                    for r, c in cells[1:]:
                        visited[r][c] = False
    return merge_ranges


def get_largest_rectangle(
        nodes: list[tuple[int,int]],
        top_left: tuple[int,int],
        priority: Literal["width","height"] = "width",
    ) -> tuple[tuple[TopLeft,TopRight],tuple[BottomLeft,BottomRight]]:
    node_set = set(nodes)
    y0, x0 = top_left

    best_score = (0, 0)
    best_corners = (
        ((top_left, top_left), (top_left, top_left)),
        ((top_left, top_left), (top_left, top_left)),
    )

    max_width = 0
    while (y0, x0 + max_width) in node_set:
        max_width += 1

    min_height = float("inf")
    for width in range(1, max_width + 1):
        height = 0
        while (y0 + height, x0 + width - 1) in node_set:
            height += 1

        min_height = min(min_height, height)
        area = width * min_height
        score = (area, min_height if priority == "height" else width)

        if score > best_score:
            best_score = score
            y_max, x_max = (y0 + int(min_height) - 1), (x0 + width - 1)
            best_corners = (
                ((y0, x0), (y0, x_max)),
                ((y_max, x0), (y_max, x_max))
            )

    return best_corners


def range_boundaries(range_string: str) -> tuple[MinCol,MinRow,MaxCol,MaxRow]:
    from openpyxl.utils import range_boundaries as boundaries
    min_col, min_row, max_col, max_row = boundaries(range_string)
    return min_col, min_row, max_col, max_row
