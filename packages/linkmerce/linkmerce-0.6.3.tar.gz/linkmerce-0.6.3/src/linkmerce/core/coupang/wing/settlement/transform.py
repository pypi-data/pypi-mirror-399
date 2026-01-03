from __future__ import annotations

from linkmerce.common.transform import JsonTransformer, DuckDBTransformer

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from typing import Literal
    from linkmerce.common.transform import JsonObject


class RocketSettlementList(JsonTransformer):
    dtype = dict
    path = ["settlementStatusReports"]

    def transform(self, obj: JsonObject, **kwargs) -> JsonObject:
        def filter_details(details: dict) -> dict:
            return {key: details.get(key) for key in self.detail_keys}

        return [dict(report, settlementStatusReportDetail = filter_details(report.get("settlementStatusReportDetail") or dict()))
                for report in self.parse(obj) if isinstance(report, dict)]

    @property
    def detail_keys(self) -> list[str]:
        return [
            "totalSalesAmount", "totalRefundedAmount", "totalTakeRateAmountWithVat", "totalSellerDiscount",
            "totalSellerFundedInstantDiscount", "totalSellerFundedDownloadDiscount", "totalPayableAmount",
            "totalMilkRunDeductionAmount", "totalAdSalesDeductionAmount", "totalAdditionalDeductionAmount",
            "totalNegativeDeductionAmount", "totalFinalCfsFeeDeductionAmount", "totalWarehousingFeeDeductionAmount",
            "totalFulfillmentFeeDeductionAmount", "totalStorageFeeDeductionAmount",
            "totalCreturnReverseShippingFeeDeductionAmount", "totalCreturnGradingFeeDeductionAmount",
            "totalVreturnHandlingFeeDeductionAmount", "totalBarcodeLabelingFeeDeductionAmount",
            "totalLastSettlementUnpaidCfsDeductionAmount", "totalPastCfsDeductionAmount",
            "totalCarryOverSettlementDeductionAmount", "totalCfsInventoryCompensationAmount",
        ]


class RocketSettlement(DuckDBTransformer):
    queries = ["create", "select", "insert"]

    def transform(self, obj: JsonObject, vendor_id: str | None = None, **kwargs):
        reports = RocketSettlementList().transform(obj)
        if reports:
            return self.insert_into_table(reports, params=dict(vendor_id=vendor_id))


class RocketSettlementDownload(DuckDBTransformer):
    queries = ["create_sales", "select_sales", "insert_sales", "create_shipping", "select_shipping", "insert_shipping"]

    def set_tables(self, tables: dict | None = None):
        base = dict(sales="coupang_rocket_sales", shipping="coupang_rocket_shipping")
        super().set_tables(dict(base, **(tables or dict())))

    def create_table(self, **kwargs):
        super().create_table(key="create_sales", table=":sales:")
        super().create_table(key="create_shipping", table=":shipping:")

    def transform(self, obj: bytes, report_type: Literal["CATEGORY_TR","WAREHOUSING_SHIPPING"], vendor_id: str | None = None, **kwargs):
        if report_type == "CATEGORY_TR":
            return self.insert_into_sales_table(obj, vendor_id)
        else:
            return self.insert_into_shipping_table(obj, vendor_id)

    def insert_into_sales_table(self, obj: bytes, vendor_id: str | None = None):
        from linkmerce.utils.excel import excel2json
        reports = excel2json(obj, header=2, warnings=False)
        if reports:
            return self.insert_into_table(
                reports, key="insert_sales", table=":sales:", values=":select_sales:", params=dict(vendor_id=vendor_id))

    def insert_into_shipping_table(self, obj: bytes, vendor_id: str | None = None):
        from linkmerce.utils.excel import filter_warnings
        from io import BytesIO
        import openpyxl
        filter_warnings()

        wb = openpyxl.load_workbook(BytesIO(obj))
        results = list()

        for sheet_name in ["입출고비", "배송비"]:
            ws = wb[sheet_name]
            headers1 = [cell.value for cell in next(ws.iter_rows(min_row=7, max_row=7))]
            headers2 = [cell.value for cell in next(ws.iter_rows(min_row=8, max_row=8))]
            headers = [(header2 if header2 else header1) for header1, header2 in zip(headers1, headers2)]

            reports = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=9, values_only=True)]
            if reports:
                results.append(self.insert_into_table(
                    reports, key="insert_shipping", table=":shipping:", values=":select_shipping:", params=dict(vendor_id=vendor_id)))
        return results
