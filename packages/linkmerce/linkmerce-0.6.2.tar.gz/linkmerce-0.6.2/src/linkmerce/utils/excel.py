from __future__ import annotations

from typing import Sequence, TypeVar, TYPE_CHECKING

if TYPE_CHECKING:
    from typing import Any, Literal, Union
    from openpyxl import Workbook, _ZipFileFileProtocol
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    from openpyxl.formatting import Rule
    from openpyxl.styles import Alignment, Border, Color, PatternFill, Font

Column = TypeVar("Column", int, str, tuple[str,...])
Row = TypeVar("Row", bound=int)
Range = TypeVar("Range", bound=str)
Ranges = TypeVar("Ranges", list, Column, Row, Range)

Width = TypeVar("Width", float, str)
Height = TypeVar("Height", float, str)
Multiple = TypeVar("Multiple", bound=str)

Node = TypeVar("Node", bound=tuple[int,int])
TopLeft = TypeVar("TopLeft", bound=Node)
TopRight = TypeVar("TopRight", bound=Node)
BottomLeft = TypeVar("BottomLeft", bound=Node)
BottomRight = TypeVar("BottomRight", bound=Node)

SINGLE_WIDTH: float = 8.43
SINGLE_HEIGHT: float = 15.0

ALIGN_CENTER = {"horizontal": "center", "vertical": "center"}


class StyleConfig(dict):
    def __init__(
            self,
            alignment: dict | None = None,
            border: dict | None = None,
            fill: dict | None = None,
            font: dict | None = None,
            number_format: str | None = None,
            hyperlink: str | None = None,
        ):
        return super().__init__(
            alignment=alignment, border=border, fill=fill, font=font,
            number_format=number_format, hyperlink=hyperlink)


class RuleConfig(dict):
    def __init__(
            self,
            operator: Literal[
                "endsWith", "containsText", "beginsWith", "lessThan", "notBetween", "lessThanOrEqual",
                "notEqual", "notContains", "between", "equal", "greaterThanOrEqual", "greaterThan", "formula"],
            formula: Sequence | None = None,
            stop_if_true: bool | None = None,
            border: dict | None = None,
            fill: dict | None = None,
            font: dict | None = None,
        ):
        return super().__init__(
            operator=operator, formula=formula, stop_if_true=stop_if_true,
            border=border, fill=fill, font=font)


class ConditionalConfig(dict):
    def __init__(
            self,
            ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
            range_type: Literal["column","row","range","auto"],
            rule: RuleConfig,
        ):
        return super().__init__(ranges=ranges, range_type=range_type, rule=rule)


class MergeConfig(dict):
    def __init__(
            self,
            ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
            range_type: Literal["column","row","range","auto"],
            mode: Literal["all","blank","same_value"],
            styles: StyleConfig | None = {"alignment": ALIGN_CENTER},
        ):
        return super().__init__(ranges=ranges, range_type=range_type, mode=mode, styles=styles)


def filter_warnings():
    import warnings
    warnings.filterwarnings("ignore", module="openpyxl.*")


###################################################################
###################### Convert Excel to JSON ######################
###################################################################

def to_unique_headers(headers: list[str]) -> list[str]:
    unique = list()
    for header in headers:
        header_str, suffix = str(header), 1
        while header_str in unique:
            header_str = f"{header}_{suffix}"
            suffix += 1
        unique.append(header_str)
    return unique


def csv2json(
        io: _ZipFileFileProtocol,
        header: int = 0,
        delimiter: str = ",",
        lineterminator: str = "\r\n",
        encoding: str | None = "utf-8",
    ) -> list[dict]:
    import os
    if isinstance(io, str) and os.path.exists(io):
        with open(io, 'r', encoding=encoding) as file:
            csv2json(file, header)

    import csv
    if isinstance(io, bytes):
        from io import BytesIO, TextIOWrapper
        io = TextIOWrapper(BytesIO(io), encoding=encoding)
    rows = list(csv.reader(io, delimiter=delimiter, lineterminator=lineterminator))
    header_row = to_unique_headers(rows[header])
    return [dict(zip(header_row, row)) for row in rows[(header+1):]]


def excel2json(
        io: _ZipFileFileProtocol,
        sheet_name: str | None = None,
        header: int = 1,
        warnings: bool = True
    ) -> list[dict]:
    from openpyxl import load_workbook
    from io import BytesIO
    if not warnings:
        filter_warnings()

    wb = load_workbook(BytesIO(io) if isinstance(io, bytes) else io)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    headers = to_unique_headers([cell.value for cell in next(ws.iter_rows(min_row=header, max_row=header))])
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=(header+1), values_only=True)]


###################################################################
################### Convert CSV or JSON to Excel ##################
###################################################################

def csv2excel(
        obj: Sequence[Sequence[Any]] | dict[str,Sequence[Sequence[Any]]],
        sheet_name: str = "Sheet1",
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        column_styles: dict[Column,StyleConfig] = dict(),
        row_styles: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit_content:",":fit_header:",":fit_values:"] | None = ":fit_content:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range,StyleConfig]] = dict(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
        zoom_scale: int | None = None,
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_styles=column_styles, row_styles=row_styles, column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, merge_cells=merge_cells, range_styles=range_styles,
        hyperlink=hyperlink, truncate=truncate, wrap_text=wrap_text,
        freeze_panes=freeze_panes, zoom_scale=zoom_scale)

    for index, (name, rows) in enumerate(obj.items()):
        _rows2sheet(wb, rows, index, name, header_rows, header_styles, **kwargs)
    return wb


def json2excel(
        obj: Sequence[dict] | dict[str,Sequence[dict]],
        sheet_name: str = "Sheet1",
        header: Literal["first","all"] | None = "first",
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        column_styles: dict[Column,StyleConfig] = dict(),
        row_styles: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit_content:",":fit_header:",":fit_values:"] | None = ":fit_content:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range,StyleConfig]] = dict(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: str | None = "A2",
        zoom_scale: int | None = None,
    ) -> Workbook:
    from openpyxl import Workbook
    wb = Workbook()
    obj = {sheet_name: obj} if isinstance(obj, Sequence) else obj
    kwargs = dict(
        column_styles=column_styles, row_styles=row_styles, column_width=column_width, row_height=row_height,
        conditional_formatting=conditional_formatting, merge_cells=merge_cells, range_styles=range_styles,
        hyperlink=hyperlink, truncate=truncate, wrap_text=wrap_text,
        freeze_panes=freeze_panes, zoom_scale=zoom_scale)

    def _get_all_keys(rows: Sequence[dict]) -> list[str]:
        keys = list()
        for row in rows:
            for key in row.keys():
                if key not in keys:
                    keys.append(key)
        return keys

    def _get_json_keys(rows: Sequence[dict], how: Literal["first","all"]) -> list[str]:
        if not rows:
            return list()
        elif how == "first":
            return list(rows[0].keys())
        else:
            return _get_all_keys(rows)

    for index, (name, rows) in enumerate(obj.items()):
        keys = _get_json_keys(rows, how=(header or "first"))
        values = [[row.get(key, None) for key in keys] for row in rows]
        csv_rows = ([keys] + values) if header else values
        header_rows = [1] if header else []
        _rows2sheet(wb, csv_rows, index, name, header_rows, header_styles, **kwargs)
    return wb


def _rows2sheet(
        wb: Workbook,
        rows: Sequence[Sequence[Any]],
        sheet_index: int,
        sheet_name: str = "Sheet1",
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig | Literal["yellow"] = "yellow",
        **kwargs
    ) -> Worksheet:
    if sheet_index == 0:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    if not rows:
        return

    for row in rows:
        ws.append(row)

    if not isinstance(header_styles, dict):
        header_styles = _yellow_headers() if header_styles == "yellow" else dict()

    style_sheet(ws, header_rows, header_styles, **kwargs)
    return ws


def save_excel_to_tempfile(wb: Workbook) -> str:
    import tempfile
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        tmp_path = tmp_file.name
        wb.save(tmp_path)
        return tmp_path


def _yellow_headers() -> StyleConfig:
    return {
        "align": ALIGN_CENTER,
        "fill": {"color": "#FFFF00", "fill_type": "solid"},
        "font": {"color": "#000000", "bold": True},
    }


###################################################################
######################### Style Worksheet #########################
###################################################################

def style_sheet(
        ws: Worksheet,
        header_rows: Sequence[Row] = [1],
        header_styles: StyleConfig = dict(),
        column_styles: dict[Column,StyleConfig] = dict(),
        row_styles: dict[Row,StyleConfig] = dict(),
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit_content:",":fit_header:",":fit_values:"] | None = ":fit_content:",
        row_height: float | Multiple | dict[Row,Height] | None = None,
        conditional_formatting: Sequence[ConditionalConfig] = list(),
        merge_cells: Sequence[MergeConfig] = list(),
        range_styles: Sequence[tuple[Range,StyleConfig]] = dict(),
        hyperlink: bool = True,
        truncate: bool = False,
        wrap_text: bool = False,
        freeze_panes: Range | None = "A2",
        zoom_scale: int | None = None,
    ):
    min_col, max_col = 'A', get_column_letter(ws.max_column)
    min_row, max_row = ((max(header_rows) + 1) if header_rows else 1), ws.max_row
    size = dict(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)

    headers = ([tuple(ws.cell(row=row_idx, column=col_idx).value for row_idx in header_rows)
        for col_idx in range(1, ws.max_column+1)] if header_rows else list())

    if truncate:
        row_height = SINGLE_HEIGHT if row_height is None else row_height
        wrap_text = True

    # STYLE CELLS BY COLUMN

    column_styles = _init_column_styles(column_styles, headers) if column_styles else dict()
    column_width = _init_column_width(column_width, headers) if column_width is not None else dict()
    fit_width = {col_idx: mode for col_idx, mode in column_width.items() if isinstance(mode, str)}

    for col_idx, column in enumerate(ws.columns, start=1):
        fit_mode = fit_width.get(col_idx)
        max_width = SINGLE_WIDTH

        for row_idx, cell in enumerate(column, start=1):
            text = str(x) if (x := cell.value) is not None else str()

            if hyperlink and text.startswith("https://"):
                cell.hyperlink = text
                cell.font = _font(color="#0000FF", underline="single")

            if wrap_text:
                cell.alignment = _alignment(wrap_text=True)

            if row_idx in header_rows:
                if header_styles:
                    style_cell(cell, **header_styles)
                if fit_mode in (":fit_content:",":fit_header:"):
                    max_width = max(max_width, get_cell_width(text))
            elif col_idx in column_styles:
                style_cell(cell, **column_styles[col_idx])
            elif row_idx in row_styles:
                style_cell(cell, **row_styles[row_idx])

            if (fit_mode in (":fit_content:",":fit_values:")) and (row_idx not in header_rows):
                max_width = max(max_width, get_cell_width(text))

        # CHANGE COLUMN WIDTH

        width = min(max_width + 2., 25.) if fit_mode else column_width.get(col_idx)
        if isinstance(width, float):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # CHANGE ROW HEIGHT

    row_height = _init_row_height(row_height) if row_height is not None else dict()

    if isinstance(row_height, dict):
        for row_idx, height in row_height.items():
            ws.row_dimensions[row_idx].height = height
    elif row_height is not None:
        for row_idx in range(1, max_row+1):
            ws.row_dimensions[row_idx].height = SINGLE_HEIGHT

    # CONDITIONAL FORMATTING

    def _to_excel_range(range_string: str) -> str:
        return to_valid_excel_range(ws, range_string)

    for config in conditional_formatting:
        ranges = get_ranges(config["ranges"], (config.get("range_type") or "auto"), **size, headers=headers)
        if ranges:
            range_string = ' '.join(map(_to_excel_range, ranges))
            ws.conditional_formatting.add(range_string, _conditional_rule(**config["rule"]))

    # MERGE CELLS

    for config in merge_cells:
        styles = config.get("styles")
        for range_string in get_ranges(config["ranges"], config.get("range_type", "auto"), **size, headers=headers):
            for merge_range in find_merge_ranges(ws, _to_excel_range(range_string), (config.get("mode") or "all")):
                ws.merge_cells(merge_range)
                if styles:
                    col_start, row_start, _, _ = range_boundaries(_to_excel_range(merge_range))
                    style_cell(ws.cell(row_start, col_start), **styles)

    # STYLE CELLS BY RANGE

    for range_string, styles in range_styles:
        col_start, row_start, col_end, row_end = range_boundaries(_to_excel_range(range_string))
        for row in ws.iter_rows(row_start, row_end, col_start, col_end):
            for cell in row:
                style_cell(cell, **styles)

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    if zoom_scale:
        ws.sheet_view.zoomScale = zoom_scale


def style_cell(
        cell: Cell,
        alignment: dict | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        number_format: str | None = None,
        hyperlink: str | None = None,
        **kwargs
    ):
    if alignment:
        cell.alignment = _alignment(**alignment)
    if border:
        cell.border = _border(**border)
    if fill:
        cell.fill = _fill(**fill)
    if font:
        cell.font = _font(**font)
    if number_format is not None:
        cell.number_format = number_format
    if hyperlink is not None:
        cell.hyperlink = hyperlink


def get_cell_width(value: str) -> float:
    try:
        # 한글: 1.8배, 공백: 1.2배, 영문/숫자: 1배
        return sum(1.8 if ord(c) > 12799 else 1.2 if c.isspace() else 1. for c in value)
    except:
        return 0.


###################################################################
########################### Column utils ##########################
###################################################################

def get_column_index(column: Column, headers: list[tuple[str,...]] = list()) -> int | None:
    if isinstance(column, int):
        return column
    elif isinstance(column, str):
        for index, header in enumerate(headers, start=1):
            if (len(header) == 1) and (column == header[0]):
                return index
    elif isinstance(column, tuple):
        try:
            return headers.index(column) + 1
        except:
            pass
    return None


def get_column_letter(column: Column, headers: list[tuple[str,...]] = list()) -> str | None:
    from openpyxl.utils import get_column_letter as get_letter
    col_idx = get_column_index(column, headers)
    return get_letter(col_idx) if isinstance(col_idx, int) else None


def colstr(col_idx: int) -> str:
    from openpyxl.utils import get_column_letter as get_letter
    return get_letter(col_idx)


###################################################################
########################### Range utils ###########################
###################################################################

def get_ranges(
        ranges: list[Union[Column,Row,Range]] | Column | Row | Range,
        range_type: Literal["column","row","range","auto"],
        min_col: str,
        max_col: str,
        min_row: int,
        max_row: int,
        headers: list[tuple[str,...]] = list(),
    ) -> list[str]:

    def _auto_detect(value: Column | Row | Range) -> Literal["column","row","range","auto"]:
        if isinstance(value, int):
            return "row"
        elif isinstance(value, str):
            return "range" if is_range_string(value) else "column"
        elif isinstance(value, tuple):
            return "column"
        else:
            return "auto"

    def _make_range_string(value: Column | Row | Range, range_type: Literal["column","row","range","auto"]) -> str:
        if range_type == "auto":
            range_type = _auto_detect(value)

        if (range_type == "column") and (column := get_column_letter(value, headers)):
            return f"{column}{min_row}:{column}{max_row}"
        elif (range_type == "row") and isinstance(value, int):
            return f"{min_col}{value}:{max_col}{value}"
        elif (range_type == "range") and isinstance(value, str) and is_range_string(value):
            return value
        else:
            raise ValueError(f"Invalid Excel range format: {value}")

    if isinstance(ranges, list):
        return [range_string for value in ranges if (range_string := _make_range_string(value, range_type))]
    else:
        return [_make_range_string(ranges, range_type)]


def is_range_string(value: str) -> bool:
    import re
    if ':' in value:
        min_col, min_row, max_col, max_row = map(bool, split_range_string(value))
        if min_col and max_col:
            return min_row or (not max_row)
        else:
            return (not min_col) and (not max_col) and (min_row and max_row)
    else:
        col, row = r"\$?[A-Z]+", r"\$?[1-9][0-9]*"
        return re.match(r"^{}{}$".format(col, row), value)


def to_valid_excel_range(ws: Worksheet, value: str) -> str:
    if ':' not in value:
        return value

    def absolute(x: str) -> Literal['$','']:
        return '$' if x.startswith('$') else ''

    min_col, min_row, max_col, max_row = split_range_string(value)
    if not min_col:
        min_col = ((absolute(min_row) or absolute(max_col)) + 'A')
    if not min_row:
        min_row = ((absolute(min_col) or absolute(max_row)) + '1')
    if not max_col:
        max_col = ((absolute(max_row) or absolute(min_col)) + get_column_letter(ws.max_column))
    if not max_row:
        max_row = ((absolute(max_col) or absolute(min_row)) + str(ws.max_row))
    return f"{min_col}{min_row}:{max_col}{max_row}"


def split_range_string(value: str) -> tuple[str,str,str,str]:
    import re
    col, row = r"\$?[A-Z]+", r"\$?[1-9][0-9]*"
    cell = r"^({})?({})?$".format(col, row)
    top_left, bottom_right = value.split(':')
    min_col, min_row = match.groups() if (match := re.search(cell, top_left)) else (None, None)
    max_col, max_row = match.groups() if (match := re.search(cell, bottom_right)) else (None, None)
    return (min_col or str()), (min_row or str()), (max_col or str()), (max_row or str())


###################################################################
########################### Style config ##########################
###################################################################

def _init_column_styles(
        column_styles: dict[Column,StyleConfig],
        headers: list[tuple[str,...]] = list(),
    ) -> dict[int,StyleConfig]:
    return {col_idx: styles for column, styles in column_styles.items()
        if ((col_idx := get_column_index(column, headers)) is not None)}


def _init_column_width(
        column_width: float | Multiple | dict[Column,Width] | Literal[":fit_content:",":fit_header:",":fit_values:"],
        headers: list[tuple[str,...]] = list(),
    ) -> dict[int, Union[float,Literal[":fit_content:",":fit_header:",":fit_values:"]]]:

    def _set_width(value: Width) -> float | Literal[":fit_content:",":fit_header:",":fit_values:"]:
        if isinstance(value, str):
            if value in (":fit_content:",":fit_header:",":fit_values:"):
                return value
            elif value.endswith('x'):
                value = SINGLE_WIDTH * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(column_width, dict):
        return {col_idx: width for column, value in column_width.items()
                if ((col_idx := get_column_index(column, headers)) is not None)
                    and ((width := _set_width(value)) is not None)}
    else:
        value = _set_width(column_width)
        if value is not None:
            return {col_idx: value for column in headers
                    if (col_idx := get_column_index(column, headers)) is not None}
        else:
            return dict()


def _init_row_height(
        row_height: dict[Row,Height] | float | Multiple | Literal["single"],
    ) -> dict[int,float] | float | None:

    def _set_height(value: Width) -> float:
        if isinstance(value, str):
            if value == "single":
                return SINGLE_HEIGHT
            elif value.endswith('x'):
                value = SINGLE_HEIGHT * float(value[:-1])
        return float(value) if isinstance(value, (float,int)) and value > 0. else None

    if isinstance(row_height, dict):
        return {row_idx: height for row_idx, value in row_height.items()
                if (height := _set_height(value)) is not None}
    else:
        return _set_height(row_height)


###################################################################
########################### Style object ##########################
###################################################################

def _alignment(**kwargs) -> Alignment:
    from openpyxl.styles import Alignment
    return Alignment(**kwargs)


def _border(**kwargs: dict) -> Border:
    from openpyxl.styles import Border, Side
    def side(color: str | None = None, **kwargs) -> Side:
        return Side(color=(_color(color) if color is not None else None), **kwargs)
    return Border(**{property: side(**config) for property, config in kwargs.items()})


def _fill(color: str | None = None, **kwargs) -> PatternFill:
    from openpyxl.styles import PatternFill
    for property, value in kwargs.items():
        if property in {"fgColor","bgColor","start_color","end_color"}:
            kwargs[property] = _color(value)
    if color is not None:
        color = _color(color)
        kwargs.update(start_color=color, end_color=color)
    return PatternFill(**kwargs)


def _font(color: str | None = None, **kwargs) -> Font:
    from openpyxl.styles import Font
    return Font(color=(_color(color) if color is not None else None), **kwargs)


def _color(rgb: Any, alpha: str = "FF") -> Color:
    from openpyxl.styles import Color
    if isinstance(rgb, str):
        return Color((alpha + rgb[1:]) if rgb.startswith('#') else rgb)
    elif isinstance(rgb, dict):
        return Color(**rgb)
    elif isinstance(rgb, Color):
        return rgb
    else:
        return None


###################################################################
###################### Conditional formatting #####################
###################################################################

def _conditional_rule(
        operator: Literal[
            "endsWith", "containsText", "beginsWith", "lessThan", "notBetween", "lessThanOrEqual",
            "notEqual", "notContains", "between", "equal", "greaterThanOrEqual", "greaterThan", "formula"],
        formula: Sequence | None = None,
        stop_if_true: bool | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        **kwargs
    ) -> Rule:
    if operator == "formula":
        return _formula_rule(formula, stop_if_true, border, fill, font)

    from openpyxl.formatting.rule import CellIsRule
    styles = dict()
    if border:
        styles["border"] = _border(**border)
    if fill:
        styles["fill"] = _fill(**fill)
    if font:
        styles["font"] = _font(**font)
    return CellIsRule(operator=operator, formula=formula, stopIfTrue=stop_if_true, **styles)


def _formula_rule(
        formula: Sequence | None = None,
        stop_if_true: bool | None = None,
        border: dict | None = None,
        fill: dict | None = None,
        font: dict | None = None,
        **kwargs
    ) -> Rule:
    from openpyxl.formatting.rule import FormulaRule
    styles = dict()
    if border:
        styles["border"] = _border(**border)
    if fill:
        styles["fill"] = _fill(**fill)
    if font:
        styles["font"] = _font(**font)
    return FormulaRule(formula=formula, stopIfTrue=stop_if_true, **styles)


###################################################################
########################### Merge cells ###########################
###################################################################

def find_merge_ranges(
        ws: Worksheet,
        range_string: Range,
        mode: Literal["all","blank","same_value"] = "all",
        priority: Literal["by_row","by_col"] = "by_row",
    ) -> list[Range]:
    if mode == "all":
        return [range_string]
    merge_ranges = list()

    from collections import deque
    min_col, min_row, max_col, max_row = range_boundaries(range_string)
    num_rows, num_cols = (max_row - min_row + 1), (max_col - min_col + 1)
    adapter = {"by_row": "width", "by_col": "height"}

    rows = [list(row) for row in ws.iter_rows(min_row, max_row, min_col, max_col, values_only=True)]
    visited = [[False for _ in range(num_cols)] for _ in range(num_rows)]

    def _bfs(row_seq: int, col_seq: int) -> list[tuple[int,int]]:
        queue, cells = deque(), [(row_seq, col_seq)]
        queue.append((row_seq, col_seq))
        visited[row_seq][col_seq] = True
        value = rows[row_seq][col_seq]

        while queue:
            r, c = queue.popleft()
            for nr, nc in [(r, c+1), (r+1, c)]: # [Right, Down]
                if (0 <= nr < num_rows) and (0 <= nc < num_cols) and (not visited[nr][nc]):
                    if (rows[nr][nc] == value) if mode == "same_value" else (rows[nr][nc] is None):
                        visited[nr][nc] = True
                        queue.append((nr, nc))
                        cells.append((nr, nc))
        return cells

    for row_seq in range(num_rows):
        for col_seq in range(num_cols):
            if (not visited[row_seq][col_seq]) and (rows[row_seq][col_seq] is not None):
                cells = _bfs(row_seq, col_seq)
                if len(cells) < 2:
                    continue

                best_corners = get_largest_rectangle(cells, top_left=cells[0], priority=adapter[priority])
                (row_start, col_start) = best_corners[0][0] # Top-Left
                (row_end, col_end) = best_corners[1][1] # Bottom-Right

                for r, c in cells:
                    if not ((row_start <= r <= row_end) and (col_start <= c <= col_end)):
                        visited[r][c] = False

                if (row_start != row_end) or (col_start != col_end):
                    merge_ranges.append(f"{colstr(col_start)}{row_start+min_row}:{colstr(col_end)}{row_end+min_col}")
                else:
                    for r, c in cells[1:]:
                        visited[r][c] = False
    return merge_ranges


def get_largest_rectangle(
        nodes: list[tuple[int,int]],
        top_left: tuple[int,int],
        priority: Literal["width","height"] = "width",
    ) -> tuple[tuple[TopLeft,TopRight],tuple[BottomLeft,BottomRight]]:
    node_set = set(nodes)
    y0, x0 = top_left

    best_score = (0, 0)
    best_corners = (
        ((top_left, top_left), (top_left, top_left)),
        ((top_left, top_left), (top_left, top_left)),
    )

    max_width = 0
    while (y0, x0 + max_width) in node_set:
        max_width += 1

    min_height = float("inf")
    for width in range(1, max_width + 1):
        height = 0
        while (y0 + height, x0 + width - 1) in node_set:
            height += 1

        min_height = min(min_height, height)
        area = width * min_height
        score = (area, min_height if priority == "height" else width)

        if score > best_score:
            best_score = score
            y_max, x_max = (y0 + int(min_height) - 1), (x0 + width - 1)
            best_corners = (
                ((y0, x0), (y0, x_max)),
                ((y_max, x0), (y_max, x_max))
            )

    return best_corners


def range_boundaries(range_string: str) -> tuple[int,int,int,int]:
    from openpyxl.utils import range_boundaries as boundaries
    min_col, min_row, max_col, max_row = boundaries(range_string)
    return min_col, min_row, max_col, max_row
