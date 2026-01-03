"""Tests for xlsxslim v2.0.0 utility."""

import pytest
from pathlib import Path
from zipfile import ZipFile
from click.testing import CliRunner

# openpyxl is only used for creating test fixtures
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

from xlsxslim import (
    main,
    find_excel_files,
    generate_output_path,
    format_size,
    col_letter,
    col_index,
    parse_cell_ref,
    get_sheet_paths,
    analyze_sheet_xml,
    optimize_sheet_xml,
    EXIT_SUCCESS,
    EXIT_FILE_NOT_FOUND,
    EXIT_MULTIPLE_FILES,
    EXIT_ALREADY_OPTIMIZED,
)


# ============================================================================
# Test Fixtures (using openpyxl to create test xlsx files)
# ============================================================================

def create_simple_file(filepath: Path) -> Path:
    """Create a simple xlsx with minimal data."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["B1"] = "Value"
    ws["A2"] = "Data"
    ws["B2"] = 123
    wb.save(filepath)
    return filepath


def create_bloated_file(filepath: Path) -> Path:
    """Create a bloated xlsx with formatting beyond data range."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Add actual data (10 rows, 3 columns)
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["C1"] = "Note"
    for i in range(2, 12):
        ws[f"A{i}"] = f"Item {i-1}"
        ws[f"B{i}"] = i * 100
        ws[f"C{i}"] = f"Note {i-1}"
    
    # Add formatting to empty cells (simulate bloat)
    fill = PatternFill(start_color="FFFF00", fill_type="solid")
    for row in range(12, 1001):
        for col in range(1, 21):
            ws.cell(row=row, column=col).fill = fill
    
    wb.save(filepath)
    return filepath


def create_empty_file(filepath: Path) -> Path:
    """Create an xlsx with only formatting, no data."""
    wb = Workbook()
    ws = wb.active
    
    # Only formatting, no values
    fill = PatternFill(start_color="CCCCCC", fill_type="solid")
    for row in range(1, 101):
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = fill
    
    wb.save(filepath)
    return filepath


def create_multi_sheet_file(filepath: Path) -> Path:
    """Create xlsx with multiple sheets."""
    wb = Workbook()
    
    # Sheet 1: Data with bloat
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = "Value"
    ws1["A2"] = 100
    fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    for row in range(3, 501):
        ws1.cell(row=row, column=1).fill = fill
    
    # Sheet 2: Empty with formatting
    ws2 = wb.create_sheet("Empty Formatted")
    for row in range(1, 101):
        ws2.cell(row=row, column=1).fill = fill
    
    # Sheet 3: Clean (no bloat)
    ws3 = wb.create_sheet("Clean")
    ws3["A1"] = "Clean"
    ws3["B1"] = "Sheet"
    
    wb.save(filepath)
    return filepath


def create_file_with_formulas(filepath: Path) -> Path:
    """Create xlsx with formulas."""
    wb = Workbook()
    ws = wb.active
    
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=SUM(A1:A2)"
    ws["B1"] = "=A1*2"
    
    # Add bloat
    fill = PatternFill(start_color="FFFFFF", fill_type="solid")
    for row in range(4, 101):
        ws.cell(row=row, column=1).fill = fill
    
    wb.save(filepath)
    return filepath


# ============================================================================
# Tests for Helper Functions
# ============================================================================

class TestHelperFunctions:
    """Tests for helper functions."""

    def test_format_size_bytes(self):
        assert format_size(500) == "500.00 B"

    def test_format_size_kb(self):
        assert format_size(1024) == "1.00 KB"

    def test_format_size_mb(self):
        assert format_size(1024 * 1024) == "1.00 MB"

    def test_format_size_gb(self):
        assert format_size(1024 * 1024 * 1024) == "1.00 GB"

    def test_col_letter_single(self):
        assert col_letter(1) == "A"
        assert col_letter(26) == "Z"

    def test_col_letter_double(self):
        assert col_letter(27) == "AA"
        assert col_letter(28) == "AB"
        assert col_letter(52) == "AZ"

    def test_col_letter_triple(self):
        assert col_letter(703) == "AAA"

    def test_col_index(self):
        assert col_index("A") == 1
        assert col_index("Z") == 26
        assert col_index("AA") == 27
        assert col_index("AZ") == 52

    def test_parse_cell_ref(self):
        assert parse_cell_ref("A1") == (1, 1)
        assert parse_cell_ref("B5") == (5, 2)
        assert parse_cell_ref("Z100") == (100, 26)
        assert parse_cell_ref("AA50") == (50, 27)

    def test_generate_output_path_default(self):
        path = Path("/some/dir/report.xlsx")
        result = generate_output_path(path, "_slim")
        assert result == Path("/some/dir/report_slim.xlsx")

    def test_generate_output_path_custom_suffix(self):
        path = Path("data.xlsx")
        result = generate_output_path(path, "_optimized")
        assert result == Path("data_optimized.xlsx")


# ============================================================================
# Tests for XML Analysis
# ============================================================================

class TestXMLAnalysis:
    """Tests for XML-based analysis functions."""

    def test_get_sheet_paths(self, tmp_dir):
        filepath = create_simple_file(tmp_dir / "simple.xlsx")
        
        with ZipFile(filepath, 'r') as zf:
            paths = get_sheet_paths(zf)
        
        assert len(paths) == 1
        assert "Sheet" in paths
        assert "xl/worksheets/sheet1.xml" in paths["Sheet"]

    def test_analyze_sheet_simple(self, tmp_dir):
        filepath = create_simple_file(tmp_dir / "simple.xlsx")
        
        with ZipFile(filepath, 'r') as zf:
            paths = get_sheet_paths(zf)
            sheet_path = list(paths.values())[0]
            stats = analyze_sheet_xml(zf, sheet_path)
        
        assert stats['data_max_row'] == 2
        assert stats['data_max_col'] == 2
        assert stats['optimized'] == False  # No bloat to remove

    def test_analyze_sheet_bloated(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "bloated.xlsx")
        
        with ZipFile(filepath, 'r') as zf:
            paths = get_sheet_paths(zf)
            sheet_path = paths["Data"]
            stats = analyze_sheet_xml(zf, sheet_path)
        
        assert stats['data_max_row'] == 11  # 10 data rows + header
        assert stats['data_max_col'] == 3   # 3 columns
        assert stats['optimized'] == True   # Has rows to remove

    def test_analyze_sheet_with_formulas(self, tmp_dir):
        filepath = create_file_with_formulas(tmp_dir / "formulas.xlsx")
        
        with ZipFile(filepath, 'r') as zf:
            paths = get_sheet_paths(zf)
            sheet_path = list(paths.values())[0]
            stats = analyze_sheet_xml(zf, sheet_path)
        
        # Formulas should be detected as data
        assert stats['data_max_row'] == 3
        assert stats['data_max_col'] == 2


# ============================================================================
# Tests for File Discovery
# ============================================================================

class TestFindExcelFiles:
    """Tests for find_excel_files function."""

    def test_find_no_files(self, tmp_dir):
        result = find_excel_files(tmp_dir)
        assert result == []

    def test_find_single_file(self, tmp_dir):
        create_simple_file(tmp_dir / "test.xlsx")
        result = find_excel_files(tmp_dir)
        
        assert len(result) == 1
        assert result[0].name == "test.xlsx"

    def test_find_multiple_files(self, tmp_dir):
        create_simple_file(tmp_dir / "a.xlsx")
        create_simple_file(tmp_dir / "b.xlsx")
        result = find_excel_files(tmp_dir)
        
        assert len(result) == 2

    def test_ignores_temp_files(self, tmp_dir):
        create_simple_file(tmp_dir / "test.xlsx")
        create_simple_file(tmp_dir / "~$test.xlsx")
        
        result = find_excel_files(tmp_dir)
        
        assert len(result) == 1
        assert result[0].name == "test.xlsx"


# ============================================================================
# Tests for CLI
# ============================================================================

class TestCLI:
    """Tests for command-line interface."""

    def test_help(self):
        runner = CliRunner()
        result = runner.invoke(main, ["--help"])
        
        assert result.exit_code == 0
        assert "Reduce Excel file size" in result.output

    def test_version(self):
        runner = CliRunner()
        result = runner.invoke(main, ["--version"])
        
        assert result.exit_code == 0
        assert "2.0.2" in result.output

    def test_file_not_found(self):
        runner = CliRunner()
        result = runner.invoke(main, ["nonexistent.xlsx"])
        
        assert result.exit_code == EXIT_FILE_NOT_FOUND

    def test_no_files_in_dir(self, tmp_dir):
        runner = CliRunner()
        with runner.isolated_filesystem(temp_dir=tmp_dir):
            result = runner.invoke(main, [])
        
        assert result.exit_code == EXIT_FILE_NOT_FOUND

    def test_multiple_files_error(self, tmp_dir):
        create_simple_file(tmp_dir / "a.xlsx")
        create_simple_file(tmp_dir / "b.xlsx")
        
        runner = CliRunner()
        with runner.isolated_filesystem(temp_dir=tmp_dir):
            # Copy files to isolated filesystem
            import shutil
            shutil.copy(tmp_dir / "a.xlsx", "a.xlsx")
            shutil.copy(tmp_dir / "b.xlsx", "b.xlsx")
            result = runner.invoke(main, [])
        
        assert result.exit_code == EXIT_MULTIPLE_FILES

    def test_explicit_file(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "bloated.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath)])
        
        assert result.exit_code == EXIT_SUCCESS
        assert "Saved:" in result.output

    def test_custom_output(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "input.xlsx")
        output_path = tmp_dir / "custom_output.xlsx"
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-o", str(output_path)])
        
        assert result.exit_code == EXIT_SUCCESS
        assert output_path.exists()

    def test_inplace(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "inplace.xlsx")
        original_size = filepath.stat().st_size
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-i"])
        
        assert result.exit_code == EXIT_SUCCESS
        # File should be smaller after optimization
        assert filepath.stat().st_size < original_size

    def test_dry_run(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "bloated.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "--dry-run"])
        
        assert result.exit_code == EXIT_SUCCESS
        assert "Dry run" in result.output
        assert "Would delete" in result.output
        
        # No output file should be created
        output_file = tmp_dir / "bloated_slim.xlsx"
        assert not output_file.exists()

    def test_quiet_mode(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "bloated.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-q"])
        
        assert result.exit_code == EXIT_SUCCESS
        # Quiet mode should have minimal output
        assert "xlsxslim v" not in result.output

    def test_verbose_mode(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "bloated.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-v"])
        
        assert result.exit_code == EXIT_SUCCESS
        assert "Used range:" in result.output

    def test_custom_suffix(self, tmp_dir):
        filepath = create_bloated_file(tmp_dir / "test.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-s", "_optimized"])
        
        assert result.exit_code == EXIT_SUCCESS
        assert (tmp_dir / "test_optimized.xlsx").exists()

    def test_already_optimized(self, tmp_dir):
        filepath = create_simple_file(tmp_dir / "clean.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath)])
        
        assert result.exit_code == EXIT_ALREADY_OPTIMIZED
        assert "already optimized" in result.output

    def test_multi_sheet(self, tmp_dir):
        filepath = create_multi_sheet_file(tmp_dir / "multi_sheet.xlsx")
        
        runner = CliRunner()
        result = runner.invoke(main, [str(filepath), "-v"])
        
        assert result.exit_code == EXIT_SUCCESS
        assert "Data" in result.output


# ============================================================================
# Tests for Round-Trip (data preservation)
# ============================================================================

class TestRoundTrip:
    """Tests for data preservation after optimization."""

    def test_data_preserved(self, tmp_dir):
        """Verify cell values are preserved after optimization."""
        original = create_bloated_file(tmp_dir / "original.xlsx")
        optimized = tmp_dir / "optimized.xlsx"
        
        runner = CliRunner()
        result = runner.invoke(main, [str(original), "-o", str(optimized)])
        assert result.exit_code == EXIT_SUCCESS
        
        # Load both files and compare data
        wb_orig = load_workbook(original)
        wb_opt = load_workbook(optimized)
        
        ws_orig = wb_orig["Data"]
        ws_opt = wb_opt["Data"]
        
        # Check all data cells
        for row in range(1, 12):
            for col in range(1, 4):
                orig_val = ws_orig.cell(row=row, column=col).value
                opt_val = ws_opt.cell(row=row, column=col).value
                assert orig_val == opt_val, f"Mismatch at row {row}, col {col}"
        
        wb_orig.close()
        wb_opt.close()

    def test_formulas_preserved(self, tmp_dir):
        """Verify formulas are preserved after optimization."""
        original = create_file_with_formulas(tmp_dir / "formulas.xlsx")
        optimized = tmp_dir / "formulas_slim.xlsx"
        
        runner = CliRunner()
        result = runner.invoke(main, [str(original), "-o", str(optimized)])
        assert result.exit_code == EXIT_SUCCESS
        
        # Load optimized file and check formulas
        wb = load_workbook(optimized)
        ws = wb.active
        
        # Check formula cells
        assert ws["A3"].value == "=SUM(A1:A2)"
        assert ws["B1"].value == "=A1*2"
        
        wb.close()

    def test_file_size_reduced(self, tmp_dir):
        """Verify file size is actually reduced."""
        original = create_bloated_file(tmp_dir / "bloated.xlsx")
        optimized = tmp_dir / "bloated_slim.xlsx"
        
        orig_size = original.stat().st_size
        
        runner = CliRunner()
        result = runner.invoke(main, [str(original), "-o", str(optimized)])
        assert result.exit_code == EXIT_SUCCESS
        
        opt_size = optimized.stat().st_size
        
        # Optimized should be smaller
        assert opt_size < orig_size
        # Should save significant space (at least 50%)
        assert opt_size < orig_size * 0.5
