import os
import datetime
import subprocess
import traceback
from io import BytesIO
import pandas as pd
from bs4 import BeautifulSoup
from xlsx2html import xlsx2html
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill, Protection, Alignment

from .smb_functions import smb_load_file_obj, smb_store_remote_file_by_obj, smb_check_file_exist, smb_copy_file_remote_to_local, smb_copy_file_local_to_remote


def create_excel_file(username, password, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=445):
    """ This function is used to create an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
    """
    target_workbook = Workbook()
    target_sheet = target_workbook.active
    target_sheet.title = sheet_name
    output_io = BytesIO()
    target_workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_name, share_name, remote_file_path, output_io, port)
    return output_io


def load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=445, **kwargs):
    """ This function is used to load data from an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        **kwargs: Additional keyword arguments to pass to openpyxl's load_workbook function.
                  Common options include:
                  - read_only(bool): Whether to open workbook in read-only mode
                  - data_only(bool): Whether to read only cell values, not formulas
                  - keep_vba(bool): Whether to preserve VBA macros
                  - keep_links(bool): Whether to preserve external links
    """
    file_obj = smb_load_file_obj(username, password, server_name, share_name, remote_file_path, port)
    target_workbook = load_workbook(file_obj, **kwargs)
    is_target_sheet_exist = check_sheet_exists(target_workbook, sheet_name)
    if is_target_sheet_exist:
        target_sheet = target_workbook[sheet_name]
    else:
        target_sheet = None

    return target_workbook, target_sheet


def save_excel_file(username, password, server_name, share_name, remote_file_path, workbook, port=445):
    """ This function is used to save an Excel file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        workbook(Workbook): This is the instance of Workbook
    """
    output_io = BytesIO()
    workbook.save(output_io)
    output_io.seek(0)

    smb_store_remote_file_by_obj(username, password, server_name, share_name, remote_file_path, output_io, port)


def check_sheet_exists(workbook, sheet_name):
    """ This function is used to check whether the sheet exists in

    Args:
        workbook(Workbook): This is the instance of Workbook
        sheet_name(str): The name of the sheet.

    """
    return sheet_name in workbook.sheetnames


def add_column_filter(worksheet):
    """ This function is used to add column filter in the worksheet.

    Args:
        worksheet(Worksheet): This is the instance of Worksheet

    """
    worksheet.auto_filter.ref = worksheet.dimensions
    return worksheet


def auto_set_column_width(worksheet):
    """ This function is used to auto set column width in the worksheet.

    Args:
        worksheet (Worksheet): This is the instance of Worksheet
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if not isinstance(cell, MergedCell):  # Ensure it's not a merged cell
                if cell.value is not None:  # Only consider non-empty cells
                    # Convert cell value to string and calculate its length
                    cell_value_str = str(cell.value)
                    max_length = max(max_length, len(cell_value_str))

        # Adjust the width with some extra padding to handle different fonts and display
        adjusted_width = max_length + 4  # Add 4 to ensure it's wide enough
        worksheet.column_dimensions[column_letter].width = adjusted_width

    return worksheet


def manage_workbook_sheet(username, password, server_name, share_name, remote_file_path, rename_sheet_dict=None, add_sheet_name_list=None, delete_sheet_name_list=None,
                          port=445, **kwargs):
    """

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        rename_sheet_dict(dict): {old_sheet_name: new_sheet_name}
        add_sheet_name_list(list):This is the list of sheet that will be created
        delete_sheet_name_list(list): This is the list of sheet that will be deleted
        **kwargs: Additional keyword arguments to pass to load_workbook function

    """
    if rename_sheet_dict is None:
        rename_sheet_dict = {}
    if add_sheet_name_list is None:
        add_sheet_name_list = []
    if delete_sheet_name_list is None:
        delete_sheet_name_list = []

    target_workbook, _ = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=port, **kwargs)

    for old_sheet_name, new_sheet_name in rename_sheet_dict.items():
        if check_sheet_exists(target_workbook, old_sheet_name):
            target_workbook[old_sheet_name].title = new_sheet_name

    for add_sheet_name in add_sheet_name_list:
        if not check_sheet_exists(target_workbook, add_sheet_name):
            target_workbook.create_sheet(add_sheet_name)

    for delete_sheet_name in delete_sheet_name_list:
        if check_sheet_exists(target_workbook, delete_sheet_name):
            target_workbook.remove(target_workbook[delete_sheet_name])

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def save_dataframe_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                              keep_header=True, port=445, auto_filter=False, auto_column_width=False, **kwargs):
    """ This function is used to save data in Dataframe format to an Excel file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        keep_header(bool): Whether to keep the header of the data.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.DataFrame): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    data_columns = list(saved_data.columns)

    if keep_header:
        start_column_number = start_cell_column
        for column_index, column_name in enumerate(data_columns):
            target_sheet.cell(row=start_cell_row, column=start_column_number + column_index, value=column_name)
        start_cell_row += 1

    for column_name in data_columns:
        start_row_number = start_cell_row
        for column_value in saved_data[column_name]:
            target_sheet.cell(row=start_row_number, column=start_cell_column, value=column_value)
            start_row_number += 1
        start_cell_column += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def save_cell_values_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data_range_dict, port=445, auto_filter=False,
                                auto_column_width=False, **kwargs):
    """ This function is used to save cell values to an Excel file according to cell locations.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data_range_dict(dict): The data to be saved in the format of {cell_location: str}.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    for cell_location, cell_value in saved_data_range_dict.items():
        target_sheet[cell_location] = cell_value

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def save_series_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data, start_cell_location='A1',
                           port=445, auto_filter=False, auto_column_width=False, **kwargs):
    """ This function is used to save data in Series or list format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        start_cell_location(str): The location of start cell. e.g. A1
        saved_data(pd.Series | list): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)
    start_cell = target_sheet[start_cell_location]
    start_cell_row = start_cell.row
    start_cell_column = start_cell.column

    for save_value in saved_data:
        target_sheet.cell(row=start_cell_row, column=start_cell_column, value=save_value)
        start_cell_row += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def save_single_value_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data, target_cell_location='A1',
                                 port=445, auto_filter=False, auto_column_width=False, **kwargs):
    """ This function is used to save single value to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        target_cell_location(str): The location of target cell. e.g. A1
        saved_data(str|int|float): The data to be saved.
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    target_sheet[target_cell_location] = saved_data

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def save_multiple_value_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data, port=445, auto_filter=False,
                                   auto_column_width=False, **kwargs):
    """ This function is used to save multiple values to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(dict): The data to be saved.e.g. {'A1': 1, 'B2': 2}
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    for cell_location, cell_value in saved_data.items():
        target_sheet[cell_location] = cell_value

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def append_flexible_dataframe_into_excel(username, password, server_name, share_name, remote_file_path, sheet_name, saved_data, column_name_dict,
                                         port=445, auto_filter=False, auto_column_width=False, **kwargs):
    """ This function is used to append data in DataFrame format to an Excel file.

    Args:

        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        saved_data(pd.Dataframe): The data to be saved.
        column_name_dict(dict): The dictionary of column name and column index. e.g. {'name':'A', 'age':'C'}
        auto_filter(bool): Whether to add column filter.
        auto_column_width(bool): Whether to auto set column width.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    max_row_number = target_sheet.max_row

    for column_name, column_index in column_name_dict.items():
        start_row_number = max_row_number + 1
        for column_value in saved_data[column_name]:
            target_sheet[f'{column_index}{start_row_number}'] = column_value
            start_row_number += 1

    if auto_filter:
        target_sheet = add_column_filter(target_sheet)

    if auto_column_width:
        auto_set_column_width(target_sheet)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def copy_worksheet(username, password, server_name, share_name, source_file_path, source_sheet_name, target_file_path, target_sheet_name, port=445,
                   remove_existed_target_sheet=False, **kwargs):
    """ This function is used to copy a worksheet to another worksheet.

    Args:
        username(str): This is the username.
        password(str): This is the password.
        server_name(str): The server name (URL).
        share_name(str): The share name of the public folder.
        source_file_path(str): The public file path for the source.
        target_file_path(str): The public file path for the target.
        port(int): The port number of the server name.
        source_sheet_name(str): The name of the source sheet.
        target_sheet_name(str): The name of the target sheet.
        remove_existed_target_sheet(bool): Whether to remove the existed target sheet.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    # Open the source workbook and target workbook
    source_workbook, source_sheet = load_excel_data(username, password, server_name, share_name, source_file_path, sheet_name=source_sheet_name, port=port, **kwargs)
    target_workbook, _ = load_excel_data(username, password, server_name, share_name, target_file_path, sheet_name=target_sheet_name, port=port, **kwargs)

    is_target_sheet_exist = check_sheet_exists(target_workbook, target_sheet_name)
    if not is_target_sheet_exist:
        target_workbook.create_sheet(target_sheet_name)
    elif is_target_sheet_exist and remove_existed_target_sheet:
        # 获取旧 Sheet 的位置索引
        idx = target_workbook.index(target_workbook[target_sheet_name])
        target_workbook.remove(target_workbook[target_sheet_name])
        # 在原位置创建新 Sheet
        target_workbook.create_sheet(target_sheet_name, index=idx)

    target_sheet = target_workbook[target_sheet_name]

    # Copy cell contents and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]

            if isinstance(cell, MergedCell):
                new_cell.value = ''
            else:
                new_cell.value = cell.value

            # Check and copy styles
            if cell.has_style:
                new_cell.font = Font(name=cell.font.name,
                                     size=cell.font.size,
                                     bold=cell.font.bold,
                                     italic=cell.font.italic,
                                     vertAlign=cell.font.vertAlign,
                                     underline=cell.font.underline,
                                     strike=cell.font.strike,
                                     color=cell.font.color)

                new_cell.border = Border(left=cell.border.left,
                                         right=cell.border.right,
                                         top=cell.border.top,
                                         bottom=cell.border.bottom)

                if cell.fill and isinstance(cell.fill, PatternFill):
                    new_cell.fill = PatternFill(start_color=cell.fill.start_color,
                                                end_color=cell.fill.end_color,
                                                fill_type=cell.fill.fill_type)

                new_cell.number_format = cell.number_format  # Copy number format

                new_cell.protection = Protection(locked=cell.protection.locked,
                                                 hidden=cell.protection.hidden)

                new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                               vertical=cell.alignment.vertical,
                                               text_rotation=cell.alignment.text_rotation,
                                               wrap_text=cell.alignment.wrap_text,
                                               shrink_to_fit=cell.alignment.shrink_to_fit,
                                               indent=cell.alignment.indent)

    # Copy column widths
    for col in source_sheet.columns:
        column_letter = get_column_letter(col[0].column)
        target_sheet.column_dimensions[column_letter].width = source_sheet.column_dimensions[column_letter].width

    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    save_excel_file(username, password, server_name, share_name, target_file_path, target_workbook, port)


def update_cell_value_alignment(worksheet, cell_start_location, cell_end_location, horizontal_alignment='center', vertical_alignment='center'):
    """ This function is used to update the alignment of the cell value.

    Args:

        horizontal_alignment(str): The horizontal alignment of the cell value. e.g. center, left, right
        vertical_alignment(str): The vertical alignment of the cell value. e.g. center, top, bottom
        worksheet(Worksheet): This is the instance of Worksheet
        cell_start_location(str): The start location of the cell. e.g. A1
        cell_end_location(str): The end location of the cell. e.g. A1
    """
    for row in worksheet[f"{cell_start_location}:{cell_end_location}"]:
        for cell in row:
            cell.alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment)


def smb_store_remote_file_by_df(username, password, server_name, share_name, remote_file_path, df_data, sheet_name, port=445):
    """ This function is used to store a pandas DataFrame to a specific sheet in a remote Excel file

    Args:
        username (str): The username for the SMB connection
        password (str): The password for the SMB connection
        server_name (str): The server name (URL), e.g., szh0fs06.apac.bosch.com
        share_name (str): The share name of the public folder, e.g., GS_ACC_CN$
        remote_file_path (str): The path to the remote file under the share name folder
        df_data(pandas.DataFrame): The DataFrame to be written to the Excel file
        sheet_name (str): The name of the sheet to write the DataFrame to
        port (int): The port number of the server (default is 445)
    """
    try:
        is_file_exist, file_obj = smb_check_file_exist(username, password, server_name, share_name, remote_file_path, port)
        if not is_file_exist:
            file_obj = create_excel_file(username, password, server_name, share_name, remote_file_path, sheet_name, port)

        with pd.ExcelWriter(file_obj, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_data.to_excel(writer, sheet_name=sheet_name, index=False)
        file_obj.seek(0)

        smb_store_remote_file_by_obj(username, password, server_name, share_name, remote_file_path, file_obj, port)

    except Exception as e:
        print(f"Ops, error is: {e}")


def delete_excel_rows(username, password, server_name, share_name, remote_file_path, sheet_name, row_number_list, port=445, **kwargs):
    """ This function is used to delete rows from an Excel file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        row_number_list(list(int)): This is the list of row numbers to be deleted. The row numbers should be in descending order.
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    row_number_list = sorted(row_number_list, reverse=True)
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    for row_number in row_number_list:
        target_sheet.delete_rows(row_number)

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def convert_html_to_pdf(input_html, output_pdf):
    """ This function is used to convert HTML file to PDF file using wkhtmltopdf.

    Args:
        input_html(str): The path to the input HTML file.
        output_pdf(str): The path to the output PDF file.
    """
    try:
        from weasyprint import HTML
    except ImportError:
        HTML = None
        print("WeasyPrint is not installed or missing dependencies. HTML to PDF functionality will be unavailable.")

    try:
        with open(input_html, "r", encoding="utf-8") as f:
            html_content = f.read()

        # additional_style = """
        #     <style>
        #         @page {
        #             size: A4 portrait;
        #         }
        #
        #         body {
        #             font-family: Arial, sans-serif;
        #             font-size: 10pt;
        #             transform: scale(0.6);
        #             transform-origin: top left;
        #         }
        #
        #         table {
        #             width: 100%;
        #             border-collapse: collapse;
        #         }
        #
        #         th, td {
        #             word-wrap: break-word;
        #         }
        #     </style>
        # """
        additional_style = """
                    <style>
                        /* 1. 纸张设置：设置为 A4 横向 (Landscape)，并留出边距 */
                        @page {
                            size: A4 landscape;
                            margin: 1cm;
                        }

                        body {
                            /* 使用 Linux 安全字体 */
                            font-family: "DejaVu Sans", "Liberation Sans", sans-serif;
                            /* 稍微缩小字号，以便塞下更多内容 */
                            font-size: 9pt; 
                        }

                        /* 2. 表格强制适配 */
                        table {
                            /* 强制表格宽度不超过页面宽度 */
                            width: 100% !important;
                            max-width: 100% !important;
                            /* 自动计算列宽，而不是用 Excel 的固定像素宽度 */
                            table-layout: auto !important; 
                            border-collapse: collapse;
                        }

                        /* 3. 单元格内容控制 */
                        td, th {
                            /* 必须允许自动换行，否则长文本会把表格撑破 */
                            white-space: normal !important; 
                            word-wrap: break-word !important;
                            overflow-wrap: break-word !important;

                            /* 确保内容可见 */
                            height: auto !important;
                            overflow: visible !important;

                            /* 加上边框方便调试，看看到底哪里溢出了 (不需要可以删掉) */
                            border: 1px solid #ccc; 
                        }

                        /* 4. 防止图片过大 */
                        img {
                            max-width: 100% !important;
                            height: auto !important;
                        }
                    </style>
                """

        soup = BeautifulSoup(html_content, "html.parser")
        if not soup.head:
            head = soup.new_tag("head")
            soup.html.insert(0, head)
        else:
            head = soup.head
        head.append(BeautifulSoup(additional_style, "html.parser"))

        HTML(string=str(soup)).write_pdf(output_pdf)
        print(f"PDF export succeeded: {output_pdf}")
    except Exception:
        print(f"PDF export failed: {traceback.format_exc()}")


def export_excel_sheet_as_pdf(username, password, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=445):
    """ This function is used to export a specific sheet from an Excel file as a pandas DataFrame.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet to be exported.
    """
    save_folder_path = "/opt/Process_Folder"
    if not os.path.exists(save_folder_path):
        os.mkdir(save_folder_path)

    file_name = os.path.basename(remote_file_path)
    local_file_path = os.path.join(save_folder_path, file_name)
    local_file_name, _ = os.path.splitext(file_name)
    local_html_file_path = os.path.join(save_folder_path, f"{local_file_name}.html")
    local_pdf_file_path = os.path.join(save_folder_path, f"{local_file_name}.pdf")
    remote_pdf_file_path = remote_file_path.replace('.xlsx', '.pdf').replace('.XLSX', '.pdf')
    remote_html_file_path = remote_file_path.replace('.xlsx', '.html').replace('.XLSX', '.html')

    smb_copy_file_remote_to_local(username, password, server_name, share_name, local_file_path, remote_file_path, port)
    xlsx2html(local_file_path, local_html_file_path, sheet=sheet_name)
    convert_html_to_pdf(local_html_file_path, local_pdf_file_path)
    smb_copy_file_local_to_remote(username, password, server_name, share_name, local_pdf_file_path, remote_pdf_file_path, port)
    smb_copy_file_local_to_remote(username, password, server_name, share_name, local_html_file_path, remote_html_file_path, port)


def locate_first_empty_cell_in_target_column(username, password, server_name, share_name, remote_file_path, port, sheet_name, column_name, row_number=1, **kwargs):
    """ This function is used to locate the first empty cell in a worksheet.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        column_name(str): The name of the column to search for empty cells.
        row_number(int): The row number to start searching from (default is 1).
        **kwargs: Additional keyword arguments to pass to load_workbook function
    """
    max_excel_rows = 1048576
    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    for row in range(row_number, max_excel_rows + 1):
        cell_value = target_sheet[f"{column_name}{row}"].value
        print('cell value: ', cell_value)
        if cell_value is None or not str(cell_value).strip():
            return f"{column_name}{row}"

    raise Exception("No empty cell found up to Excel's row limit.")


def set_column_date_format(username, password, server_name, share_name, remote_file_path, sheet_name='Sheet1', port=445, column_index_list=None, start_row=2, end_row=None,
                           date_format='yyyy-mm-dd', **kwargs):
    """ This function is used to set the date format for a specific column in an Excel file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        remote_file_path(str): This is the public file path that file will be saved under share name folder
        port(int): This is the port number of the server name
        sheet_name(str): The name of the sheet.
        column_index_list(list): The list of column indices to set the date format (1 for column A, 2 for column B, etc.). If None, it will apply to all columns.
        start_row(int): The starting row number for applying the date format (default is 2).
        end_row(int): The ending row number for applying the date format (default is None, which means it will apply to all rows until the end of the column).
        date_format(str): The date format to apply to the column (default is yyyy-mm-dd).
        **kwargs: Additional keyword arguments to pass to load_workbook function

    """
    if column_index_list is None:
        column_index_list = []

    target_workbook, target_sheet = load_excel_data(username, password, server_name, share_name, remote_file_path, sheet_name=sheet_name, port=port, **kwargs)

    if end_row is None:
        end_row = target_sheet.max_row

    for column_index in column_index_list:
        for row in target_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=column_index, max_col=column_index):
            for cell in row:
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = date_format

    save_excel_file(username, password, server_name, share_name, remote_file_path, target_workbook, port)


def transfer_xls_into_xlsx(username, password, server_name, share_name, xls_file_path, xlsx_file_path, sheet_name='Sheet1', port=445, encoding='utf-8'):
    """ This function is used to transfer an xls file into xlsx file.

    Args:
        username(str): This is the username
        password(str): This is the password
        server_name(str):This is the server name (url), e.g. szh0fs06.apac.bosch.com
        share_name(str): This is the share_name of public folder, e.g. GS_ACC_CN$
        xls_file_path(str): This is the public file path of the xls file that will be saved under share name folder
        xlsx_file_path(str): This is the public file path of the xlsx file that will be saved under share name folder
        sheet_name(str): The name of the sheet.
        port(int): This is the port number of the server name
        encoding(str): The encoding of the xls file, default is 'utf-8'.
    """
    xls_file_obj = smb_load_file_obj(username, password, server_name, share_name, xls_file_path, port)
    try:
        xls_file_data = pd.read_csv(xls_file_obj, sep='\t', encoding=encoding, dtype=str)
    except:
        xls_file_data = pd.read_excel(xls_file_obj, dtype=str)

    file_obj = BytesIO()

    with pd.ExcelWriter(file_obj, engine='xlsxwriter') as writer:
        xls_file_data.to_excel(writer, sheet_name=sheet_name, index=False, float_format='%.2f')

    file_obj.seek(0)

    smb_store_remote_file_by_obj(username, password, server_name, share_name, xlsx_file_path, file_obj, port)
