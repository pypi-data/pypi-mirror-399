"""
This module provides the ExcelReader class to read data from Excel files with .xlsx format.

The ExcelReader class offers methods to read specific cells, rows, and convert the row data to JSON format.
Additionally, it provides functionalities to count the total number of rows and nodes in the JSON data.

Classes:
--------
ExcelReader

Methods:
--------
read_cells(sheet_name: str, cell_references: List[str]) -> List[Any]:
    Reads specific cells from a given sheet.

read_rows(sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> List[List[Any]]:
    Reads data from a range of rows in a given sheet, with the option to remove a specified number of rows from the end.

rows_to_json(sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> str:
    Converts row data to structured JSON format.

get_row_count(data: List[List[Any]]) -> int:
    Returns the total number of rows in the provided data.

get_node_count(json_data: str) -> int:
    Returns the total number of nodes in the provided JSON data.
"""
import json
from typing import List, Any
from openpyxl import load_workbook


class ExcelReader:
    """
    A class to read data from Excel files with .xlsx format.

    Attributes:
    -----------
    file_path : str
        The path to the Excel file.

    Methods:
    --------
    read_cell(sheet_name: str, cell_reference: str) -> str:
        Reads a specific cell from a given sheet.

    read_cells(sheet_name: str, cell_references: List[str]) -> List[Any]:
        Reads specific cells from a given sheet.

    read_rows(sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> List[List[Any]]:
        Reads data from a range of rows in a given sheet, with the option to remove a specified number of rows from the end.

    rows_to_json(sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> str:
        Converts row data to structured JSON format.

    get_row_count(data: List[List[Any]]) -> int:
        Returns the total number of rows in the provided data.
    """

    def __init__(self, file_path: str):
        """
        Initializes the ExcelReader with the path to the Excel file.

        Parameters:
        -----------
        file_path : str
            The path to the Excel file.
        """
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path)

    def read_cell(self, sheet_name: str, cell_reference: str) -> str:
        """
        Reads a specific cell from a given sheet.

        Parameters:
        -----------
        sheet_name : str
            The name of the sheet to read from.
        cell_reference : str
            A list of cell references to read (e.g., ['A1', 'B2']).

        Returns:
        --------
        str
            The value from the specified cell as a string.
        """
        sheet = self.workbook[sheet_name]
        data = sheet[cell_reference].value
        return str(data) if data is not None else ""

    def read_cells(self, sheet_name: str, cell_references: List[str]) -> List[Any]:
        """
        Reads specific cells from a given sheet.

        Parameters:
        -----------
        sheet_name : str
            The name of the sheet to read from.
        cell_references : List[str]
            A list of cell references to read (e.g., ['A1', 'B2']).

        Returns:
        --------
        List[Any]
            A list of values from the specified cells.
        """
        sheet = self.workbook[sheet_name]
        data = [sheet[cell].value for cell in cell_references]
        return data

    def read_rows(self, sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> List[List[Any]]:
        """
        Reads data from a range of rows in a given sheet, with the option to remove a specified number of rows from the end.

        Parameters:
        -----------
        sheet_name : str
            The name of the sheet to read from.
        start_col : str
            The starting column letter.
        end_col : str
            The ending column letter.
        start_row : int, optional
            The row to start reading from (default is 1).
        rows_to_remove : int, optional
            The number of rows to remove from the end (default is 0).

        Returns:
        --------
        List[List[Any]]
            A list of lists where each inner list contains the data from one row.
        """
        sheet = self.workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_col=self._col_to_index(start_col),
                                   max_col=self._col_to_index(end_col),
                                   min_row=start_row):
            row_data = [str(cell.value) for cell in row if cell.value is not None]
            data.append(row_data)

        if rows_to_remove > 0:
            data = self._remove_rows_from_end(data, rows_to_remove)

        return data

    def rows_to_json(self, sheet_name: str, start_col: str, end_col: str, start_row: int = 1, rows_to_remove: int = 0) -> str:
        """
        Converts row data to structured JSON format.

        Parameters:
        -----------
        sheet_name : str
            The name of the sheet to read from.
        start_col : str
            The starting column letter.
        end_col : str
            The ending column letter.
        start_row : int, optional
            The row to start reading from (default is 1).
        rows_to_remove : int, optional
            The number of rows to remove from the end (default is 0).

        Returns:
        --------
        str
            JSON string representing the structured row data.
        """
        row_data = self.read_rows(sheet_name, start_col, end_col, start_row, rows_to_remove)
        json_data = {"row_" + str(i + start_row): row_data[i] for i in range(len(row_data))}
        return json.dumps(json_data, indent=4, ensure_ascii=False)

    def get_row_count(self, data: List[List[Any]]) -> int:
        """
        Returns the total number of rows in the provided data.

        Parameters:
        -----------
        data : List[List[Any]]
            The list of rows to count.

        Returns:
        --------
        int
            The total number of rows.
        """
        return len(data)

    def _col_to_index(self, col: str) -> int:
        """
        Converts a column letter to a 1-based index.

        Parameters:
        -----------
        col : str
            The column letter (e.g., 'A', 'B', ..., 'Z', 'AA', etc.).

        Returns:
        --------
        int
            The 1-based index of the column.
        """
        index = 0
        for char in col:
            index = index * 26 + (ord(char.upper()) - ord('A')) + 1
        return index

    def _remove_rows_from_end(self, data: List[List[Any]], rows_to_remove: int) -> List[List[Any]]:
        """
        Removes a specified number of rows from the end of the data.

        Parameters:
        -----------
        data : List[List[Any]]
            The list of rows to process.
        rows_to_remove : int
            The number of rows to remove from the end.

        Returns:
        --------
        List[List[Any]]
            The modified list of rows with the specified number of rows removed from the end.
        """
        non_empty_row_index = len(data)
        while non_empty_row_index > 0 and not any(data[non_empty_row_index - 1]):
            non_empty_row_index -= 1

        return data[:non_empty_row_index - rows_to_remove] if rows_to_remove > 0 else data
