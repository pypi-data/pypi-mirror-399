"""
This module defines a Sharepoint class that facilitates interactions with a SharePoint site.
It provides methods for authenticating with the site, listing files in a specified document
library folder, downloading files, and saving them to a local directory. The class is designed
to encapsulate all necessary functionalities for handling files on a SharePoint site, making it
suitable for scripts or applications that require automated access to SharePoint resources.

The Sharepoint class uses the SharePlum library to communicate with SharePoint, handling common
tasks such as authentication, file retrieval, and file management. This includes methods to
authenticate users, fetch file lists from specific library folders, download individual files,
and save them locally. The class is initialized with user credentials and site details, which
are used throughout the class to manage SharePoint interactions.

Usage:
    After creating an instance of the Sharepoint class with the necessary credentials and site details,
    users can call methods to list files in a folder, download a specific file, or retrieve and save
    all files from a folder to a local directory. This makes it easy to integrate SharePoint file
    management into automated workflows or systems.

Example:
    sharepoint_details = {
        "username": "john@do.e",
        "password": "johndoe",
        "site_url": "https://site_url",
        "site_name": "department123",
        "document_library": "Shared documents"
    }
    sp = Sharepoint(**sharepoint_details)
    sp.download_files("FolderName", "C:\\LocalPath")
"""

import os

import math

import traceback

from pathlib import PurePath

from io import BytesIO

from typing import Optional, List, Dict, Any, Union

from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

import pandas as pd

from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File


class Sharepoint:
    """
    A class to interact with a SharePoint site, enabling authentication, file listing,
    downloading, uploading, and saving functionalities within a specified SharePoint document library.

    Attributes:
        tenant (str): Tenant name or domain (e.g. 'yourtenant.onmicrosoft.com').
        client_id (str): The Azure AD application (client) ID.
        thumbprint (str): The certificate thumbprint registered in Azure AD.
        cert_path (str): Path to the PEM file containing the private key and certificate.
        site_url (str): Base URL of the SharePoint tenant (e.g. 'https://contoso.sharepoint.com').
        site_name (str): Name of the SharePoint site (e.g. 'MyTeamSite').
        document_library (str): Name of the document library within the site.
    """

    def __init__(
            self, tenant: str, client_id: str, thumbprint: str, cert_path: str, site_url: str, site_name: str, document_library: str
    ):
        """Initializes the Sharepoint class with credentials and site details."""
        self.tenant = tenant
        self.client_id = client_id
        self.thumbprint = thumbprint
        self.cert_path = cert_path
        self.site_url = site_url
        self.site_name = site_name
        self.document_library = document_library
        self.ctx = self._auth()

    def _auth(self):
        """
        Authenticates to the SharePoint site and returns the client context.

        Returns:
            Optional[ClientContext]: A ClientContext object for interacting with the SharePoint site if authentication is successful,
                            otherwise None.
        """
        try:
            site_full_url = f"{self.site_url}/teams/{self.site_name}"

            ctx = ClientContext(site_full_url).with_client_certificate(
                tenant=self.tenant,
                client_id=self.client_id,
                thumbprint=self.thumbprint,
                cert_path=self.cert_path
            )

            web = ctx.web
            ctx.load(web)
            ctx.execute_query()

            print(f"Authenticated successfully. Site Title: {web.properties['Title']}")

            return ctx
        except Exception as e:
            print(f"Failed to authenticate: {e}")
            return None

    def fetch_files_list(self, folder_name: str) -> Optional[List[dict]]:
        """
        Retrieves a list of files from a specified folder within the document library.

        Args:
            folder_name (str): The name of the folder within the document library.

        Returns:
            list: A list of file dictionaries in the specified folder, or None if an error occurs or if the site is not authenticated.
        """
        if self.ctx:
            try:
                folder_url = f"/teams/{self.site_name}/{self.document_library}/{folder_name}"
                folder = self.ctx.web.get_folder_by_server_relative_url(folder_url)
                files = folder.files

                self.ctx.load(files)
                self.ctx.execute_query()

                files_list = [{"Name": file.name} for file in files]

                return files_list

            except Exception as e:

                print(f"Error retrieving files: {e}")

                return None

        return None

    def fetch_file_content(self, file_name: str, folder_name: str) -> Optional[bytes]:
        """
        Downloads a file from a specified folder within the document library.

        Args:
            file_name (str): The name of the file to be downloaded.
            folder_name (str): The name of the folder where the file is located.

        Returns:
            Optional[bytes]: The binary content of the file if successful, otherwise None.
        """
        if self.ctx:
            try:
                file_url = f"/teams/{self.site_name}/{self.document_library}/{folder_name}/{file_name}"
                file = self.ctx.web.get_file_by_server_relative_url(file_url)
                file_content = file.read().execute_query()
                return file_content.value
            except Exception as e:
                print(f"Failed to download file: {e}")
                return None
        return None

    def fetch_file_using_open_binary(self, file_name: str, folder_name: str) -> Optional[bytes]:
        """
        Downloads a file using the open_binary method from SharePoint.
        """
        if self.ctx:
            try:
                file_url = f"/teams/{self.site_name}/{self.document_library}/{folder_name}/{file_name}"

                file_content = File.open_binary(self.ctx, file_url)

                return file_content.content

            except Exception:
                print("Failed to download file:")

                traceback.print_exc()

                return None

        return None

    def _write_file(self, folder_destination: str, file_name: str, file_content: bytes):
        """
        Saves the binary content of a file to a specified local destination.

        Args:
            folder_destination (str): The local folder path where the file will be saved.
            file_name (str): The name of the file to be saved.
            file_content (bytes): The binary content of the file.
        """
        file_directory_path = PurePath(folder_destination, file_name)
        with open(file_directory_path, "wb") as file:
            file.write(file_content)

    def download_file(self, folder: str, filename: str, folder_destination: str):
        """
        Downloads a specified file from a specified folder and saves it to a local destination.

        Args:
            folder (str): The name of the folder in the document library containing the file.
            filename (str): The name of the file to download.
            folder_destination (str): The local folder path where the downloaded file will be saved.
        """
        file_content = self.fetch_file_content(filename, folder)
        if file_content:
            self._write_file(folder_destination, filename, file_content)
        else:
            print(f"Failed to download {filename}")

    def download_files(self, folder: str, folder_destination: str):
        """
        Downloads all files from a specified folder and saves them to a local destination.

        Args:
            folder (str): The name of the folder in the document library containing the files.
            folder_destination (str): The local folder path where the downloaded files will be saved.
        """
        files_list = self.fetch_files_list(folder)
        if files_list:
            for file in files_list:
                file_content = self.fetch_file_content(file["Name"], folder)
                if file_content:
                    self._write_file(folder_destination, file["Name"], file_content)
                else:
                    print(f"Failed to download {file['Name']}")
        else:
            print(f"No files found in folder {folder}")

    def upload_file(self, folder_name: str, file_path: str, file_name: Optional[str] = None):
        """
        Uploads a single file to a specified folder within the document library.

        Args:
            folder_name (str): The name of the folder within the document library.
            file_path (str): The local path to the file to be uploaded.
            file_name (Optional[str]): The name to give the file in SharePoint. If not provided, uses the name from file_path.
        """
        if self.ctx:
            try:
                if file_name is None:
                    file_name = os.path.basename(file_path)

                folder_url = f"/teams/{self.site_name}/{self.document_library}/{folder_name}"
                target_folder = self.ctx.web.get_folder_by_server_relative_url(folder_url)

                with open(file_path, 'rb') as content_file:
                    file_content = content_file.read()

                target_folder.upload_file(file_name, file_content).execute_query()
                print(f"File '{file_name}' uploaded successfully to '{folder_url}'.")
            except Exception as e:
                print(f"Failed to upload file '{file_name}': {e}")

    def upload_files(self, folder_name: str, files: List[str]):
        """
        Uploads multiple files to a specified folder within the document library.

        Args:
            folder_name (str): The name of the folder within the document library.
            files (List[str]): A list of local file paths to be uploaded.
        """
        if self.ctx:
            for file_path in files:
                try:
                    file_name = os.path.basename(file_path)
                    self.upload_file(folder_name, file_path, file_name)
                except Exception as e:
                    print(f"Failed to upload file '{file_path}': {e}")

    def upload_file_from_bytes(self, binary_content: bytes, file_name: str, folder_name: str):
        """
        Uploads a file to SharePoint directly from a bytes object.

        Args:
            binary_content (bytes): The binary content of the file.
            file_name (str): The name to give the file in SharePoint.
            folder_name (str): The folder in the document library where the file will be uploaded.
        """

        if self.ctx:
            try:
                folder_url = f"/teams/{self.site_name}/{self.document_library}/{folder_name}"

                target_folder = self.ctx.web.get_folder_by_server_relative_url(folder_url)

                target_folder.upload_file(file_name, binary_content).execute_query()

                print(f"File '{file_name}' uploaded successfully to '{folder_url}'.")

            except Exception as e:
                print(f"Failed to upload file '{file_name}': {e}")

    def append_row_to_sharepoint_excel(
        self,
        required_headers: Optional[List[str]] = None,
        folder_name: str = "",
        excel_file_name: str = "",
        sheet_name: str = "",
        new_rows: Union[Dict, List[Dict]] = None,
    ) -> None:
        """
        • Appends one or more rows to an existing Excel file.
        • Sorts and formats based on provided parameters.
        """

        # Ensure new_rows is a list of dicts
        if isinstance(new_rows, dict):
            new_rows = [new_rows]

        elif not isinstance(new_rows, list) or not all(isinstance(r, dict) for r in new_rows):
            raise TypeError("new_rows must be a dict or a list of dicts.")

        # 1. Pull file
        binary_file = self.fetch_file_using_open_binary(excel_file_name, folder_name)
        if binary_file is None:
            raise FileNotFoundError(f"File '{excel_file_name}' not found in folder '{folder_name}'.")

        wb = load_workbook(BytesIO(binary_file))

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{excel_file_name}'")

        ws = wb[sheet_name]

        # 2. Validate headers
        if required_headers:
            current_headers = [cell.value for cell in ws[1]]

            if current_headers != required_headers:
                raise ValueError(
                    f"Header mismatch in sheet '{sheet_name}'!\n"
                    f"Expected: {required_headers}\n"
                    f"Found:    {current_headers}"
                )

        # 2.5 Clean up empty rows before appending
        for row_idx in range(ws.max_row, 1, -1):  # Start from bottom, skip header
            row_values = [cell.value for cell in ws[row_idx]]

            if all(cell is None for cell in row_values):
                ws.delete_rows(row_idx)

        # 3. Append each new row
        headers = [header.value for header in ws[1]]

        for row_dict in new_rows:
            ws.append([row_dict.get(header, "") for header in headers])

        # 4. Save and upload
        temp_stream = BytesIO()

        wb.save(temp_stream)

        temp_stream.seek(0)

        self.upload_file_from_bytes(temp_stream.getvalue(), excel_file_name, folder_name)

    def format_and_sort_excel_file(
        self,
        folder_name: str,
        excel_file_name: str,
        sheet_name: str,
        sorting_keys: Optional[List[Dict[str, Any]]] = None,
        font_config: Optional[Dict[int, Dict[str, Any]]] = None,
        bold_rows: Optional[List[int]] = None,
        italic_rows: Optional[List[int]] = None,
        align_horizontal: str = "center",
        align_vertical: str = "center",
        column_widths: Any = "auto",
        freeze_panes: Optional[str] = None,
    ):
        """
        Sorts and formats an Excel worksheet based on provided styling and sorting rules.

        Params:
            folder_name: Name of the folder where the file resides
            excel_file_name: Name of the excel file
            sheet_name: Name of the sheet that will be sorted
            sorting_keys: List of dicts like [{"key": "A", "ascending": True, "type": "datetime"}]
            bold_rows: List of row numbers to bold (defaults to [1])
            italic_rows: List of row numbers to italicize
            font_config: Dict of row -> font config (overrides bold/italic)
            align_horizontal: Horizontal text alignment
            align_vertical: Vertical text alignment
            column_widths: "auto" or an int to represent a pixel value
            freeze_panes: E.g., "A2" to freeze header row

        Returns:
            Modified worksheet
        """

        # Step 1 - Fetch the file to update from SharePoint and load it as a workbook
        # This ensures we don't override any other sheets in the excel file
        binary_file = self.fetch_file_using_open_binary(excel_file_name, folder_name)
        if binary_file is None:
            raise FileNotFoundError(f"File '{excel_file_name}' not found in folder '{folder_name}'.")

        wb = load_workbook(BytesIO(binary_file))
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in '{excel_file_name}'")

        ws = wb[sheet_name]

        # Step 2 - Read data into DataFrame
        rows = list(ws.iter_rows(values_only=True))
        header, *data_rows = rows
        df = pd.DataFrame(data_rows, columns=header)

        # Step 3 – Prepare sorting logic
        # For each sorting instruction, we:
        # - Extract the column to sort by (using letter, index, or name)
        # - Convert the column values to the desired data type if specified (str, int, float, datetime)
        # - Track which columns to sort and in which order (ascending or descending)
        #
        # This ensures the DataFrame is sorted correctly, even when types like dates or numbers need conversion.
        if sorting_keys:
            sort_columns = []
            ascending_flags = []

            for item in sorting_keys:
                key = item.get("key")
                ascending = item.get("ascending", True)
                dtype = item.get("type")

                if isinstance(key, int):
                    col_name = header[key]

                elif isinstance(key, str) and key.isalpha():
                    col_name = header[ord(key.upper()) - ord("A")]

                else:
                    col_name = key

                sort_columns.append(col_name)
                ascending_flags.append(ascending)

                if dtype == "datetime":
                    df[col_name] = pd.to_datetime(df[col_name], dayfirst=True, errors="coerce")

                elif dtype == "int":
                    df[col_name] = pd.to_numeric(df[col_name], errors="coerce", downcast="integer")

                elif dtype == "float":
                    df[col_name] = pd.to_numeric(df[col_name], errors="coerce", downcast="float")

                elif dtype == "str":
                    df[col_name] = df[col_name].astype(str)

            # Step 4 – Sort
            df.sort_values(by=sort_columns, ascending=ascending_flags, inplace=True)

        # Step 5 - Overwrite worksheet
        ws.delete_rows(1, ws.max_row)

        ws.append(header)

        for _, row in df.iterrows():
            ws.append(list(row))

        # Step 6 – Adjust column widths and apply wrapping if needed
        #
        # If column_widths is "auto":
        # - Calculate the max content length in each column and set the column width accordingly (+2 for padding)
        #
        # If column_widths is a single int:
        # - Use it as a global max width across all columns
        # - If content fits, set width based on actual content length
        # - If content exceeds the max width clamp column width and enable wrap_text for that column's cells
        #
        # Then, for wrapped cells, auto-adjust the row height:
        # - Estimate how many lines the wrapped text would occupy and set row height accordingly to ensure all content is visible
        if column_widths in (None, "auto"):
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)

                ws.column_dimensions[col[0].column_letter].width = max_len + 2

        elif isinstance(column_widths, int):
            for col in ws.columns:
                col_letter = col[0].column_letter

                max_len = max(len(str(cell.value or "")) for cell in col)

                # If content fits, auto-size
                if max_len + 2 <= column_widths:
                    ws.column_dimensions[col_letter].width = max_len + 2

                # Else, cap width and enable wrap
                else:
                    ws.column_dimensions[col_letter].width = column_widths

                    for cell in col:
                        cell.alignment = Alignment(wrap_text=True)

            # Here we handle row height
            for row in ws.iter_rows():
                max_line_count = 1

                for cell in row:
                    if cell.value and cell.alignment and cell.alignment.wrap_text:
                        col_letter = cell.column_letter
                        col_width = ws.column_dimensions[col_letter].width or 10
                        chars_per_line = col_width * 1.2
                        lines = str(cell.value).split("\n")
                        line_count = sum(math.ceil(len(line) / chars_per_line) for line in lines)
                        max_line_count = max(max_line_count, line_count)

                ws.row_dimensions[row[0].row].height = max_line_count * 20

        else:
            raise ValueError(f"Column width provided with incorrect datatype - datatype int expected, instead column width is of datatype {type(column_widths)}")

        # Step 7 - Freeze panes if needed
        if freeze_panes:
            ws.freeze_panes = freeze_panes

        # Step 8 – Apply base formatting
        # For each cell in the worksheet:
        # - Apply font styling based on either a custom `font_config` (row-specific) or default to bold/italic based on row number (e.g., header rows)
        # - Set horizontal and vertical alignment for consistent layout
        # - Disable text wrapping by default (wrapping will be handled later if needed)
        #
        # This ensures a clean, uniform look across the sheet while allowing for custom styling where defined.
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                if font_config and row_idx in font_config:
                    config = font_config[row_idx]

                    cell.font = Font(
                        name=config.get("name", "Calibri"),
                        size=config.get("size", 11),
                        bold=config.get("bold", False),
                        italic=config.get("italic", False),
                    )

                else:
                    cell.font = Font(
                        bold=row_idx in bold_rows if bold_rows else False,
                        italic=row_idx in italic_rows if italic_rows else False,
                    )

                cell.alignment = Alignment(
                    horizontal=align_horizontal,
                    vertical=align_vertical,
                    wrap_text=cell.alignment.wrap_text
                )

        # Step 9 - Save and re-upload
        temp_stream = BytesIO()

        wb.save(temp_stream)

        temp_stream.seek(0)

        self.upload_file_from_bytes(temp_stream.getvalue(), excel_file_name, folder_name)
