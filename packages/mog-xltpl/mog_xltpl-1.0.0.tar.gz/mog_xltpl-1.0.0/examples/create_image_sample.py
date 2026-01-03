from pathlib import Path
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

root = Path('examples')
root.mkdir(exist_ok=True)

thin = Side(style='thin', color='BBBBBB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Step 1: Create template with Jinja2 filter syntax
wb = Workbook()
ws = wb.active
ws.title = 'template'

ws.merge_cells('A1:D1')
ws['A1'] = '画像埋め込みサンプル（フィルター版）'
ws['A1'].font = Font(size=15, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

ws['A3'] = '氏名'
ws['B3'] = '{{ name }}'
ws['A4'] = '日付'
ws['B4'] = '{{ issue_date }}'
ws['A5'] = 'メモ'
ws['B5'] = '{{ note | default("-") }}'

ws['A7'] = '写真'

# Template image должна быть на той же ячейке, что и фильтр
# B7 будет содержать как фильтр, так и изображение
# Add a placeholder image to B7 - this is required for xltpl to work!
# IMPORTANT: Use two-cell anchor for proper positioning
img_path = root / 'images' / '0.jpg'
if img_path.exists():
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    img = XLImage(str(img_path))
    img.width = 120
    img.height = 140
    # Create two-cell anchor explicitly for B7
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=6)  # B7 (0-indexed)
    anchor.to = AnchorMarker(col=2, row=7)
    img.anchor = anchor
    ws.add_image(img)

# Put the filter in the SAME cell as the image
ws['B7'] = '{{ avatar_path | img(120, 140) }}'
ws['A7'].alignment = Alignment(vertical='top')

for col in 'ABCD':
    ws.column_dimensions[col].width = 18
for row in range(3, 16):
    for col in 'AB':
        ws[f'{col}{row}'].border = border

# 備考
ws.merge_cells('A17:D19')
ws['A17'] = '{{ remarks | default("{{ avatar_path | img(120, 140) }} でフィルター経由で画像を埋め込めます") }}'
ws['A17'].alignment = Alignment(vertical='top', wrap_text=True)
ws['A17'].border = border

# Save template
img_tpl = root / 'sample_image.xlsx'
wb.save(img_tpl)
print('template created:', img_tpl)

# Step 2: Render with data and add image directly after
def create_sample_with_image():
    # Load template fresh
    wb = load_workbook(img_tpl)
    ws = wb.active
    
    # Set data
    ws['B3'] = '山田太郎'
    ws['B4'] = date.today().isoformat()
    ws['B5'] = '画像埋め込みの例'
    ws['A17'] = 'フィルター {{ avatar_path | img(width=120, height=140) }} で画像サイズを指定できます。'
    
    # Add image to B7 position
    img_path = root / 'images' / '0.jpg'
    if img_path.exists():
        img = XLImage(str(img_path))
        img.width = 120
        img.height = 140
        ws.add_image(img, 'B7')
    
    img_out = root / 'sample_image_result.xlsx'
    wb.save(img_out)
    print('rendered sample with image:', img_out)

create_sample_with_image()

