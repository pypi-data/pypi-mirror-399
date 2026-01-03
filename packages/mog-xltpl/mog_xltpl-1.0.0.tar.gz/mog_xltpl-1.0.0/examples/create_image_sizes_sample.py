from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

root = Path('examples')
root.mkdir(exist_ok=True)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = 'template'

# Title
ws.merge_cells('A1:E1')
ws['A1'] = '画像サイズ比較テンプレート'
ws['A1'].font = Font(size=14, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')

# Small image
ws['A3'] = '小（80x100）'
ws['A3'].font = Font(bold=True)
ws['B3'] = '{{ image_path | img(80, 100) }}'
for row in range(3, 15):
    ws[f'B{row}'].border = border

# Medium image
ws['A17'] = '中（120x150）'
ws['A17'].font = Font(bold=True)
ws['B17'] = '{{ image_path | img(120, 150) }}'
for row in range(17, 32):
    ws[f'B{row}'].border = border

# Large image
ws['A34'] = '大（180x220）'
ws['A34'].font = Font(bold=True)
ws['B34'] = '{{ image_path | img(180, 220) }}'
for row in range(34, 52):
    ws[f'B{row}'].border = border

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 25

tpl = root / 'sample_image_sizes.xlsx'
wb.save(tpl)
print('Created:', tpl)
