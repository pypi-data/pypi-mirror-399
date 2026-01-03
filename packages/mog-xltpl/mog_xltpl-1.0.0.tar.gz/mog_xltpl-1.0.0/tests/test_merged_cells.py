"""Tests for merged cells functionality"""
import pytest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from xltpl.writerx import BookWriter as BookWriterx


@pytest.fixture
def template_with_merged_cells(tmp_path):
    """Create a test template with merged cells"""
    tpl_path = tmp_path / "merged_cell_template.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "MergedTest"
    
    # Normal cell
    ws['A1'] = "Normal: {{normal_value}}"
    
    # Merged cell A3:C3 with template variable
    ws.merge_cells('A3:C3')
    ws['A3'] = "Merged: {{merged_value}}"
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Merged cell A5:B6 with template variable
    ws.merge_cells('A5:B6')
    ws['A5'] = "Title: {{title}}"
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Normal cell after merged
    ws['D5'] = "Status: {{status}}"
    
    # Horizontal merged cell E7:G7
    ws.merge_cells('E7:G7')
    ws['E7'] = "{{description}}"
    
    wb.save(str(tpl_path))
    return tpl_path


def test_merged_cell_substitution(template_with_merged_cells, tmp_path):
    """Test that merged cells are correctly substituted with values"""
    tpl = template_with_merged_cells
    out = tmp_path / "merged_cell_output.xlsx"
    
    # Variables to substitute
    vars_data = {
        'normal_value': 'Normal Text',
        'merged_value': 'Merged Cell Text',
        'title': 'Test Title',
        'status': 'Active',
        'description': 'This is a long description in a merged cell'
    }
    
    writer = BookWriterx(str(tpl))
    payloads = [vars_data]
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists(), "Output file should exist"
    
    # Load output and verify substitutions
    wb_out = load_workbook(str(out))
    ws = wb_out['MergedTest']
    
    # Check normal cell
    assert ws['A1'].value == 'Normal: Normal Text', f"Normal cell should be substituted, got '{ws['A1'].value}'"
    
    # Check merged cells - the value should be in the top-left cell of the merged range
    assert ws['A3'].value == 'Merged: Merged Cell Text', f"Merged cell A3:C3 should be substituted, got '{ws['A3'].value}'"
    assert ws['A5'].value == 'Title: Test Title', f"Merged cell A5:B6 should be substituted, got '{ws['A5'].value}'"
    assert ws['D5'].value == 'Status: Active', f"Normal cell after merged should be substituted, got '{ws['D5'].value}'"
    assert ws['E7'].value == 'This is a long description in a merged cell', f"Horizontal merged cell should be substituted, got '{ws['E7'].value}'"


def test_merged_cell_ranges_preserved(template_with_merged_cells, tmp_path):
    """Test that merged cell ranges are preserved after rendering"""
    tpl = template_with_merged_cells
    out = tmp_path / "merged_ranges_preserved.xlsx"
    
    # Load template to get original merged ranges
    wb_tpl = load_workbook(str(tpl))
    ws_tpl = wb_tpl['MergedTest']
    original_merged = sorted([str(merge) for merge in ws_tpl.merged_cells.ranges])
    
    # Render with variables
    vars_data = {
        'normal_value': 'Test',
        'merged_value': 'Test Merged',
        'title': 'Title',
        'status': 'OK',
        'description': 'Description'
    }
    
    writer = BookWriterx(str(tpl))
    payloads = [vars_data]
    writer.render_book(payloads)
    writer.save(str(out))
    
    # Load output and check merged ranges
    wb_out = load_workbook(str(out))
    ws_out = wb_out['MergedTest']
    output_merged = sorted([str(merge) for merge in ws_out.merged_cells.ranges])
    
    # Verify that merged ranges are preserved
    assert output_merged == original_merged, f"Merged ranges should be preserved.\nOriginal: {original_merged}\nOutput: {output_merged}"


def test_merged_cell_with_loops(tmp_path):
    """Test merged cells inside loops"""
    tpl_path = tmp_path / "merged_loop_template.xlsx"
    
    # Create template with loop and merged cells
    wb = Workbook()
    ws = wb.active
    ws.title = "LoopMerged"
    
    ws['A1'] = "{% for item in items %}"
    ws['A2'] = "{{loop.index}}"
    
    # Merged cell in loop
    ws.merge_cells('B2:D2')
    ws['B2'] = "{{item.name}}"
    
    ws['E2'] = "{{item.value}}"
    ws['A3'] = "{% endfor %}"
    
    wb.save(str(tpl_path))
    
    # Render
    out = tmp_path / "merged_loop_output.xlsx"
    vars_data = {
        'items': [
            {'name': 'Item A', 'value': 100},
            {'name': 'Item B', 'value': 200},
            {'name': 'Item C', 'value': 300}
        ]
    }
    
    writer = BookWriterx(str(tpl_path))
    payloads = [vars_data]
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify output
    wb_out = load_workbook(str(out))
    ws_out = wb_out['LoopMerged']
    
    # Check if loop created rows with values
    # Note: exact behavior depends on xltpl's loop implementation
    # At minimum, check that some values were rendered
    values_found = False
    for row in ws_out.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and 'Item' in str(cell.value):
                values_found = True
                break
    
    assert values_found, "Loop should have rendered items with merged cells"


def test_empty_merged_cell(tmp_path):
    """Test merged cells that are initially empty"""
    tpl_path = tmp_path / "empty_merged_template.xlsx"
    
    # Create template
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Header"
    
    # Empty merged cell
    ws.merge_cells('A2:C2')
    ws['A2'] = "{{value}}"
    
    wb.save(str(tpl_path))
    
    # Render
    out = tmp_path / "empty_merged_output.xlsx"
    vars_data = {'value': 'Filled Value'}
    
    writer = BookWriterx(str(tpl_path))
    payloads = [vars_data]
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify
    wb_out = load_workbook(str(out))
    ws_out = wb_out['Sheet']
    
    assert ws_out['A2'].value == 'Filled Value', f"Empty merged cell should be filled, got '{ws_out['A2'].value}'"
