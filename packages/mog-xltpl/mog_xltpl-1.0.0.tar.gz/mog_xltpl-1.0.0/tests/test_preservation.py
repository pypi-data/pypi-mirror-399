"""Tests for template preservation features"""
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from xltpl.writerx import BookWriter as BookWriterx


def test_all_sheets_rendered_with_empty_payloads(tmp_path):
    """Test that all sheets are rendered when payloads is None or empty"""
    tpl = Path(__file__).resolve().parent.parent / "examples" / "box.xlsx"
    out = tmp_path / "all_sheets_output.xlsx"
    
    writer = BookWriterx(str(tpl))
    # Render with empty payloads - should render all sheets
    writer.render_sheets(payloads=None)
    writer.save(str(out))
    
    assert out.exists()
    
    # Load template and output to compare
    wb_tpl = load_workbook(str(tpl))
    wb_out = load_workbook(str(out))
    
    # Get all template sheet names
    template_sheet_names = [sheet.title for sheet in wb_tpl.worksheets]
    output_sheet_names = [sheet.title for sheet in wb_out.worksheets]
    
    # All template sheets should be in output
    assert len(output_sheet_names) == len(template_sheet_names), \
        f"Expected {len(template_sheet_names)} sheets, got {len(output_sheet_names)}"
    
    assert set(output_sheet_names) == set(template_sheet_names), \
        f"Sheet names mismatch. Template: {template_sheet_names}, Output: {output_sheet_names}"


def test_all_sheets_rendered_with_empty_list(tmp_path):
    """Test that all sheets are rendered when payloads is an empty list"""
    tpl = Path(__file__).resolve().parent.parent / "examples" / "box.xlsx"
    out = tmp_path / "all_sheets_empty_list.xlsx"
    
    writer = BookWriterx(str(tpl))
    # Render with empty list
    writer.render_sheets(payloads=[])
    writer.save(str(out))
    
    assert out.exists()
    
    wb_tpl = load_workbook(str(tpl))
    wb_out = load_workbook(str(out))
    
    template_sheet_names = [sheet.title for sheet in wb_tpl.worksheets]
    output_sheet_names = [sheet.title for sheet in wb_out.worksheets]
    
    assert len(output_sheet_names) == len(template_sheet_names)


def test_template_images_preserved(tmp_path):
    """Test that template images are preserved in output"""
    # Create a test template with an image
    tpl_path = tmp_path / "template_with_image.xlsx"
    out_path = tmp_path / "output_with_image.xlsx"
    
    # Create template workbook with an image
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "Test"
    
    # Add an image to template
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if test_image_path.exists():
        img = OpenpyxlImage(str(test_image_path))
        img.anchor = "C3"
        ws.add_image(img)
        wb.save(str(tpl_path))
        
        # Render the template
        writer = BookWriterx(str(tpl_path))
        writer.render_sheets([{}])  # Empty payload
        writer.save(str(out_path))
        
        assert out_path.exists()
        
        # Check that image is preserved
        wb_out = load_workbook(str(out_path))
        ws_out = wb_out.active
        
        assert len(ws_out._images) > 0, "Template image should be preserved in output"
        assert ws_out._images[0].anchor is not None, "Image should have anchor position"


def test_template_images_with_dynamic_images(tmp_path):
    """Test that both template images and dynamic images appear in output"""
    tpl_path = tmp_path / "template_with_image2.xlsx"
    out_path = tmp_path / "output_with_both_images.xlsx"
    
    # Create template workbook with an image
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = "{{ 'Test' }}"
    ws['A2'] = "{{ image_path | img(100, 100) }}"  # Dynamic image
    
    # Add a template image
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if test_image_path.exists():
        img = OpenpyxlImage(str(test_image_path))
        img.anchor = "D4"  # Different position
        ws.add_image(img)
        wb.save(str(tpl_path))
        
        # Render with dynamic image
        writer = BookWriterx(str(tpl_path))
        payload = {"image_path": str(test_image_path)}
        writer.render_sheets([payload])
        writer.save(str(out_path))
        
        assert out_path.exists()
        
        # Check that both images are present
        wb_out = load_workbook(str(out_path))
        ws_out = wb_out.active
        
        # Should have both template image and dynamic image
        assert len(ws_out._images) >= 2, \
            f"Output should contain both template and dynamic images, got {len(ws_out._images)}"


def test_multiple_sheets_preserve_individual_images(tmp_path):
    """Test that each sheet preserves its own template images"""
    tpl_path = tmp_path / "multi_sheet_with_images.xlsx"
    out_path = tmp_path / "multi_sheet_output.xlsx"
    
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not test_image_path.exists():
        return  # Skip if no test image
    
    # Create template with multiple sheets, each with an image
    wb = Workbook()
    
    # Sheet 1
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1['A1'] = "Sheet 1"
    img1 = OpenpyxlImage(str(test_image_path))
    img1.anchor = "B2"
    ws1.add_image(img1)
    
    # Sheet 2
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = "Sheet 2"
    img2 = OpenpyxlImage(str(test_image_path))
    img2.anchor = "C3"
    ws2.add_image(img2)
    
    wb.save(str(tpl_path))
    
    # Render all sheets
    writer = BookWriterx(str(tpl_path))
    writer.render_sheets([])  # Empty list - should render all sheets
    writer.save(str(out_path))
    
    assert out_path.exists()
    
    # Check each sheet has its image
    wb_out = load_workbook(str(out_path))
    
    assert len(wb_out.worksheets) == 2, "Should have 2 sheets in output"
    
    sheet1_out = wb_out["Sheet1"]
    sheet2_out = wb_out["Sheet2"]
    
    assert len(sheet1_out._images) >= 1, "Sheet1 should have template image"
    assert len(sheet2_out._images) >= 1, "Sheet2 should have template image"


def test_partial_sheets_rendered_includes_unprocessed(tmp_path):
    """Test that unprocessed sheets are copied from template when only some sheets are rendered"""
    tpl_path = tmp_path / "multi_sheet_partial.xlsx"
    out_path = tmp_path / "multi_sheet_partial_output.xlsx"
    
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not test_image_path.exists():
        return  # Skip if no test image
    
    # Create template with 3 sheets
    wb = Workbook()
    
    # Sheet 1 - will be rendered
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1['A1'] = "{{ name }}"
    
    # Sheet 2 - will NOT be rendered
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = "Unprocessed Sheet"
    ws2['B1'] = "This should remain as-is"
    
    # Sheet 3 - will NOT be rendered
    ws3 = wb.create_sheet("Sheet3")
    ws3['A1'] = "Another Unprocessed"
    ws3['C1'] = "Also unchanged"
    
    wb.save(str(tpl_path))
    
    # Render ONLY Sheet1
    writer = BookWriterx(str(tpl_path))
    payload = {"name": "Test Name", "sheet_name": "Sheet1"}
    writer.render_sheets([payload])
    writer.save(str(out_path))
    
    assert out_path.exists()
    
    # Check that all 3 sheets are in output
    wb_out = load_workbook(str(out_path))
    
    assert len(wb_out.worksheets) == 3, \
        f"Should have 3 sheets in output (1 rendered + 2 unprocessed), got {len(wb_out.worksheets)}"
    
    output_sheet_names = [sheet.title for sheet in wb_out.worksheets]
    assert "Sheet1" in output_sheet_names, "Sheet1 should be in output"
    assert "Sheet2" in output_sheet_names, "Unprocessed Sheet2 should be in output"
    assert "Sheet3" in output_sheet_names, "Unprocessed Sheet3 should be in output"
    
    # Verify unprocessed sheets have original content
    sheet2_out = wb_out["Sheet2"]
    assert sheet2_out['A1'].value == "Unprocessed Sheet", "Sheet2 should have original content"
    assert sheet2_out['B1'].value == "This should remain as-is", "Sheet2 cell B1 should be unchanged"
    
    sheet3_out = wb_out["Sheet3"]
    assert sheet3_out['A1'].value == "Another Unprocessed", "Sheet3 should have original content"
    assert sheet3_out['C1'].value == "Also unchanged", "Sheet3 cell C1 should be unchanged"

