from pathlib import Path
import pytest
import time
import gc

from xltpl import cli


@pytest.fixture(autouse=True)
def cleanup_com_between_tests():
    """Ensure COM cleanup between tests to avoid RPC_E_DISCONNECTED errors."""
    yield
    # Wait for Excel COM to fully release after each test
    gc.collect()
    time.sleep(1.5)


def test_cli_simple_yaml(tmp_path):
    """Test CLI with template, output, and yaml vars file using a minimal template."""
    # Create a minimal template
    from openpyxl import Workbook
    
    template = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name: {{ name }}'
    ws['A2'] = 'Date: {{ date }}'
    wb.save(template)
    
    # Create vars.yaml
    yaml_file = tmp_path / "vars.yaml"
    yaml_file.write_text("""vars:
  name: "Test User"
  date: "2025-12-30"
""", encoding="utf-8")
    
    output = tmp_path / "output.xlsx"
    
    # Run CLI
    cli.main([str(template), str(output), str(yaml_file)])
    
    # Verify output exists
    assert output.exists()
    
    # Verify content
    from openpyxl import load_workbook
    wb_out = load_workbook(output)
    ws_out = wb_out.active
    assert ws_out['A1'].value == 'Name: Test User'
    assert ws_out['A2'].value == 'Date: 2025-12-30'


def test_cli_taskfile_placeholder_normalization(tmp_path):
        """Taskfile-style placeholders {{ .name }} are normalized before rendering."""
        from openpyxl import Workbook, load_workbook

        template = tmp_path / "tpl.xlsx"
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '{{ greeting }}'
        wb.save(template)

        yaml_file = tmp_path / "vars.yaml"
        yaml_file.write_text("""vars:
    name: "World"
    greeting: "Hello {{ .name }}"
""", encoding="utf-8")

        output = tmp_path / "out.xlsx"
        cli.main([str(template), str(output), str(yaml_file)])

        wb_out = load_workbook(output)
        ws_out = wb_out.active
        assert ws_out['A1'].value == 'Hello World'


def test_cli_with_highlight_output(tmp_path):
    """Ensure highlight output is produced when option is given."""
    from openpyxl import Workbook, load_workbook

    template = tmp_path / "tpl.xlsx"
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Name: {{ name }}'
    wb.save(template)

    yaml_file = tmp_path / "vars.yaml"
    yaml_file.write_text("""vars:
  name: "Highlight User"
""", encoding="utf-8")

    output = tmp_path / "out.xlsx"
    highlight_output = output.with_name(f"{output.stem}_highlight{output.suffix}")

    cli.main([
        str(template),
        str(output),
        str(yaml_file),
        "--highlight-output",
        "--highlight-color", "FFFF9999",
    ])

    assert output.exists()
    assert highlight_output.exists()

    wb_out = load_workbook(highlight_output)
    ws_out = wb_out.active
    assert ws_out['A1'].value == 'Name: Highlight User'


def test_cli_sheet_name_templating(tmp_path):
        """Sheet names can be templated (e.g., {{ sheet_name }}) and rendered from vars."""
        from openpyxl import Workbook, load_workbook

        template = tmp_path / "tpl.xlsx"
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws3 = wb.create_sheet("{{ sheet_name }}")
        ws3["A1"] = "Hello {{ sheet_name }}"  # content also templated
        wb.save(template)

        yaml_file = tmp_path / "vars.yaml"
        yaml_file.write_text("""vars:
    sheet_name: "hogehoge"
""", encoding="utf-8")

        output = tmp_path / "out.xlsx"
        cli.main([str(template), str(output), str(yaml_file)])

        wb_out = load_workbook(output)
        assert wb_out.sheetnames == ["Sheet1", "Sheet2", "hogehoge"]
        assert wb_out["hogehoge"]["A1"].value == "Hello hogehoge"


def test_cli_missing_template(tmp_path):
    """Test CLI with missing template file."""
    yaml_file = tmp_path / "vars.yaml"
    yaml_file.write_text("vars:\n  name: test\n", encoding="utf-8")
    
    with pytest.raises(SystemExit) as exc_info:
        cli.main(["nonexistent.xlsx", "output.xlsx", str(yaml_file)])
    
    assert "Template file not found" in str(exc_info.value)


def test_cli_missing_yaml(tmp_path):
    """Test CLI with missing YAML file."""
    examples_dir = Path(__file__).resolve().parent.parent / "examples"
    template = examples_dir / "sample_invoice.xlsx"
    
    with pytest.raises(SystemExit) as exc_info:
        cli.main([str(template), "output.xlsx", "nonexistent.yaml"])
    
    assert "YAML file not found" in str(exc_info.value)


def test_cli_yaml_without_vars(tmp_path):
    """Test CLI with YAML file that doesn't contain vars."""
    examples_dir = Path(__file__).resolve().parent.parent / "examples"
    template = examples_dir / "sample_invoice.xlsx"
    
    yaml_file = tmp_path / "bad.yaml"
    yaml_file.write_text("other: value\n", encoding="utf-8")
    
    output = tmp_path / "output.xlsx"
    
    with pytest.raises(SystemExit) as exc_info:
        cli.main([str(template), str(output), str(yaml_file)])
    
    assert "must contain 'vars' section" in str(exc_info.value)
