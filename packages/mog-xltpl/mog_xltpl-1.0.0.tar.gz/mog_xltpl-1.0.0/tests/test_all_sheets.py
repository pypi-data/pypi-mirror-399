"""Tests for rendering all sheets with same variables and preserving images"""
import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from xltpl.writerx import BookWriter as BookWriterx
import pytest


@pytest.fixture
def template_with_multiple_sheets_and_images(tmp_path):
    """Create a test template with multiple sheets and images"""
    tpl_path = tmp_path / "multi_sheet_template.xlsx"
    
    # Create workbook with 2 sheets
    wb = Workbook()
    
    # First sheet
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1['A1'] = "Name: {{name}}"
    ws1['A2'] = "Date: {{date}}"
    
    # Add image to first sheet
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not test_image_path.exists():
        pytest.skip("Test image missing: examples/images/0.jpg")

    img1 = OpenpyxlImage(str(test_image_path))
    img1.anchor = "C3"
    ws1.add_image(img1)
    
    # Second sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2['B1'] = "Value: {{value}}"
    ws2['B2'] = "Total: {{total}}"
    
    # Add image to second sheet
    img2 = OpenpyxlImage(str(test_image_path))
    img2.anchor = "D2"
    ws2.add_image(img2)
    
    wb.save(str(tpl_path))
    return tpl_path


def test_all_sheets_substituted_with_same_vars(template_with_multiple_sheets_and_images, tmp_path):
    """Test that all sheets receive the same variable substitutions"""
    tpl = template_with_multiple_sheets_and_images
    out = tmp_path / "all_sheets_substituted.xlsx"
    
    # Variables to substitute
    vars_data = {
        'name': 'Test User',
        'date': '2025-12-31',
        'value': '12345',
        'total': '99999'
    }
    
    writer = BookWriterx(str(tpl))
    
    # Create payload for each sheet with same vars
    payloads = []
    for sheet_state in writer.sheet_resource_map.sheet_resources:
        payload = dict(vars_data)
        payload['tpl_index'] = sheet_state.index
        payload['sheet_name'] = sheet_state.name
        payloads.append(payload)
    
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists()
    
    # Load output and verify substitutions in all sheets
    wb_out = load_workbook(str(out))
    
    # Check Sheet1
    ws1 = wb_out['Sheet1']
    assert ws1['A1'].value == 'Name: Test User', f"Expected 'Name: Test User', got '{ws1['A1'].value}'"
    assert ws1['A2'].value == 'Date: 2025-12-31', f"Expected 'Date: 2025-12-31', got '{ws1['A2'].value}'"
    
    # Check Sheet2
    ws2 = wb_out['Sheet2']
    assert ws2['B1'].value == 'Value: 12345', f"Expected 'Value: 12345', got '{ws2['B1'].value}'"
    assert ws2['B2'].value == 'Total: 99999', f"Expected 'Total: 99999', got '{ws2['B2'].value}'"


def test_template_images_preserved_in_all_sheets(template_with_multiple_sheets_and_images, tmp_path):
    """Test that template images are preserved in all sheets after rendering"""
    tpl = template_with_multiple_sheets_and_images
    out = tmp_path / "images_preserved.xlsx"
    
    # Load template to count original images
    wb_tpl = load_workbook(str(tpl))
    sheet1_images = len(wb_tpl['Sheet1']._images)
    sheet2_images = len(wb_tpl['Sheet2']._images)
    
    # Render with variables
    vars_data = {
        'name': 'Test',
        'date': '2025-12-31',
        'value': '123',
        'total': '456'
    }
    
    writer = BookWriterx(str(tpl))
    
    # Create payload for each sheet
    payloads = []
    for sheet_state in writer.sheet_resource_map.sheet_resources:
        payload = dict(vars_data)
        payload['tpl_index'] = sheet_state.index
        payload['sheet_name'] = sheet_state.name
        payloads.append(payload)
    
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists()
    
    # Load output and verify images are preserved
    wb_out = load_workbook(str(out))
    
    # Check Sheet1 images
    ws1_out = wb_out['Sheet1']
    assert len(ws1_out._images) == sheet1_images, \
        f"Sheet1: Expected {sheet1_images} images, got {len(ws1_out._images)}"
    
    # Check Sheet2 images
    ws2_out = wb_out['Sheet2']
    assert len(ws2_out._images) == sheet2_images, \
        f"Sheet2: Expected {sheet2_images} images, got {len(ws2_out._images)}"


def test_images_preserved_with_no_substitutions(tmp_path):
    """Test that images are preserved even when no substitutions occur"""
    tpl_path = tmp_path / "image_only_template.xlsx"
    
    # Create simple template with just an image
    wb = Workbook()
    ws = wb.active
    ws.title = "ImageSheet"
    ws['A1'] = "Static text"
    
    # Add image
    test_image_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not test_image_path.exists():
        pytest.skip("Test image missing: examples/images/0.jpg")

    img = OpenpyxlImage(str(test_image_path))
    img.anchor = "B2"
    ws.add_image(img)
    wb.save(str(tpl_path))
    
    out = tmp_path / "image_preserved_output.xlsx"
    
    # Render with empty payload
    writer = BookWriterx(str(tpl_path))
    payloads = []
    for sheet_state in writer.sheet_resource_map.sheet_resources:
        payload = {'tpl_index': sheet_state.index, 'sheet_name': sheet_state.name}
        payloads.append(payload)
    
    writer.render_book(payloads)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify image is preserved
    wb_out = load_workbook(str(out))
    ws_out = wb_out['ImageSheet']
    assert len(ws_out._images) == 1, \
        f"Expected 1 image, got {len(ws_out._images)}"
    assert ws_out['A1'].value == "Static text"


def test_static_image_template_multi_sheet_substitution(tmp_path):
    tpl = Path(__file__).resolve().parent.parent / "examples" / "static_image_2_temp.xlsx"
    assert tpl.exists(), "Template missing: examples/static_image_2_temp.xlsx"

    out = tmp_path / "static_image_2_rendered.xlsx"

    writer = BookWriterx(str(tpl))
    vars_data = {
        "MODEL": "Car-01",
        "STRING": "Rendered text",
    }

    payloads = []
    for sheet_state in writer.sheet_resource_map.sheet_resources:
        payload = dict(vars_data)
        payload["tpl_index"] = sheet_state.index
        payload["sheet_name"] = sheet_state.name
        payloads.append(payload)

    writer.render_book(payloads)
    writer.save(str(out))

    artifact_dir = Path(__file__).resolve().parent / "artifacts"
    artifact_dir.mkdir(exist_ok=True)
    shutil.copy(out, artifact_dir / "static_image_2_rendered.xlsx")

    assert out.exists()

    wb_out = load_workbook(str(out))
    for sheet_name in wb_out.sheetnames:
        ws = wb_out[sheet_name]
        assert ws["A1"].value == "Car-01"
        assert ws["C1"].value == "Rendered text"
        assert len(ws._images) == 1
