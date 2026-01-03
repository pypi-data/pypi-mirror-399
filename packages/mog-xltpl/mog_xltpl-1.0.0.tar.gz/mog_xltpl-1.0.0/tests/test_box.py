from pathlib import Path
from datetime import datetime
import sys

from xltpl.writer import BookWriter
from xltpl.writerx import BookWriter as BookWriterx

# ensure examples directory is importable
EXAMPLES_DIR = Path(__file__).resolve().parent.parent / "examples"
if str(EXAMPLES_DIR) not in sys.path:
    sys.path.insert(0, str(EXAMPLES_DIR))
from box import get_items


def _build_payload():
    now = datetime.now()
    items = get_items()
    person = {
        "address": "No Where",
        "name": "No Name",
        "fm": 333,
        "date": now,
        "rows": items,
        "items": items,
        "sheet_name": "box",
    }
    return person


def _render_box(writer):
    person = _build_payload()
    top_left = None
    # 1周だけ描画して軽量化
    person["tpl_name"] = "top"
    top_box = writer.render_sheet(person, top_left)

    person["tpl_name"] = "left"
    left_box = writer.render_sheet(person, (top_box.bottom, top_box.left))

    person["tpl_name"] = "list1"
    middle_box = writer.render_sheet(person, (left_box.top, left_box.right))

    person["tpl_name"] = "right"
    writer.render_sheet(person, (middle_box.top, middle_box.right))


def test_box_xls(tmp_path):
    tpl = Path(__file__).resolve().parent.parent / "examples" / "box.xls"
    out = tmp_path / "box_out.xls"
    writer = BookWriter(str(tpl))
    _render_box(writer)
    writer.save(str(out))
    assert out.exists()


def test_box_xlsx(tmp_path):
    tpl = Path(__file__).resolve().parent.parent / "examples" / "box.xlsx"
    out = tmp_path / "box_out.xlsx"
    writer = BookWriterx(str(tpl))
    _render_box(writer)
    writer.save(str(out))
    assert out.exists()


# ========== Image filter tests ==========

def test_img_filter_with_size(tmp_path):
    """Test img filter with width and height parameters"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    
    # Create template with placeholder image
    tpl = tmp_path / "tpl_with_size.xlsx"
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Image with Size'
    ws['B2'] = '{{ img_path | img(150, 120) }}'
    
    # Add placeholder image
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    placeholder_img = XLImage(str(img_path))
    placeholder_img.width = 200  # Large placeholder to see difference
    placeholder_img.height = 250
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=1)  # B2
    anchor.to = AnchorMarker(col=2, row=3)
    placeholder_img.anchor = anchor
    ws.add_image(placeholder_img)
    
    wb.save(tpl)
    
    out = tmp_path / "test_img_filter_with_size.xlsx"
    
    payload = [{
        'img_path': str(img_path)
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    # Verify output file
    assert out.exists()
    
    # Check that image exists in output
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    assert len(ws_out._images) > 0, "Output should contain at least one image"
    
    # Verify the template had placeholder
    wb_tpl = load_workbook(tpl)
    ws_tpl = wb_tpl.active
    assert len(ws_tpl._images) > 0, "Template should have placeholder image"


def test_img_filter_without_size(tmp_path):
    """Test img filter without size parameters (auto-size)"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    
    tpl = tmp_path / "tpl_auto_size.xlsx"
    
    # Create simple template with placeholder
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Image Auto Size'
    ws['B2'] = '{{ img_path | img }}'
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    # Add placeholder image
    placeholder_img = XLImage(str(img_path))
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=1)  # B2
    anchor.to = AnchorMarker(col=2, row=3)
    placeholder_img.anchor = anchor
    ws.add_image(placeholder_img)
    
    wb.save(tpl)
    
    out = tmp_path / "result_auto_size.xlsx"
    
    payload = [{
        'img_path': str(img_path)
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify output file has image
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    assert len(ws_out._images) > 0, "Output should contain image"


def test_img_filter_with_keyword_args(tmp_path):
    """Test img filter with keyword arguments"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    
    tpl = tmp_path / "tpl_keyword_args.xlsx"
    
    # Create simple template with placeholder
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Image with Keywords'
    ws['B2'] = '{{ img_path | img(width=150, height=200) }}'
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    # Add placeholder image
    placeholder_img = XLImage(str(img_path))
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=1)  # B2
    anchor.to = AnchorMarker(col=2, row=3)
    placeholder_img.anchor = anchor
    ws.add_image(placeholder_img)
    
    wb.save(tpl)
    
    out = tmp_path / "result_keyword_args.xlsx"
    
    payload = [{
        'img_path': str(img_path)
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify output file has image
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    assert len(ws_out._images) > 0, "Output should contain image"


def test_sample_invoice_render(tmp_path):
    """Test sample invoice rendering"""
    tpl = Path(__file__).resolve().parent.parent / "examples" / "sample_invoice.xlsx"
    out = tmp_path / "sample_invoice_result.xlsx"
    
    if not tpl.exists():
        return
    
    payload = [{
        'customer': {
            'name': 'テスト太郎',
            'company': 'テスト会社'
        },
        'issue_date': '2025-12-29',
        'note': 'テスト用',
        'items': [
            {'name': 'テスト商品A', 'qty': 2, 'price': 1000, 'tax_rate': 0.1,
             'subtotal': 2000, 'tax_amount': 200, 'total': 2200},
            {'name': 'テスト商品B', 'qty': 1, 'price': 5000, 'tax_rate': 0.1,
             'subtotal': 5000, 'tax_amount': 500, 'total': 5500},
        ],
        'remarks': 'テスト備考'
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    assert out.exists()


def test_sample_image_render(tmp_path):
    """Test sample image rendering"""
    tpl = Path(__file__).resolve().parent.parent / "examples" / "sample_image_template.xlsx"
    out = tmp_path / "sample_image_test_result.xlsx"
    
    if not tpl.exists():
        return
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    payload = [{
        'name': 'テスト太郎',
        'issue_date': '2025-12-29',
        'note': 'テスト',
        'avatar_path': str(img_path),
        'remarks': 'テスト備考'
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    assert out.exists()


def test_dual_output_with_highlight(tmp_path):
    """Test dual output with highlight color"""
    tpl = Path(__file__).resolve().parent.parent / "examples" / "sample_invoice.xlsx"
    out_normal = tmp_path / "invoice_normal.xlsx"
    
    if not tpl.exists():
        return
    
    payload = [{
        'customer': {
            'name': 'テスト太郎',
            'company': 'テスト会社'
        },
        'issue_date': '2025-12-29',
        'note': 'テスト',
        'items': [
            {'name': 'テスト商品', 'qty': 1, 'price': 1000, 'tax_rate': 0.1,
             'subtotal': 1000, 'tax_amount': 100, 'total': 1100},
        ],
        'remarks': 'テスト'
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_and_save(
        payload,
        str(out_normal),
        dual_output=True,
        highlight_color='FFFF9999'
    )
    # Normal file should exist
    assert out_normal.exists()
    # Highlight file (with _highlight suffix)
    out_highlight = tmp_path / "invoice_normal_highlight.xlsx"
    assert out_highlight.exists()


def test_img_filter_multiple_rows(tmp_path):
    """Test img filter with multiple rows - each row gets its own image"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    
    tpl = tmp_path / "tpl_multi_rows.xlsx"
    wb = Workbook()
    ws = wb.active
    
    # Create template with two rows of data
    ws['A1'] = 'Name'
    ws['B1'] = 'Avatar'
    ws['A2'] = '{{ name }}'
    ws['B2'] = '{{ avatar | img(100, 100) }}'
    ws['A3'] = '{{ name }}'
    ws['B3'] = '{{ avatar | img(100, 100) }}'
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    # Add placeholder image to B2
    placeholder_img = XLImage(str(img_path))
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=1)  # B2
    anchor.to = AnchorMarker(col=2, row=2)
    placeholder_img.anchor = anchor
    ws.add_image(placeholder_img)
    
    # Add placeholder image to B3
    placeholder_img2 = XLImage(str(img_path))
    anchor2 = TwoCellAnchor()
    anchor2._from = AnchorMarker(col=1, row=2)  # B3
    anchor2.to = AnchorMarker(col=2, row=3)
    placeholder_img2.anchor = anchor2
    ws.add_image(placeholder_img2)
    
    wb.save(tpl)
    
    out = tmp_path / "result_multi_rows.xlsx"
    
    payload = [
        {'name': 'Person 1', 'avatar': str(img_path)},
        {'name': 'Person 2', 'avatar': str(img_path)},
    ]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify output has images for both rows
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    # Should have at least 2 images (one for each row)
    assert len(ws_out._images) >= 2, f"Output should contain at least 2 images, got {len(ws_out._images)}"


def test_img_filter_image_path_verification(tmp_path):
    """Test that img filter correctly processes the specified image path"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    
    tpl = tmp_path / "tpl_path_verify.xlsx"
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Avatar'
    ws['B2'] = '{{ img_path | img(120, 140) }}'
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    # Add placeholder image
    placeholder_img = XLImage(str(img_path))
    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=1, row=1)  # B2
    anchor.to = AnchorMarker(col=2, row=2)
    placeholder_img.anchor = anchor
    ws.add_image(placeholder_img)
    
    wb.save(tpl)
    
    out = tmp_path / "result_path_verify.xlsx"
    
    payload = [{
        'img_path': str(img_path)
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify image exists in output
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    assert len(ws_out._images) > 0, "Output should contain image"
    
    # Verify the image has expected ref/path
    img = ws_out._images[0]
    assert img.ref is not None, "Image should have ref"
    # The ref should point to the media location in the workbook
    assert img.path.startswith('/xl/media/'), f"Image path should be in /xl/media/, got {img.path}"


def test_img_filter_without_placeholder(tmp_path):
    """Test that img tag works without placeholder images in template.
    
    Uses {% img %} tag syntax.
    """
    from openpyxl import Workbook, load_workbook
    
    tpl = tmp_path / "tpl_no_placeholder.xlsx"
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Avatar'
    ws['B2'] = '{% img avatar_path, 120, 140 %}'  # Use {% img %} tag with width and height
    
    # NO placeholder image - just the template tag
    wb.save(tpl)
    
    out = tmp_path / "result_no_placeholder.xlsx"
    
    img_path = Path(__file__).resolve().parent.parent / "examples" / "images" / "0.jpg"
    if not img_path.exists():
        return
    
    payload = [{
        'avatar_path': str(img_path)
    }]
    
    writer = BookWriterx(str(tpl))
    writer.render_book(payload)
    writer.save(str(out))
    
    assert out.exists()
    
    # Verify image exists in output despite no placeholder
    wb_out = load_workbook(out)
    ws_out = wb_out.active
    assert len(ws_out._images) > 0, "Output should contain image even without placeholder"
    
    # Verify the image has expected dimensions
    img = ws_out._images[0]
    # Note: openpyxl may adjust dimensions, so we just check they're set
    assert img.width > 0, "Image width should be set"
    assert img.height > 0, "Image height should be set"


