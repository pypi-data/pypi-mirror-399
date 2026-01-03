# -*- coding: utf-8 -*-

from .patchx import *
from openpyxl import load_workbook
from openpyxl.utils.cell import absolute_coordinate
import zipfile
from openpyxl.cell.rich_text import CellRichText, TextBlock
from .basex import SheetBase, BookBase
from .writermixin import SheetMixin, BookMixin, Box
from .utils import tag_test, parse_cell_tag
from .xlnode import Tree, Row, Cell, EmptyCell, Node, create_cell
from .jinja import JinjaEnvx
from .nodemap import NodeMap
from .sheetresource import SheetResourceMap
from .richtexthandler import rich_handlerx
from .mergerx import Merger
from .config import config
from .celltag import CellTag
from .image import img_cache
from pathlib import Path
import tempfile
import os

class SheetWriter(SheetBase, SheetMixin):

    def __init__(self, bookwriter, sheet_resource, sheet_name):
        self.bookwriter = bookwriter
        self.workbook = bookwriter.workbook
        self.merger = sheet_resource.merger
        self.rdsheet = sheet_resource.rdsheet
        # Render directly onto the template sheet (no recreation) to preserve images/relations
        self.wtsheet = self.rdsheet
        # Apply requested sheet name (already validated in CLI)
        if sheet_name and sheet_name != self.wtsheet.title:
            self.wtsheet.title = sheet_name
        self.highlight = getattr(bookwriter, 'highlight', False)
        self.highlight_color = getattr(bookwriter, 'highlight_color', None)
        self.log_path = getattr(bookwriter, 'log_path', None)
        self.copy_sheet_settings()
        self.wtrows = set()
        self.wtcols = set()
        self.box = Box(0, 0)
        # Track rendered cell values for direct XML serialization
        self.rendered_cells = {}  # (row, col) -> value


class BookWriter(BookBase, BookMixin):
    sheet_writer_cls = SheetWriter

    def __init__(self, fname, debug=False, highlight=False, highlight_color=None, log_path=None):
        config.debug = debug
        self.debug = debug
        self.highlight = highlight
        self.highlight_color = highlight_color
        self.log_path = log_path
        self.template_path = Path(fname)
        self.load(fname)

    def load(self, fname):
        # IMPORTANT: Load template with openpyxl only for cell/structure analysis
        # We will use COM to actually open/save the template to preserve all Excel features
        with zipfile.ZipFile(fname) as zf:
            has_vba = any(name.endswith('vbaProject.bin') for name in zf.namelist())
        
        # Load template with openpyxl for analysis only (not for saving)
        self.workbook = load_workbook(fname, rich_text=True, keep_vba=has_vba)
        self.font_map = {}
        self.node_map = NodeMap()
        self.jinja_env = JinjaEnvx(self.node_map)
        from .mergerx import Merger
        self.merger_cls = Merger
        self.sheet_writer_map = {}
        # Track template sheets; we keep them in place
        self.template_sheets = {}  # name -> sheet mapping
        self.processed_sheet_names = set()
        
        # Track original template image count per sheet for image merging
        self.template_image_counts = {}
        for sheet in self.workbook.worksheets:
            self.template_image_counts[sheet.title] = len(sheet._images)
        
        self.sheet_resource_map = SheetResourceMap(self, self.jinja_env)
        for index, rdsheet in enumerate(self.workbook.worksheets):
            self.template_sheets[rdsheet.title] = rdsheet
            self.sheet_resource_map.add(rdsheet, rdsheet.title, index)

    def build(self, sheet, index, merger):
        tree = Tree(index, self.node_map)
        max_row = max(sheet.max_row, merger.image_merger.max_row)
        max_col = max(sheet.max_column, merger.image_merger.max_col)
        
        # Get merged cell ranges to identify member cells (non-top-left cells of merged ranges)
        merged_ranges = list(sheet.merged_cells.ranges)
        merged_member_cells = set()
        for merged_range in merged_ranges:
            # Add all cells except the top-left one
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if not (row == merged_range.min_row and col == merged_range.min_col):
                        merged_member_cells.add((row, col))
        
        for rowx in range(1, max_row + 1):
            row_node = Row(rowx)
            tree.add_child(row_node)
            for colx in range(1, max_col + 1):
                # Skip member cells of merged ranges (they should be EmptyCell)
                if (rowx, colx) in merged_member_cells:
                    cell_node = EmptyCell(rowx, colx)
                    tree.add_child(cell_node)
                    continue
                    
                sheet_cell = sheet._cells.get((rowx, colx))
                if not sheet_cell:
                    cell_node = EmptyCell(rowx, colx)
                    tree.add_child(cell_node)
                    continue
                cell_tag_map = None
                if sheet_cell.comment:
                    comment = sheet_cell.comment.text
                    if tag_test(comment):
                        _,cell_tag_map = parse_cell_tag(comment)
                value = sheet_cell._value
                data_type = sheet_cell.data_type
                if data_type == 's':
                    rich_text = None
                    if isinstance(value, CellRichText):
                        #print(value)
                        rich_text = value
                        value = str(rich_text)
                    if not tag_test(value):
                        if rich_text:
                            cell_node = Cell(sheet_cell, rowx, colx, rich_text, data_type)
                        else:
                            cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                    else:
                        font = self.get_font(sheet_cell._style.fontId)
                        cell_node = create_cell(sheet_cell, rowx, colx, value, rich_text, data_type, font, rich_handlerx)
                else:
                    cell_node = Cell(sheet_cell, rowx, colx, value, data_type)
                if cell_tag_map:
                    cell_tag = CellTag(cell_tag_map)
                    cell_node.extend_cell_tag(cell_tag)
                    if colx==1:
                        row_node.cell_tag = cell_tag
                tree.add_child(cell_node)
        tree.add_child(Node())#
        return tree

    def cleanup_defined_names(self):
        self.workbook.custom_doc_props = ()
        # Custom Document Properties cause invalid file error
        sheet_cnt = len(self.workbook.worksheets)
        valid_names = {}
        for k, v in self.workbook.defined_names.items():
            if v.localSheetId:
                if int(v.localSheetId) < sheet_cnt:
                    valid_names[k] = v
            else:
                valid_names[k] = v
        self.workbook.defined_names = valid_names

    def _normalize_print_areas(self):
        # openpyxl can drop sheet qualifiers on print areas when rendering in-place
        for ws in self.workbook.worksheets:
            area = ws.print_area
            if not area:
                continue
            if "!" in area and "$" in area:
                # Already qualified with sheet name and absolute refs
                continue
            ranges = []
            for rng in str(area).split(","):
                rng = rng.strip()
                if not rng:
                    continue
                if "!" in rng:
                    ranges.append(rng)
                    continue
                if ":" in rng:
                    start, end = rng.split(":", 1)
                    ranges.append(f"{absolute_coordinate(start)}:{absolute_coordinate(end)}")
                else:
                    ranges.append(absolute_coordinate(rng))
            fixed = ",".join(ranges)
            ws.print_area = f"'{ws.title}'!{fixed}"

    def save(self, fname):
        """
        Save using COM to preserve all Excel template features.
        openpyxl is used only for rendering logic, not for saving.
        """
        import os
        import tempfile
        import time
        from pathlib import Path
        
        output_ext = os.path.splitext(fname)[1].lower()

        try:
            import win32com.client
        except ImportError:
            raise RuntimeError(
                "pywin32 is required for saving. Please install: pip install pywin32\n"
                "This is necessary to preserve all Excel template features."
            )
        
        excel = None
        template_wb = None
        try:
            # Create fresh Excel instance with retry for COM cleanup timing
            max_retries = 5
            retry_delay = 1.0
            for attempt in range(max_retries):
                try:
                    excel = win32com.client.Dispatch("Excel.Application")
                    break
                except Exception as e:
                    if attempt < max_retries - 1:
                        import gc
                        gc.collect()
                        time.sleep(retry_delay)
                    else:
                        raise RuntimeError(f"Failed to create Excel instance after {max_retries} attempts: {e}")
            
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.AskToUpdateLinks = False
            excel.AlertBeforeOverwriting = False
            
            # Open the ORIGINAL template via COM (not openpyxl)
            # This preserves VBA, shapes, complex formatting, etc.
            template_wb = excel.Workbooks.Open(
                str(self.template_path.absolute()),
                UpdateLinks=0,
                ReadOnly=False,
            )
            
            # Apply rendered values from openpyxl workbook to COM workbook
            for sheet_idx in range(len(self.workbook.worksheets)):
                if sheet_idx >= template_wb.Worksheets.Count:
                    break
                
                src_sheet = self.workbook.worksheets[sheet_idx]  # openpyxl with rendered values
                dst_sheet = template_wb.Worksheets(sheet_idx + 1)  # COM original template
                
                # Apply sheet name from openpyxl (Jinja-rendered) to COM workbook
                if src_sheet.title != dst_sheet.Name:
                    try:
                        dst_sheet.Name = src_sheet.title
                    except Exception as e:
                        print(f"[WARNING] Could not rename sheet to '{src_sheet.title}': {e}")
                
                # Apply cell values
                for row in src_sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            try:
                                value = cell.value
                                if isinstance(value, (CellRichText, TextBlock)):
                                    value = str(value)
                                dst_sheet.Cells(cell.row, cell.column).Value = value
                            except Exception as e:
                                pass
                
                # Add dynamic images
                try:
                    template_image_count = self.template_image_counts.get(src_sheet.title, 0)
                    total_images = len(src_sheet._images)
                    
                    if total_images > template_image_count:
                        for img_idx in range(template_image_count, total_images):
                            img = src_sheet._images[img_idx]
                            if hasattr(img.anchor, '_from'):
                                from_marker = img.anchor._from
                                row_num = from_marker.row + 1
                                col_num = from_marker.col + 1
                                
                                # Create temp image file
                                temp_dir = tempfile.mkdtemp()
                                temp_img_path = Path(temp_dir) / f"img_{sheet_idx}_{from_marker.row}_{from_marker.col}.png"
                                
                                if hasattr(img, '_data'):
                                    with open(temp_img_path, 'wb') as f:
                                        f.write(img._data())
                                    
                                    try:
                                        target_cell = dst_sheet.Cells(row_num, col_num)
                                        dst_sheet.Shapes.AddPicture(
                                            Filename=str(temp_img_path.absolute()),
                                            LinkToFile=False,
                                            SaveWithDocument=True,
                                            Left=target_cell.Left,
                                            Top=target_cell.Top,
                                            Width=int(img.width) if img.width else -1,
                                            Height=int(img.height) if img.height else -1
                                        )
                                    except Exception as e:
                                        print(f"[ERROR] Failed to add image: {e}")
                                        import traceback
                                        traceback.print_exc()
                                    finally:
                                        try:
                                            os.remove(temp_img_path)
                                            os.rmdir(temp_dir)
                                        except:
                                            pass
                except Exception as e:
                    pass
            
            # Save via COM
            output_path_abs = os.path.abspath(fname)
            if os.path.exists(output_path_abs):
                os.remove(output_path_abs)
            
            format_map = {
                '.xlsm': 52,
                '.xlsx': 51,
                '.xltx': 54,
                '.xltm': 53,
                '.xls': 56,
            }
            file_format = format_map.get(output_ext, 51)
            
            template_wb.SaveAs(output_path_abs, FileFormat=file_format)
            
        except Exception as e:
            print(f"[ERROR] Save failed: {e}")
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"Failed to save workbook: {e}")
        
        finally:
            # Cleanup
            if template_wb:
                try:
                    template_wb.Close(SaveChanges=False)
                except:
                    pass
                time.sleep(0.3)
            
            if excel:
                try:
                    while excel.Workbooks.Count > 0:
                        excel.Workbooks(1).Close(SaveChanges=False)
                        time.sleep(0.2)
                except:
                    pass
                try:
                    excel.Quit()
                except:
                    pass
                time.sleep(1.0)
                excel = None
                template_wb = None
                
                import gc
                gc.collect()
                time.sleep(0.5)
        
        img_cache.clear()
        self.sheet_writer_map.clear()
        
        try:
            self.workbook.close()
        except:
            pass

