import copy
from openpyxl.cell.cell import MergedCell
import xlrd
import xlwt
from openpyxl.utils.datetime import to_excel
from openpyxl.cell.cell import NUMERIC_TYPES, TIME_TYPES, STRING_TYPES
from openpyxl.styles import PatternFill
import csv
from pathlib import Path
import csv
from pathlib import Path
BOOL_TYPE = bool

def get_type(value):
    if isinstance(value, NUMERIC_TYPES):
        dt = xlrd.XL_CELL_NUMBER
    elif isinstance(value, STRING_TYPES):
        dt = xlrd.XL_CELL_TEXT
    elif isinstance(value, TIME_TYPES):
        dt = xlrd.XL_CELL_DATE
        return to_excel(value), dt
    elif isinstance(value, BOOL_TYPE):
        dt = xlrd.XL_CELL_BOOLEAN
    else:
        return str(value), xlrd.XL_CELL_TEXT
    return value, dt

class Base(object):

    def __init__(self, sheet_writer, cell_node, value, data_type):
        self.sheet_writer = sheet_writer
        self.cell_node = cell_node
        self.value = value
        self.data_type = data_type

    @property
    def rdsheet(self):
        return self.sheet_writer.rdsheet

    @property
    def wtsheet(self):
        return self.sheet_writer.wtsheet

    @property
    def source_cell(self):
        return self.cell_node.sheet_cell

    @property
    def target_cell(self):
        return None

    @property
    def rdcolx(self):
        return self.cell_node.colx

    @property
    def rdrowx(self):
        return self.cell_node.rowx

    @property
    def wtcolx(self):
        return self.sheet_writer.box.right

    @property
    def wtrowx(self):
        return self.sheet_writer.box.bottom

    def apply_filters(self):
        if hasattr(self.cell_node, 'filters') and self.cell_node.filters:
            for (filter, args) in self.cell_node.filters:
                #print('args', args)
                filter(self, *args)
            self.cell_node.filters.clear()

    def _log_cell_write(self):
        """Log cell write to CSV if sheet_writer has log_path set"""
        log_path = getattr(self.sheet_writer, 'log_path', None)
        if not log_path:
            return
        # Excel column letter
        col_letter = self._col_to_letter(self.wtcolx)
        row_number = self.wtrowx + 1  # 1-based for display
        address = f"{col_letter}{row_number}"
        value = self.value if self.value is not None else self.source_cell.value if hasattr(self.source_cell, 'value') else ''
        sheet_name = getattr(self.wtsheet, 'name', None) or getattr(self.wtsheet, 'title', 'Sheet')
        # Append to CSV
        log_file = Path(log_path)
        file_exists = log_file.exists()
        with log_file.open('a', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(['sheet', 'address', 'row', 'col', 'value'])
            writer.writerow([sheet_name, address, row_number, self.wtcolx, str(value)])

    @staticmethod
    def _col_to_letter(col_idx):
        """Convert 0-based column index to Excel letter (A, B, ..., Z, AA, ...)"""
        result = []
        col_idx += 1  # 1-based
        while col_idx > 0:
            col_idx -= 1
            result.append(chr(col_idx % 26 + ord('A')))
            col_idx //= 26
        return ''.join(reversed(result))

    # highlight helper for both backends
    def _maybe_highlight_xls(self, style):
        if not getattr(self.sheet_writer, "highlight", False):
            return style
        # xlwt color name map; fallback to red if unknown
        color_name = getattr(self.sheet_writer, "highlight_color", "red") or "red"
        colour_map = xlwt.Style.colour_map
        colour_idx = colour_map.get(color_name, colour_map.get("red", 2))
        pattern = xlwt.Pattern()
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = colour_idx
        style.pattern = pattern
        return style

    def _maybe_highlight_xlsx(self, target_cell):
        if not getattr(self.sheet_writer, "highlight", False):
            return target_cell
        color_hex = getattr(self.sheet_writer, "highlight_color", None) or "FFFF9999"
        target_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        return target_cell


class CellContextX(Base):

    def __init__(self, sheet_writer, cell_node, value, data_type):
        super().__init__(sheet_writer, cell_node, value, data_type)
        self._target_cell = None

    @property
    def target_cell(self):
        if self._target_cell:
            return self._target_cell
        source = self.source_cell
        wtcolx = self.wtcolx
        wtrowx = self.wtrowx
        value = self.value
        data_type = self.data_type
        target = self.wtsheet.cell(column=wtcolx, row=wtrowx)

        # If target is a merged placeholder cell, redirect writes to the top-left real cell
        if isinstance(target, MergedCell):
            for merged_range in self.wtsheet.merged_cells.ranges:
                if merged_range.min_row <= wtrowx <= merged_range.max_row \
                        and merged_range.min_col <= wtcolx <= merged_range.max_col:
                    target = self.wtsheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    break

        if value is None:
            target._value = source._value
            target.data_type = source.data_type
        elif isinstance(value, (str, bytes)) and value.startswith('='):
            target.value = value
        elif data_type:
            target._value = value
            target.data_type = data_type
        else:
            target.value = value
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.hyperlink:
            target.hyperlink = copy.copy(source.hyperlink)
        # apply highlight on xlsx
        self._maybe_highlight_xlsx(target)
        self._target_cell = target
        return self._target_cell

    def get_style(self):
        return self.target_cell.style

    def finish(self):
        self.apply_filters()
        self._log_cell_write()



class CellContext(Base):

    def __init__(self, sheet_writer, cell_node, value, data_type):
        super().__init__(sheet_writer, cell_node, value, data_type)
        self._style = None

    def get_style(self):
        if self._style:
            return self._style
        source_cell = self.source_cell
        if source_cell.xf_index is not None:
            style = self.sheet_writer.style_list[source_cell.xf_index]
        else:
            style = self.style_list[0]
        self._style = copy.copy(style)
        return self._style

    def set_cell(self):
        cty = self.data_type
        if cty == xlrd.XL_CELL_EMPTY:
            return
        source_cell = self.source_cell
        rdrowx = self.rdrowx
        rdcolx = self.rdcolx
        wtrowx = self.wtrowx
        wtcolx = self.wtcolx
        value = self.value
        style = self._maybe_highlight_xls(self.get_style())

        if value is None:
            value = source_cell.value
            cty = source_cell.ctype
        if cty is None:
            value, cty = get_type(value)

        wtrow = self.wtsheet.row(wtrowx)
        if cty == xlrd.XL_CELL_TEXT:
            if isinstance(value, (list, tuple)):
                wtrow.set_cell_rich_text(wtcolx, value, style)
            elif value.startswith('='):
                try:
                    formula = xlwt.Formula(value[1:])
                    wtrow.set_cell_formula(wtcolx, formula, style)
                except BaseException as e:
                    wtrow.set_cell_text(wtcolx, value, style)
            else:
                wtrow.set_cell_text(wtcolx, value, style)
        elif cty == xlrd.XL_CELL_NUMBER or cty == xlrd.XL_CELL_DATE:
            wtrow.set_cell_number(wtcolx, value, style)
        elif cty == xlrd.XL_CELL_BLANK:
            wtrow.set_cell_blank(wtcolx, style)
        elif cty == xlrd.XL_CELL_BOOLEAN:
            wtrow.set_cell_boolean(wtcolx, value, style)
        elif cty == xlrd.XL_CELL_ERROR:
            wtrow.set_cell_error(wtcolx, value, style)
        else:
            raise Exception(
                "Unknown xlrd cell type %r with value %r at (sheet=%r,rowx=%r,colx=%r)" \
                % (cty, value, self.rdsheet.name, rdrowx, rdcolx)
            )
    def finish(self):
        self.apply_filters()
        self.set_cell()
        self._log_cell_write()
