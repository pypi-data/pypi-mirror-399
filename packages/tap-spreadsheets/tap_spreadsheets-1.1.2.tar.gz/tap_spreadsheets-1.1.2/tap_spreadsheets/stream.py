"""SpreadsheetStream class."""

from __future__ import annotations

import typing as t
import re
import fnmatch
import csv
from openpyxl import load_workbook
from singer_sdk.streams import Stream
from singer_sdk import typing as th
from decimal import Decimal
from datetime import datetime, date, time, timedelta, timezone
from tap_spreadsheets.storage import Storage
import os
from urllib.parse import urlparse


if t.TYPE_CHECKING:
    from singer_sdk.helpers.types import Context

SDC_INCREMENTAL_KEY = "_sdc_last_modified"
SDC_FILENAME = "_sdc_filename"
SDC_WORKSHEET = "_sdc_worksheet"
SDC_STREAM = "_sdc_stream"


def normalize_path(path: str) -> str:
    """Convert file:/// URIs or relative paths to absolute FS paths."""
    if path.startswith("file://"):
        return urlparse(path).path  # gives /data/...
    return os.path.abspath(path)


def to_iso8601(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat()


def parse_bookmark(val: str | None) -> datetime | None:
    if not val:
        return None
    clean = val.replace("Z", "+00:00")
    return datetime.fromisoformat(clean).astimezone(timezone.utc)


class SpreadsheetStream(Stream):
    """Stream class for spreadsheet (CSV/Excel) files."""

    def __init__(self, tap, file_cfg: dict) -> None:

        self.tap = tap

        self.file_cfg = file_cfg

        self.format: str = file_cfg["format"].lower()
        self.worksheet_ref: str | int | None = file_cfg.get("worksheet")
        self.table_name = file_cfg.get("table_name")

        if not self.table_name:
            self.table_name = "spreadsheet_stream"
            if self.format == "excel" and self.worksheet_ref:
                self.table_name = f"{self.table_name}_{self.worksheet_ref}"

        super().__init__(tap, name=self.table_name)

        self.state_partitioning_keys = [SDC_FILENAME, SDC_STREAM, SDC_WORKSHEET]
        self.replication_key = SDC_INCREMENTAL_KEY
        self.forced_replication_method = "INCREMENTAL"

        self.primary_keys = [
            self._stem_header(n, 0) for n in file_cfg.get("primary_keys", [])
        ]
        self.drop_empty = file_cfg.get("drop_empty", True)
        self.skip_columns = file_cfg.get("skip_columns", 0)
        self.skip_rows = file_cfg.get("skip_rows", 0)
        self.sample_rows = file_cfg.get("sample_rows", 100)
        self.column_headers = file_cfg.get("column_headers")

        self._schema = None
        self._headers: list[str] = []
        self.storage = Storage(self.file_cfg["path"])

    @property
    def is_sorted(self) -> bool:
        """The stream returns records in order."""
        return False

    def get_partition_name(self, filepath: str) -> str:
        return self.storage.normalize_path(filepath)

    def _stem_header(self, h: t.Any, idx: int) -> str:
        """Normalize header names to safe identifiers."""
        # If we accidentally get a Cell or other object, extract its value
        if hasattr(h, "value"):
            h = h.value

        if h is None or str(h).strip() == "":
            return f"col_{idx}"

        h = str(h)
        import unicodedata, re

        h = unicodedata.normalize("NFKD", h).encode("ascii", "ignore").decode()
        h = h.replace("\n", " ").replace("/", " ")
        h = h.lower()
        h = re.sub(r"[^a-z0-9]+", "_", h)
        h = re.sub(r"_+", "_", h)
        h = h.strip("_")

        return h or f"col_{idx}"

    def _infer_type(self, col_values: list[t.Any]):
        """Infer a JSON schema type from sample values with safe fallback."""
        if not col_values:
            return th.StringType()

        norm = []
        for v in col_values:
            if isinstance(v, Decimal):
                norm.append(float(v))
            else:
                norm.append(v)
        col_values = norm

        if all(isinstance(v, int) for v in col_values):
            return th.IntegerType()
        if all(isinstance(v, (int, float)) for v in col_values):
            return th.NumberType()
        if all(isinstance(v, str) for v in col_values):
            return th.StringType()
        if any(isinstance(v, (int, float)) for v in col_values):
            return th.NumberType()
        return th.StringType()

    def _coerce_value(self, v: t.Any, expected_type: str | None = None) -> t.Any:
        if v is None:
            return None
        if expected_type == "integer":
            try:
                return int(v)
            except Exception:
                return None
        if expected_type == "number":
            try:
                return float(v)
            except Exception:
                return None
        if expected_type == "string":
            return str(v)
        if isinstance(v, Decimal):
            return float(v)
        if isinstance(v, (datetime, date, time)):
            return v.isoformat()
        if isinstance(v, timedelta):
            return v.total_seconds()
        return v

    def _extract_headers_excel(self, file: str) -> list[str]:
        with self.storage.open(file, "rb") as fh:
            wb = load_workbook(fh, read_only=True, data_only=True)

            ws = None
            if isinstance(self.worksheet_ref, int):
                if self.worksheet_ref < len(wb.worksheets):
                    ws = wb.worksheets[self.worksheet_ref]
            elif isinstance(self.worksheet_ref, str):
                if self.worksheet_ref in wb.sheetnames:
                    ws = wb[self.worksheet_ref]
                else:
                    pattern = self.worksheet_ref
                    regex = (
                        re.compile(fnmatch.translate(pattern))
                        if any(ch in pattern for ch in ["*", "?"])
                        else re.compile(pattern)
                    )
                    matches = [name for name in wb.sheetnames if regex.match(name)]
                    if matches:
                        ws = wb[matches[0]]
            if ws is None:
                self.logger.warning(
                    "No matching worksheet found in %s. Skipping file.", file
                )
                return []  # skip schema for this file

            header_row = ws.iter_rows(
                min_row=self.skip_rows + 1,
                max_row=self.skip_rows + 1,
                values_only=True,
            )
            raw_headers = next(header_row)
            return [self._stem_header(h, i) for i, h in enumerate(raw_headers)][
                self.skip_columns :
            ]

    def _iter_excel(self, file: str):
        """Iterate data rows (excluding header) from all matched worksheets."""
        with self.storage.open(file, "rb") as fh:
            wb = load_workbook(fh, read_only=True, data_only=True)

            worksheets = []
            if isinstance(self.worksheet_ref, int):
                try:
                    worksheets = [wb.worksheets[self.worksheet_ref]]
                except IndexError:
                    self.logger.warning(
                        "Worksheet index %s out of range in %s. Skipping file.",
                        self.worksheet_ref,
                        file,
                    )
                    return  # skip file

            elif isinstance(self.worksheet_ref, str):
                if self.worksheet_ref in wb.sheetnames:
                    worksheets = [wb[self.worksheet_ref]]
                else:
                    pattern = self.worksheet_ref
                    regex = (
                        re.compile(fnmatch.translate(pattern))
                        if any(ch in pattern for ch in ["*", "?"])
                        else re.compile(pattern)
                    )
                    matches = [name for name in wb.sheetnames if regex.match(name)]
                    if not matches:
                        self.logger.warning(
                            "No worksheets match '%s' in %s. Skipping file. Available: %s",
                            pattern,
                            file,
                            wb.sheetnames,
                        )
                        return  # skip file
                    worksheets = [wb[name] for name in matches]
            else:
                self.logger.warning(
                    "Invalid worksheet_ref %s. Skipping file %s",
                    self.worksheet_ref,
                    file,
                )
                return

            for ws in worksheets:
                start_row = self.skip_rows + 2
                for row in ws.iter_rows(min_row=start_row, values_only=True):
                    yield row

    def _detect_csv_dialect(self, file: str) -> tuple[str, str]:
        """Detect delimiter and quotechar for a CSV file, with config override and fallback."""
        delimiter: str = self.file_cfg.get("delimiter", "")
        quotechar: str = self.file_cfg.get("quote_char", "")

        with self.storage.open(file, "rt") as fh:
            sample = fh.read(4096)

        if not delimiter or not quotechar:
            try:
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample)
                delimiter = delimiter or str(dialect.delimiter)
                quotechar = quotechar or str(dialect.quotechar)
            except csv.Error:
                delimiter = delimiter or ","
                quotechar = quotechar or '"'

        return delimiter, quotechar

    def _extract_headers_csv(self, file: str) -> list[str]:
        """Extract and normalize headers from a CSV file."""
        delimiter, quotechar = self._detect_csv_dialect(file)

        with self.storage.open(file, "rt") as fh:
            reader = csv.reader(fh, delimiter=delimiter, quotechar=quotechar)

            # Skip configured rows before header
            for _ in range(self.skip_rows):
                next(reader, None)

            try:
                raw_headers = next(reader)
            except StopIteration:
                raise ValueError(f"No header row found in {file}")

        headers: list[str] = []
        for i, h in enumerate(raw_headers):
            if h is None or str(h).strip() == "":
                headers.append(f"col_{i}")
            elif str(h).strip().isnumeric():
                headers.append(f"col_{i}")
            else:
                headers.append(self._stem_header(h, i))

        return headers[self.skip_columns :]

    def _iter_csv(self, file: str):
        """Iterate data rows (excluding header) from a CSV file."""
        delimiter, quotechar = self._detect_csv_dialect(file)

        with self.storage.open(file, "rt") as fh:
            reader = csv.reader(fh, delimiter=delimiter, quotechar=quotechar)

            # skip configured number of rows + the header
            for _ in range(self.skip_rows + 1):
                next(reader, None)

            for row in reader:
                yield row

    def get_files(self):
        files = self.storage.glob()
        if not files or len(files) == 0:
            self.logger.warning(
                "No %s files found for %s", self.format, self.file_cfg["path"]
            )
        return files or []

    @property
    def schema(self) -> dict:
        if self._schema:
            return self._schema

        props = []

        files = self.get_files()
        if len(files):
            sample_file = files[0]
            rows = (
                self._iter_excel(sample_file)
                if self.format == "excel"
                else self._iter_csv(sample_file)
            )

            if self.column_headers:
                headers = [
                    self._stem_header(h, i) for i, h in enumerate(self.column_headers)
                ]
            else:
                if self.format == "excel":
                    headers = self._extract_headers_excel(sample_file)
                else:
                    headers = self._extract_headers_csv(sample_file)

            self._headers = headers

            samples = []
            for i, row in enumerate(rows):
                if i >= self.sample_rows:
                    break
                samples.append(row)

            types: dict[str, th.JSONTypeHelper] = {}
            for i, h in enumerate(headers):
                col_idx = i + self.skip_columns
                col_values = [
                    r[col_idx]
                    for r in samples
                    if col_idx < len(r) and r[col_idx] not in (None, "")
                ]
                types[h] = self._infer_type(col_values)

            # build schema with replication key
            props = [th.Property(name.lower(), tpe) for name, tpe in types.items()]

        props.append(
            th.Property(
                SDC_INCREMENTAL_KEY,
                th.DateTimeType(nullable=True),
                description="Replication checkpoint (file mtime or row date)",
            )
        )
        props.append(
            th.Property(
                SDC_FILENAME,
                th.StringType(nullable=True),
                description="Filename reference",
            ),
        )
        props.append(
            th.Property(
                SDC_STREAM,
                th.StringType(nullable=True),
                description="Stream (table_name) reference",
            )
        )
        props.append(
            th.Property(
                SDC_WORKSHEET,
                th.StringType(nullable=True),
                description="Worksheet reference (null for CSV)",
            )
        )
        self._schema = th.PropertiesList(*props).to_dict()

        overrides = self.file_cfg.get("schema_overrides", {})
        for col, props in overrides.items():
            if col in self._schema["properties"]:
                # Fix: replace Python None with the JSON Schema string "null"
                if "type" in props:
                    props["type"] = ["null" if t is None else t for t in props["type"]]
                self._schema["properties"][col].update(props)

        return self._schema

    def get_partition_context(self, filepath: str) -> dict[str, t.Any]:
        """Return the one true partition context for this file."""
        return {
            SDC_FILENAME: self.get_partition_name(filepath),
            SDC_STREAM: self.table_name,
            SDC_WORKSHEET: str(self.worksheet_ref) if self.format == "excel" else None,
        }

    def process_file(
        self,
        filepath: str,
        headers: list[str],
        expected_types: dict[str, t.Any],
        context: Context,
    ) -> list[dict]:
        """Process one file with state awareness and return its records."""

        # load bookmark
        last_bookmark = self.get_starting_replication_key_value(context)
        bookmark_dt = parse_bookmark(last_bookmark)

        info = self.storage.describe(filepath)
        mtime = info.mtime

        self.logger.debug(
            "Partition context: %s, last_bookmark=%s, mtime=%s",
            context,
            bookmark_dt,
            mtime,
        )

        # skip if already processed
        if bookmark_dt and mtime <= bookmark_dt:
            self.logger.info(
                "Skipping %s (mtime=%s <= bookmark=%s)", filepath, mtime, bookmark_dt
            )
            return []

        # parse file rows
        rows = (
            self._iter_excel(filepath)
            if self.format == "excel"
            else self._iter_csv(filepath)
        )
        records: list[dict] = []

        for row in rows:
            record = {
                h: self._coerce_value(
                    (
                        row[i + self.skip_columns]
                        if i + self.skip_columns < len(row)
                        else None
                    ),
                    (
                        "integer"
                        if "integer" in expected_types.get(h, [])
                        else (
                            "number"
                            if "number" in expected_types.get(h, [])
                            else "string"
                        )
                    ),
                )
                for i, h in enumerate(headers)
            }
            if self.drop_empty and any(
                record.get(pk) in (None, "") for pk in self.primary_keys
            ):
                continue

            record[SDC_INCREMENTAL_KEY] = to_iso8601(mtime)
            record[SDC_FILENAME] = filepath
            record[SDC_STREAM] = self.table_name
            record[SDC_WORKSHEET] = (
                str(self.worksheet_ref) if self.format == "excel" else None
            )
            records.append(record)

        if records:
            self.logger.info("Processed %d rows from %s", len(records), filepath)
            self._increment_stream_state(
                {SDC_INCREMENTAL_KEY: to_iso8601(mtime)},
                context=context,
            )

        return records

    def get_records(self, context: Context | None) -> t.Iterable[dict]:
        """Yield records for all files matching this stream's glob."""
        if not self._headers:
            _ = self.schema
        headers = self._headers

        expected_types = {
            name: schema_def.get("type", ["string"])
            for name, schema_def in self.schema["properties"].items()
        }

        files = self.get_files()
        if not files:
            self.logger.warning(
                "No %s files found for %s", self.format, self.file_cfg["path"]
            )
            yield from []

        for filepath in sorted(files):
            yield from self.process_file(
                filepath, headers, expected_types, context or {}
            )
