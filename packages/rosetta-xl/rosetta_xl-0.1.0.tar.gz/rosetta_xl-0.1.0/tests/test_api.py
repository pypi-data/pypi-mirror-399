"""Tests for the Rosetta API."""

import io
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from rosetta.api import app


@pytest.fixture
def client():
    """Create a test client for the API."""
    return TestClient(app)


def create_mock_translate_file(sample_excel_bytes):
    """Create a mock translate_file that creates an actual output file."""
    def mock_translate(input_file, output_file, target_lang, source_lang=None, context=None, sheets=None):
        # Create an actual output file (copy of input)
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "[TR] Hello"
        ws["A2"] = "[TR] World"
        wb.save(output_file)

        return {
            "cell_count": 2,
            "rich_text_cells": 0,
            "dropdown_count": 0,
            "status": "completed",
        }
    return mock_translate


@pytest.fixture
def sample_excel_bytes():
    """Create a simple Excel file in memory."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["A2"] = "World"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@pytest.fixture
def empty_excel_bytes():
    """Create an Excel file with no text content."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 123  # Number, not text
    ws["A2"] = "=SUM(1,2)"  # Formula

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


class TestHealthCheck:
    """Tests for the health check endpoint."""

    def test_root_returns_ok(self, client):
        """GET / should return status ok."""
        response = client.get("/")
        assert response.status_code == 200
        assert response.json() == {"status": "ok", "service": "rosetta"}


class TestTranslateEndpoint:
    """Tests for the /translate endpoint."""

    def test_missing_file_returns_422(self, client):
        """POST /translate without file should return 422."""
        response = client.post("/translate", data={"target_lang": "french"})
        assert response.status_code == 422

    def test_missing_target_lang_returns_422(self, client, sample_excel_bytes):
        """POST /translate without target_lang should return 422."""
        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
        )
        assert response.status_code == 422

    def test_invalid_file_type_returns_400(self, client):
        """POST /translate with non-Excel file should return 400."""
        response = client.post(
            "/translate",
            files={"file": ("test.txt", b"Hello world")},
            data={"target_lang": "french"},
        )
        assert response.status_code == 400
        assert "Invalid file type" in response.json()["detail"]

    def test_empty_filename_returns_error(self, client, sample_excel_bytes):
        """POST /translate with empty filename should return error."""
        response = client.post(
            "/translate",
            files={"file": ("", sample_excel_bytes)},
            data={"target_lang": "french"},
        )
        # FastAPI returns 422 for validation errors
        assert response.status_code in (400, 422)

    def test_no_translatable_content_returns_400(self, client, empty_excel_bytes):
        """POST /translate with no text content should return 400."""
        response = client.post(
            "/translate",
            files={"file": ("empty.xlsx", empty_excel_bytes)},
            data={"target_lang": "french"},
        )
        assert response.status_code == 400
        assert "No translatable content" in response.json()["detail"]

    @patch("rosetta.api.app.translate_file")
    def test_successful_translation(self, mock_translate, client, sample_excel_bytes):
        """POST /translate with valid file should return translated file."""
        # Mock needs to create actual output file
        mock_translate.side_effect = create_mock_translate_file(sample_excel_bytes)

        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
            data={"target_lang": "french"},
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert response.headers["x-cells-translated"] == "2"
        assert "test_french.xlsx" in response.headers.get("content-disposition", "")

    @patch("rosetta.api.app.translate_file")
    def test_translation_with_source_lang(self, mock_translate, client, sample_excel_bytes):
        """POST /translate with source_lang should pass it to translate_file."""
        mock_translate.side_effect = create_mock_translate_file(sample_excel_bytes)

        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
            data={"target_lang": "french", "source_lang": "english"},
        )

        assert response.status_code == 200
        # Verify source_lang was passed
        call_kwargs = mock_translate.call_args.kwargs
        assert call_kwargs["source_lang"] == "english"

    @patch("rosetta.api.app.translate_file")
    def test_translation_with_context(self, mock_translate, client, sample_excel_bytes):
        """POST /translate with context should pass it to translate_file."""
        mock_translate.side_effect = create_mock_translate_file(sample_excel_bytes)

        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
            data={
                "target_lang": "french",
                "context": "Medical terminology",
            },
        )

        assert response.status_code == 200
        call_kwargs = mock_translate.call_args.kwargs
        assert call_kwargs["context"] == "Medical terminology"

    @patch("rosetta.api.app.count_cells")
    @patch("rosetta.api.app.translate_file")
    def test_translation_with_sheets(self, mock_translate, mock_count, client, sample_excel_bytes):
        """POST /translate with sheets param should filter sheets."""
        mock_count.return_value = 2  # Pretend we found cells
        mock_translate.side_effect = create_mock_translate_file(sample_excel_bytes)

        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
            data={
                "target_lang": "french",
                "sheets": "Sheet1, Sheet2",
            },
        )

        assert response.status_code == 200
        call_kwargs = mock_translate.call_args.kwargs
        assert call_kwargs["sheets"] == {"Sheet1", "Sheet2"}

    @patch("rosetta.api.app.translate_file")
    def test_translation_error_returns_500(self, mock_translate, client, sample_excel_bytes):
        """Translation errors should return 500."""
        mock_translate.side_effect = Exception("API error")

        response = client.post(
            "/translate",
            files={"file": ("test.xlsx", sample_excel_bytes)},
            data={"target_lang": "french"},
        )

        assert response.status_code == 500
        assert "Translation failed" in response.json()["detail"]


class TestFileSizeLimits:
    """Tests for file size validation."""

    def test_large_file_returns_400(self, client):
        """Files over 10MB should be rejected."""
        # Create a file larger than 10MB
        large_content = b"x" * (11 * 1024 * 1024)

        response = client.post(
            "/translate",
            files={"file": ("large.xlsx", large_content)},
            data={"target_lang": "french"},
        )

        assert response.status_code == 400
        assert "File too large" in response.json()["detail"]
