"""Pytest fixtures for Rosetta tests."""

from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from rosetta.core.config import Config
from rosetta.models import Cell, TranslationBatch
from rosetta.services import Translator


@pytest.fixture
def mock_config():
    """Create a mock configuration."""
    config = Config(
        anthropic_api_key="test-key",
        model="claude-sonnet-4-20250514",
        batch_size=50,
        max_retries=3,
    )
    return config


@pytest.fixture
def mock_translator(mock_config):
    """Create a mock translator that returns predictable translations.

    The mock simply prepends '[TR] ' to each input text.
    """
    with patch.object(Translator, "__init__", lambda self, config: None):
        translator = Translator(mock_config)

        def mock_translate_batch(batch: TranslationBatch) -> list[str]:
            """Mock translation: prepend [TR] to each text."""
            return [f"[TR] {cell.value}" for cell in batch.cells]

        translator.translate_batch = MagicMock(side_effect=mock_translate_batch)
        yield translator


@pytest.fixture
def simple_excel_file(tmp_path):
    """Create a simple Excel file with basic text cells."""
    file_path = tmp_path / "simple.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Hello"
    ws["A2"] = "World"
    ws["B1"] = "Bonjour"
    ws["B2"] = "Monde"

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_with_formulas(tmp_path):
    """Create an Excel file with formulas (should be skipped)."""
    file_path = tmp_path / "formulas.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Price"
    ws["A2"] = 100
    ws["A3"] = 200
    ws["A4"] = "=SUM(A2:A3)"  # Formula should be skipped
    ws["B1"] = "Total:"

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_with_dropdown(tmp_path):
    """Create an Excel file with inline dropdown validation."""
    file_path = tmp_path / "dropdown.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Status"
    ws["A2"] = "Yes"

    # Add inline dropdown validation
    dv = DataValidation(type="list", formula1='"Yes,No,Maybe"', allow_blank=True)
    dv.add("A2:A10")
    ws.add_data_validation(dv)

    wb.save(file_path)
    return file_path


@pytest.fixture
def excel_with_multiple_sheets(tmp_path):
    """Create an Excel file with multiple sheets."""
    file_path = tmp_path / "multi_sheet.xlsx"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Hello"
    ws1["A2"] = "World"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Bonjour"
    ws2["A2"] = "Monde"

    ws3 = wb.create_sheet("Sheet3")
    ws3["A1"] = "Hola"
    ws3["A2"] = "Mundo"

    wb.save(file_path)
    return file_path
