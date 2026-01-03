"""Excel file extraction service."""

from pathlib import Path
from typing import Iterator, Optional

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell as OpenpyxlCell

from rosetta.core.exceptions import ExcelError
from rosetta.models import Cell


class ExcelExtractor:
    """Extracts translatable text from Excel files."""

    def __init__(self, file_path: Path, sheets: Optional[set[str]] = None) -> None:
        """Initialize the extractor with an Excel file.

        Args:
            file_path: Path to the Excel file.
            sheets: Optional set of sheet names to extract from.
                    If None, all sheets are extracted.
        """
        self.file_path = file_path
        self.sheets_filter = sheets
        try:
            self.workbook = load_workbook(file_path)
        except Exception as e:
            raise ExcelError(f"Failed to load Excel file: {e}") from e

    @property
    def sheet_names(self) -> list[str]:
        """Return all sheet names in the workbook."""
        return self.workbook.sheetnames

    def extract_cells(self) -> Iterator[Cell]:
        """Extract all cells with translatable text content.

        Yields cells that:
        - Have text content (strings)
        - Are not formulas
        - Are not empty
        - Are in the selected sheets (if filter is specified)
        """
        for sheet_name in self.workbook.sheetnames:
            # Skip sheets not in the filter (if filter is specified)
            if self.sheets_filter is not None and sheet_name not in self.sheets_filter:
                continue

            sheet = self.workbook[sheet_name]

            for row in sheet.iter_rows():
                for openpyxl_cell in row:
                    if self._is_translatable(openpyxl_cell):
                        yield self._to_cell_model(openpyxl_cell, sheet_name)

    def _is_translatable(self, cell: OpenpyxlCell) -> bool:
        """Check if a cell contains translatable content."""
        # Skip empty cells
        if cell.value is None:
            return False

        # Skip non-string values (numbers, dates, etc.)
        if not isinstance(cell.value, str):
            return False

        # Skip empty strings or whitespace-only
        if not cell.value.strip():
            return False

        # Skip formulas (they start with '=')
        if cell.value.startswith("="):
            return False

        return True

    def _to_cell_model(self, openpyxl_cell: OpenpyxlCell, sheet_name: str) -> Cell:
        """Convert openpyxl cell to our Cell model."""
        return Cell(
            sheet=sheet_name,
            row=openpyxl_cell.row,
            col=openpyxl_cell.column,
            value=str(openpyxl_cell.value),
            original_value=str(openpyxl_cell.value),
        )

    def close(self) -> None:
        """Close the workbook."""
        self.workbook.close()

    def __enter__(self) -> "ExcelExtractor":
        return self

    def __exit__(self, *args: object) -> None:
        self.close()
