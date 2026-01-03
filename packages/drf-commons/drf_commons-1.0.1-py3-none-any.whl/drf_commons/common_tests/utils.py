"""
Test utilities for DRF Commons library tests.
"""

import csv
import io
import logging
from contextlib import contextmanager
from unittest.mock import patch

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import HttpResponse
from django.test import override_settings

from drf_commons.current_user.utils import USER_ATTR_NAME, _set_current_user, _thread_locals, get_current_user
from drf_commons.common_conf.settings import clear_settings_cache


def mock_current_user(user):
    """Set current user in thread-local storage for testing."""
    _set_current_user(user)


@contextmanager
def temporary_current_user(user):
    """Context manager for temporary current user."""
    original_user = None
    try:
        original_user = get_current_user()
        mock_current_user(user)
        yield
    finally:
        if original_user:
            mock_current_user(original_user)


def clear_current_user():
    """Clear current user from thread-local storage for testing."""
    try:
        if hasattr(_thread_locals, USER_ATTR_NAME):
            delattr(_thread_locals, USER_ATTR_NAME)
    except ImportError:
        pass


def create_test_file(
    filename="test.txt", content="test content", content_type="text/plain"
):
    """Create test file for upload testing."""
    if isinstance(content, str):
        content = content.encode("utf-8")
    return SimpleUploadedFile(filename, content, content_type=content_type)


def create_csv_file(headers, rows, filename="test.csv"):
    """Create CSV file for import testing."""

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    content = output.getvalue().encode("utf-8")
    return SimpleUploadedFile(filename, content, content_type="text/csv")


def create_excel_file(headers, rows, filename="test.xlsx"):
    """Create Excel file for import testing."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write headers
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    # Write rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=value)

    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    content = output.getvalue()

    return SimpleUploadedFile(
        filename,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def assert_response_success(response, expected_status=200):
    """Assert response is successful with DRF Commons format."""
    assert response.status_code == expected_status
    assert "data" in response.data
    assert "message" in response.data
    assert "success" in response.data
    assert response.data["success"] is True


def assert_response_error(response, expected_status=400):
    """Assert response is error with DRF Commons format."""
    assert response.status_code == expected_status
    assert "message" in response.data
    assert "success" in response.data
    assert response.data["success"] is False


def assert_model_fields_set(instance, user, check_created=True, check_updated=True):
    """Assert model instance has correct user tracking fields set."""
    if check_created and hasattr(instance, "created_by"):
        assert instance.created_by == user
    if check_updated and hasattr(instance, "updated_by"):
        assert instance.updated_by == user
    if hasattr(instance, "created_at"):
        assert instance.created_at is not None
    if hasattr(instance, "updated_at"):
        assert instance.updated_at is not None


def assert_queryset_equal(qs1, qs2, msg=None):
    """Assert two querysets contain the same objects."""
    list1 = list(qs1.order_by("pk"))
    list2 = list(qs2.order_by("pk"))
    assert list1 == list2, msg


@contextmanager
def mock_file_download_response():
    """Mock file download response for testing."""
    with patch("django.http.HttpResponse") as mock_response:
        mock_response.return_value = HttpResponse()
        yield mock_response


@contextmanager
def override_debug_settings(**settings_dict):
    """Override DRF Commons debug settings for testing."""
    with override_settings(**settings_dict):
        # Clear any cached settings
        try:
            clear_settings_cache()
        except ImportError:
            pass
        yield


@contextmanager
def capture_logs(logger_name, level=logging.DEBUG):
    """Capture log messages for testing."""
    log_capture = io.StringIO()
    handler = logging.StreamHandler(log_capture)
    handler.setLevel(level)

    logger = logging.getLogger(logger_name)
    original_level = logger.level
    logger.setLevel(level)
    logger.addHandler(handler)

    try:
        yield log_capture
    finally:
        logger.removeHandler(handler)
        logger.setLevel(original_level)


def assert_log_contains(log_output, expected_text):
    """Assert log output contains expected text."""
    content = (
        log_output.getvalue() if hasattr(log_output, "getvalue") else str(log_output)
    )
    assert (
        expected_text in content
    ), f"Expected '{expected_text}' in log output: {content}"


def assert_no_logs(log_output):
    """Assert no log messages were captured."""
    content = (
        log_output.getvalue() if hasattr(log_output, "getvalue") else str(log_output)
    )
    assert not content.strip(), f"Expected no log output, got: {content}"
