"""
Django management command to generate import template files from viewset configurations.

Usage:
    python manage.py generate_import_template accounts.StudentViewSet
    python manage.py generate_import_template myapp.MyViewSet --filename custom_template.xlsx
"""

import importlib
import re
from pathlib import Path
from typing import Any, Dict, List

from django.conf import settings as django_settings
from django.core.management.base import BaseCommand, CommandError

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class Command(BaseCommand):
    help = """
    Generate Excel import template file from a ViewSet's import_file_config.

    Examples:
        python manage.py generate_import_template accounts.StudentViewSet
        python manage.py generate_import_template myapp.MyViewSet --filename custom_template.xlsx
        python manage.py generate_import_template myapp.MyViewSet --format csv
    """

    def add_arguments(self, parser):
        parser.add_argument(
            "viewset_path",
            type=str,
            help='ViewSet path in format "app.ViewSetClass" (e.g., "accounts.StudentViewSet")',
        )
        parser.add_argument(
            "--filename",
            type=str,
            help="Custom filename for the template (default: use import_template_name from viewset)",
        )
        parser.add_argument(
            "--format",
            choices=["xlsx", "csv"],
            default="xlsx",
            help="Output format (default: xlsx)",
        )
        parser.add_argument(
            "--include-examples",
            action="store_true",
            help="Include example data rows in the template",
        )
        parser.add_argument(
            "--order-by",
            choices=["config", "required-first", "alphabetic"],
            default="required-first",
            help="Column ordering: config (preserve config order), required-first (required fields first, then optional), or alphabetic (A-Z)",
        )

    def handle(self, *args, **options):
        viewset_path = options["viewset_path"]
        filename = options.get("filename")
        format_type = options["format"]
        include_examples = options["include_examples"]
        order_by = options["order_by"]

        try:
            # Load the viewset class
            viewset_class = self.load_viewset_class(viewset_path)

            # Validate viewset has required attributes
            self.validate_viewset(viewset_class, viewset_path)

            # Extract columns from config
            columns_ordered, column_required_status = self.extract_columns_from_config(
                viewset_class.import_file_config
            )

            if not columns_ordered:
                raise CommandError("No columns found in import_file_config")

            # Use import_template_name if no filename provided
            if not filename:
                if (
                    hasattr(viewset_class, "import_template_name")
                    and viewset_class.import_template_name
                ):
                    filename = viewset_class.import_template_name
                else:
                    filename = self.generate_filename(viewset_class, format_type)

            # Ensure filename has correct extension
            if not filename.endswith(f".{format_type}"):
                filename = f"{filename.rsplit('.', 1)[0]}.{format_type}"

            # Create template file (always overwrite)
            template_path = self.create_template_file(
                columns_ordered,
                column_required_status,
                filename,
                format_type,
                include_examples,
                order_by,
            )

            self.stdout.write(
                self.style.SUCCESS(f"✅ Template created successfully: {template_path}")
            )

            # Show column summary with color coding info
            required_count = sum(
                1 for is_req in column_required_status.values() if is_req
            )
            optional_count = len(columns_ordered) - required_count

            self.stdout.write(f"📋 Template contains {len(columns_ordered)} columns:")
            self.stdout.write(f"   🔴 {required_count} required fields (red headers)")
            self.stdout.write(f"   🟢 {optional_count} optional fields (green headers)")
            self.stdout.write(f"   📊 Column order: {order_by}")
            self.stdout.write("")

            # Show columns in the order they appear in template
            display_columns = []
            if order_by == "alphabetic":
                display_columns = sorted(columns_ordered)
            elif order_by == "required-first":
                required_cols = [
                    col
                    for col in columns_ordered
                    if column_required_status.get(col, True)
                ]
                optional_cols = [
                    col
                    for col in columns_ordered
                    if not column_required_status.get(col, True)
                ]
                display_columns = required_cols + optional_cols
            else:  # config order
                display_columns = columns_ordered

            for i, column in enumerate(display_columns, 1):
                is_required = column_required_status.get(column, True)
                status_icon = "🔴" if is_required else "🟢"
                status_text = "required" if is_required else "optional"
                self.stdout.write(f"  {i:2d}. {status_icon} {column} ({status_text})")

            # Show usage info
            self.stdout.write("\n📖 Usage:")
            self.stdout.write(
                "   1. Download template: GET /api/.../download-import-template/"
            )
            self.stdout.write("   2. Fill in your data following the column headers")
            self.stdout.write("   3. Upload via: POST /api/.../import-from-file/")

        except Exception as e:
            raise CommandError(f"Failed to generate template: {str(e)}")

    def load_viewset_class(self, viewset_path: str):
        """Load viewset class from module path."""
        try:
            if "." not in viewset_path:
                raise CommandError(
                    "Invalid viewset path format. Use 'app.ViewSetClass' "
                    "(e.g., 'accounts.StudentViewSet')"
                )

            # Split module and class name
            module_parts = viewset_path.split(".")
            class_name = module_parts[-1]
            module_path = ".".join(module_parts[:-1])

            # Try common patterns for viewset locations
            possible_modules = [
                f"{module_path}.views.{class_name.lower().replace('viewset', '')}",
                f"{module_path}.views",
                f"{module_path}.viewsets",
                module_path,
            ]

            viewset_class = None
            for module_name in possible_modules:
                try:
                    module = importlib.import_module(module_name)
                    if hasattr(module, class_name):
                        viewset_class = getattr(module, class_name)
                        break
                except ImportError:
                    continue

            if not viewset_class:
                raise CommandError(
                    f"ViewSet class '{class_name}' not found. "
                    f"Tried modules: {possible_modules}"
                )

            return viewset_class

        except ImportError as e:
            raise CommandError(f"Cannot import module: {str(e)}")

    def validate_viewset(self, viewset_class, viewset_path: str):
        """Validate that viewset has required attributes."""
        if not hasattr(viewset_class, "import_file_config"):
            raise CommandError(
                f"ViewSet '{viewset_path}' does not have 'import_file_config' attribute. "
                f"Make sure it uses FileImportMixin."
            )

        config = viewset_class.import_file_config
        if not config:
            raise CommandError(
                f"ViewSet '{viewset_path}' has empty 'import_file_config'. "
                f"Please define the import configuration."
            )

        if not isinstance(config, dict):
            raise CommandError(
                f"ViewSet '{viewset_path}' import_file_config must be a dictionary."
            )

        required_keys = ["order", "models"]
        for key in required_keys:
            if key not in config:
                raise CommandError(
                    f"ViewSet '{viewset_path}' import_file_config missing required key: '{key}'"
                )

    def extract_columns_from_config(
        self, config: dict
    ) -> tuple[List[str], Dict[str, bool]]:
        """Extract all column names and their required status from the import configuration."""
        columns_ordered = []  # Preserve order from config
        column_required_status = {}  # column_name -> is_required

        self.stdout.write(f"🔍 Processing config with order: {config['order']}")

        for step_key in config["order"]:
            model_config = config["models"][step_key]

            self.stdout.write(
                f"📋 Processing step '{step_key}' (model: {model_config.get('model', 'Unknown')})"
            )

            # Direct columns - preserve order from config
            if "direct_columns" in model_config:
                for field_name, column_name in model_config["direct_columns"].items():
                    if column_name not in column_required_status:  # Avoid duplicates
                        columns_ordered.append(column_name)
                        # Check if field is required by examining model field constraints
                        is_required = self._is_field_required(
                            model_config["model"], field_name, model_config
                        )
                        column_required_status[column_name] = is_required
                        status = "required" if is_required else "optional"
                        self.stdout.write(
                            f"  ✓ Added column: '{column_name}' (direct, {status})"
                        )

            # Transformed columns - check if transformation target is required
            if "transformed_columns" in model_config:
                for field_name, transform_spec in model_config[
                    "transformed_columns"
                ].items():
                    column_name = transform_spec["column"]
                    if column_name not in column_required_status:  # Avoid duplicates
                        columns_ordered.append(column_name)
                        is_required = self._is_field_required(
                            model_config["model"], field_name, model_config
                        )
                        column_required_status[column_name] = is_required
                        status = "required" if is_required else "optional"
                        self.stdout.write(
                            f"  ✓ Added column: '{column_name}' (transformed, {status})"
                        )

            # Lookup fields - check if lookup field is required
            if "lookup_fields" in model_config:
                for field_name, lookup_spec in model_config["lookup_fields"].items():
                    column_name = lookup_spec["column"]
                    if column_name not in column_required_status:  # Avoid duplicates
                        columns_ordered.append(column_name)
                        is_required = self._is_field_required(
                            model_config["model"], field_name, model_config
                        )
                        column_required_status[column_name] = is_required
                        status = "required" if is_required else "optional"
                        self.stdout.write(
                            f"  ✓ Added column: '{column_name}' (lookup, {status})"
                        )

            # Computed fields - only include "if_empty" mode fields that have columns
            if "computed_fields" in model_config:
                for field_name, compute_spec in model_config["computed_fields"].items():
                    mode = compute_spec.get("mode", "if_empty")

                    if mode == "if_empty" and "column" in compute_spec:
                        # Hybrid field: can be provided or generated
                        column_name = compute_spec["column"]
                        if (
                            column_name not in column_required_status
                        ):  # Avoid duplicates
                            columns_ordered.append(column_name)
                            # These are always optional since they can be generated
                            column_required_status[column_name] = False
                            self.stdout.write(
                                f"  ✓ Added column: '{column_name}' (computed-hybrid, optional)"
                            )
                    elif mode == "always":
                        # Fully generated field: no column needed
                        self.stdout.write(
                            f"  ⚡ Skipped field: '{field_name}' (computed-always, no column needed)"
                        )

        return columns_ordered, column_required_status

    def _is_field_required(
        self, model_path: str, field_name: str, model_config: Dict[str, Any] = None
    ) -> bool:
        """Determine if a field is required based on config required_fields and computed fields."""
        # If field is in computed_fields, it's always optional (can be generated)
        if model_config and "computed_fields" in model_config:
            if field_name in model_config["computed_fields"]:
                return False

        # Check if field is in required_fields list
        if model_config and "required_fields" in model_config:
            return field_name in model_config["required_fields"]

        # Default to optional if not explicitly listed as required
        return False

    def generate_filename(self, viewset_class, format_type: str) -> str:
        """Generate a filename based on the viewset class name."""
        class_name = viewset_class.__name__
        # Convert CamelCase to snake_case

        snake_case = re.sub(r"(?<!^)(?=[A-Z])", "_", class_name).lower()
        # Remove 'viewset' suffix if present
        snake_case = snake_case.replace("_viewset", "").replace("viewset", "")
        return f"{snake_case}_import_template.{format_type}"

    def create_template_file(
        self,
        columns_ordered: List[str],
        column_required_status: Dict[str, bool],
        filename: str,
        format_type: str,
        include_examples: bool,
        order_by: str,
    ) -> str:
        """Create the template file with headers and auto-sized columns."""
        # Ensure static/import-templates directory exists
        template_dir = Path(django_settings.BASE_DIR) / "static" / "import-templates"
        template_dir.mkdir(parents=True, exist_ok=True)

        template_path = template_dir / filename

        # Apply column ordering based on user preference
        if order_by == "alphabetic":
            sorted_columns = sorted(columns_ordered)
        elif order_by == "required-first":
            # Sort by required status first, then preserve original order within each group
            required_cols = [
                col for col in columns_ordered if column_required_status.get(col, True)
            ]
            optional_cols = [
                col
                for col in columns_ordered
                if not column_required_status.get(col, True)
            ]
            sorted_columns = required_cols + optional_cols
        else:  # order_by == 'config'
            # Preserve original config order
            sorted_columns = columns_ordered

        # Create DataFrame with headers
        if include_examples:
            # Add example rows
            example_data = []
            for i in range(3):  # 3 example rows
                row = {}
                for col in sorted_columns:
                    row[col] = f"Example {col.lower().replace(' ', '_')} {i+1}"
                example_data.append(row)

            df = pd.DataFrame(example_data, columns=sorted_columns)
        else:
            # Empty template with just headers
            df = pd.DataFrame(columns=sorted_columns)

        # Save file based on format
        if format_type == "xlsx":
            # Create Excel with auto-sized columns

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Import Template"

            # Add legend first
            legend_data = [
                ["LEGEND:", "", "", ""],
                ["🔴 Red headers = Required fields", "", "", ""],
                ["🟢 Green headers = Optional fields", "", "", ""],
                ["", "", "", ""],  # Empty row for spacing
            ]

            for row_data in legend_data:
                ws.append(row_data)

            # Style legend

            legend_font = Font(bold=True, size=10)
            for row_num in range(1, 5):  # Legend rows
                for cell in ws[row_num]:
                    cell.font = legend_font

            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Style headers with different colors for required vs optional fields
            required_font = Font(bold=True, color="FFFFFF")
            required_fill = PatternFill(
                start_color="DC3545", end_color="DC3545", fill_type="solid"
            )  # Red for required

            optional_font = Font(bold=True, color="FFFFFF")
            optional_fill = PatternFill(
                start_color="28A745", end_color="28A745", fill_type="solid"
            )  # Green for optional

            for cell in ws[5]:  # Header row (after legend)
                column_name = cell.value
                is_required = column_required_status.get(
                    column_name, True
                )  # Default to required if unknown

                if is_required:
                    cell.font = required_font
                    cell.fill = required_fill
                else:
                    cell.font = optional_font
                    cell.fill = optional_fill

            # Auto-size columns based on content
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass

                # Add some padding and set minimum width
                adjusted_width = max(max_length + 2, 12)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(template_path)

        else:  # csv
            df.to_csv(template_path, index=False)

        return str(template_path)
