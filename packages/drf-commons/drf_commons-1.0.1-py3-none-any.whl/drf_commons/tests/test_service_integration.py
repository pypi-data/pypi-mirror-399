"""
Service integration tests for import/export functionality.

Tests services with real files and database operations.
"""

import os
import tempfile
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings

import openpyxl
import pandas as pd

from drf_commons.common_tests.factories import UserFactory
from drf_commons.services.export_file.service import ExportService
from drf_commons.services.import_from_file.service import FileImportService

User = get_user_model()


class ImportServiceIntegrationTests(TestCase):
    """Test import service with real files and database operations."""

    def setUp(self):
        self.user = UserFactory()
        self.config = {
            "file_format": "xlsx",
            "order": ["users"],
            "models": {
                "users": {
                    "model": "auth.User",
                    "unique_by": ["username"],
                    "update_if_exists": True,
                    "direct_columns": {
                        "username": "username",
                        "email": "email",
                        "first_name": "first_name"
                    },
                }
            },
        }

    def create_test_excel_file(self, data_rows):
        """Create real Excel file with test data matching service expectations."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "users"

        # Add title rows (service expects header on row 5)
        sheet.cell(row=1, column=1, value="User Import Template")
        sheet.cell(row=2, column=1, value="Fill in the data below")
        sheet.cell(row=3, column=1, value="")
        sheet.cell(row=4, column=1, value="")

        # Headers on row 5
        headers = ["username", "email", "first_name"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=5, column=col, value=header)

        # Data rows starting from row 6
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to BytesIO
        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)
        return file_buffer

    def test_import_new_users_from_excel(self):
        """Test importing new users from Excel file creates database records."""
        test_data = [
            ["testuser1", "test1@example.com", "Test"],
            ["testuser2", "test2@example.com", "User"],
        ]

        excel_file = self.create_test_excel_file(test_data)
        uploaded_file = SimpleUploadedFile(
            "test_users.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        service = FileImportService(self.config)
        results = service.import_file(uploaded_file)

        # Verify database state
        self.assertEqual(User.objects.filter(username__startswith="testuser").count(), 2)

        user1 = User.objects.get(username="testuser1")
        self.assertEqual(user1.email, "test1@example.com")
        self.assertEqual(user1.first_name, "Test")

        # Verify results structure
        self.assertEqual(len(results["rows"]), 2)
        self.assertEqual(results["summary"]["created"], 2)
        self.assertEqual(results["summary"]["failed"], 0)

    def test_import_update_existing_users(self):
        """Test importing existing users updates database records."""
        # Create existing user
        existing_user = UserFactory(username="testuser1", email="old@example.com")

        test_data = [
            ["testuser1", "new@example.com", "Updated"],
        ]

        excel_file = self.create_test_excel_file(test_data)
        uploaded_file = SimpleUploadedFile(
            "update_users.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        service = FileImportService(self.config)
        service.import_file(uploaded_file)

        # Verify update occurred
        existing_user.refresh_from_db()
        self.assertEqual(existing_user.email, "new@example.com")
        self.assertEqual(existing_user.first_name, "Updated")

        # Verify no new user created
        self.assertEqual(User.objects.filter(username="testuser1").count(), 1)

    def test_import_handles_data_validation_errors(self):
        """Test import service handles invalid data gracefully."""
        test_data = [
            ["validuser", "valid@example.com", "Valid"],  # Valid user first
            ["", "invalid-email", "Test"],  # Empty username, invalid email
        ]

        excel_file = self.create_test_excel_file(test_data)
        uploaded_file = SimpleUploadedFile(
            "mixed_data.xlsx",
            excel_file.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        service = FileImportService(self.config)
        results = service.import_file(uploaded_file)

        # Verify results structure
        self.assertEqual(len(results["rows"]), 2)

        # Check that we have some failures due to validation
        self.assertGreater(results["summary"]["failed"], 0)

    def test_import_csv_file_format(self):
        """Test import service works with CSV files."""
        config = dict(self.config)
        config["file_format"] = "csv"
        config["models"]["users"]["direct_columns"] = {
            "username": "username",
            "email": "email",
            "first_name": "first_name"
        }

        # Create CSV content
        csv_content = "username,email,first_name\ntestuser,test@example.com,Test"
        uploaded_file = SimpleUploadedFile(
            "test_users.csv",
            csv_content.encode('utf-8'),
            content_type="text/csv"
        )

        service = FileImportService(config)
        results = service.import_file(uploaded_file)

        # Verify CSV import worked
        self.assertEqual(User.objects.filter(username="testuser").count(), 1)
        self.assertEqual(results["summary"]["created"], 1)


class ExportServiceIntegrationTests(TestCase):
    """Test export service with real data and file generation."""

    def setUp(self):
        # Create test users
        self.users = [
            UserFactory(username="user1", email="user1@test.com", first_name="First"),
            UserFactory(username="user2", email="user2@test.com", first_name="Second"),
        ]

        self.test_data = [
            {"username": "user1", "email": "user1@test.com", "first_name": "First"},
            {"username": "user2", "email": "user2@test.com", "first_name": "Second"},
        ]

        self.includes = ["username", "email", "first_name"]
        self.column_config = {
            "username": {"label": "Username"},
            "email": {"label": "Email Address"},
            "first_name": {"label": "First Name"},
        }
        self.export_headers = ["Username", "Email Address", "First Name"]

    def test_export_csv_generates_valid_file(self):
        """Test CSV export generates downloadable file with correct data."""
        service = ExportService()

        response = service.export_csv(
            data_rows=self.test_data,
            includes=self.includes,
            column_config=self.column_config,
            filename="test_export.csv",
            export_headers=self.export_headers,
            document_titles=["User Export"],
        )

        # Verify response headers
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("attachment", response["Content-Disposition"])
        self.assertIn("test_export.csv", response["Content-Disposition"])

        # Verify CSV content
        content = response.content.decode('utf-8')
        self.assertIn("Username,Email Address,First Name", content)
        self.assertIn("user1,user1@test.com,First", content)
        self.assertIn("user2,user2@test.com,Second", content)

    def test_export_xlsx_generates_valid_file(self):
        """Test Excel export generates valid workbook with data."""
        service = ExportService()

        response = service.export_xlsx(
            data_rows=self.test_data,
            includes=self.includes,
            column_config=self.column_config,
            filename="test_export.xlsx",
            export_headers=[],  # No export headers
            document_titles=[],  # No document titles
        )

        # Verify response headers
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.assertIn("test_export.xlsx", response["Content-Disposition"])

        # Verify Excel content by loading with openpyxl
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active

        # Check headers in row 1 (no export headers or titles)
        self.assertEqual(sheet.cell(row=1, column=1).value, "Username")
        self.assertEqual(sheet.cell(row=1, column=2).value, "Email Address")

        # Check data in row 2
        self.assertEqual(sheet.cell(row=2, column=1).value, "user1")
        self.assertEqual(sheet.cell(row=2, column=2).value, "user1@test.com")

    def test_export_pdf_generates_valid_file(self):
        """Test PDF export generates valid PDF with data."""
        service = ExportService()

        response = service.export_pdf(
            data_rows=self.test_data,
            includes=self.includes,
            column_config=self.column_config,
            filename="test_export.pdf",
            export_headers=self.export_headers,
            document_titles=["User Export"],
        )

        # Verify response headers
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn("test_export.pdf", response["Content-Disposition"])

        # Verify PDF content starts with PDF signature
        content = response.content
        self.assertTrue(content.startswith(b"%PDF"))

    def test_export_data_processing_handles_missing_fields(self):
        """Test export data processing handles missing fields gracefully."""
        incomplete_data = [
            {"username": "user1", "email": "user1@test.com"},  # Missing first_name
            {"username": "user2", "first_name": "Second"},     # Missing email
        ]

        service = ExportService()
        processed = service.process_export_data(
            provided_data=incomplete_data,
            includes=self.includes,
            column_config=self.column_config,
        )

        # Verify processing handles missing fields
        self.assertEqual(len(processed["table_data"]), 2)
        self.assertIn("user1", str(processed["table_data"]))
        self.assertIn("user2", str(processed["table_data"]))


class ImportExportWorkflowIntegrationTests(TestCase):
    """Test complete import-export workflow integration."""

    def test_import_then_export_workflow(self):
        """Test importing data then exporting it maintains data integrity."""
        # Step 1: Import data
        import_config = {
            "file_format": "csv",
            "order": ["users"],
            "models": {
                "users": {
                    "model": "auth.User",
                    "unique_by": ["username"],
                    "update_if_exists": False,
                    "direct_columns": {
                        "username": "username",
                        "email": "email",
                        "first_name": "first_name"
                    },
                }
            },
        }

        csv_content = "username,email,first_name\nworkflow_user,workflow@test.com,Workflow"
        uploaded_file = SimpleUploadedFile(
            "workflow_test.csv",
            csv_content.encode('utf-8'),
            content_type="text/csv"
        )

        import_service = FileImportService(import_config)
        import_results = import_service.import_file(uploaded_file)

        # Verify import
        self.assertEqual(import_results["summary"]["created"], 1)
        imported_user = User.objects.get(username="workflow_user")

        # Step 2: Export the imported data
        export_data = [{
            "username": imported_user.username,
            "email": imported_user.email,
            "first_name": imported_user.first_name,
        }]

        export_service = ExportService()
        csv_response = export_service.export_csv(
            data_rows=export_data,
            includes=["username", "email", "first_name"],
            column_config={
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
            },
            filename="exported_workflow.csv",
            export_headers=["Username", "Email", "First Name"],
            document_titles=["Workflow Export"],
        )

        # Verify export contains original data
        export_content = csv_response.content.decode('utf-8')
        self.assertIn("workflow_user", export_content)
        self.assertIn("workflow@test.com", export_content)
        self.assertIn("Workflow", export_content)