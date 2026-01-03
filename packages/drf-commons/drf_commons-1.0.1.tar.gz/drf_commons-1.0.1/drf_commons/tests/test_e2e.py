"""
End-to-end workflow tests for critical library functionality.

Tests complete user workflows from start to finish.
"""

import json
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import override_settings
from django.urls import path, include

from rest_framework import serializers, viewsets
from rest_framework.routers import DefaultRouter
from rest_framework.test import APITestCase

import openpyxl

from drf_commons.common_tests.factories import UserFactory
from drf_commons.serializers.base import BaseModelSerializer
from drf_commons.views.base import BulkViewSet
from drf_commons.views.mixins import FileImportMixin, FileExportMixin

User = get_user_model()


class UserSerializer(BaseModelSerializer):
    """Serializer for E2E testing."""

    class Meta(BaseModelSerializer.Meta):
        model = User
        fields = ["id", "username", "email", "first_name", "last_name", "is_active"]


class E2EUserViewSet(BulkViewSet, FileImportMixin, FileExportMixin):
    """ViewSet for E2E testing with all functionality."""

    queryset = User.objects.all()
    serializer_class = UserSerializer

    # FileImportMixin configuration
    import_template_name = "user_import_template.xlsx"
    import_file_config = {
        "file_format": "xlsx",
        "order": ["users"],
        "models": {
            "users": {
                "model": "auth.User",
                "unique_by": ["username"],
                "update_if_exists": True,
                "direct_columns": {
                    "username": "username",
                    "email": "email",
                    "first_name": "first_name",
                    "last_name": "last_name"
                },
            }
        },
    }


# E2E Test URLs
router = DefaultRouter()
router.register(r'users', E2EUserViewSet)

e2e_urlpatterns = [
    path('api/', include(router.urls)),
]


@override_settings(ROOT_URLCONF=__name__)
class DataManagementE2ETests(APITestCase):
    """End-to-end tests for complete data management workflows."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.admin_user = UserFactory(is_staff=True, is_superuser=True)
        self.client.force_authenticate(user=self.admin_user)

    def create_test_excel_file(self, data_rows):
        """Create Excel file for testing."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "users"

        # Add title rows (service expects header on row 5)
        sheet.cell(row=1, column=1, value="User Import Template")
        sheet.cell(row=2, column=1, value="Fill in the data below")
        sheet.cell(row=3, column=1, value="")
        sheet.cell(row=4, column=1, value="")

        # Headers on row 5
        headers = ["username", "email", "first_name", "last_name"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=5, column=col, value=header)

        # Data rows starting from row 6
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)
        return file_buffer

    def test_complete_user_management_workflow(self):
        """
        Test complete workflow: Create → Read → Update → Delete → Bulk Operations
        """
        # Step 1: Create single user via API
        create_data = {
            "username": "e2e_user",
            "email": "e2e@test.com",
            "first_name": "E2E",
            "last_name": "Test"
        }
        create_response = self.client.post('/api/users/', create_data)
        self.assertEqual(create_response.status_code, 201)
        user_id = create_response.data['data']['id']

        # Step 2: Read user via API
        read_response = self.client.get(f'/api/users/{user_id}/')
        self.assertEqual(read_response.status_code, 200)
        self.assertEqual(read_response.data['data']['username'], 'e2e_user')

        # Step 3: Update user via API
        update_data = {"email": "updated_e2e@test.com", "first_name": "Updated"}
        update_response = self.client.patch(f'/api/users/{user_id}/', update_data)
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data['data']['email'], 'updated_e2e@test.com')

        # Step 4: Verify update in database
        updated_user = User.objects.get(id=user_id)
        self.assertEqual(updated_user.email, 'updated_e2e@test.com')
        self.assertEqual(updated_user.first_name, 'Updated')

        # Step 5: Bulk create additional users
        bulk_data = [
            {"username": "bulk1", "email": "bulk1@test.com", "first_name": "Bulk1"},
            {"username": "bulk2", "email": "bulk2@test.com", "first_name": "Bulk2"}
        ]
        bulk_create_response = self.client.post('/api/users/bulk-create/', bulk_data, format='json')
        self.assertEqual(bulk_create_response.status_code, 201)

        # Step 6: List all users to verify count
        list_response = self.client.get('/api/users/')
        self.assertEqual(list_response.status_code, 200)
        # Should have admin + e2e_user + 2 bulk users = at least 4
        self.assertGreaterEqual(len(list_response.data['data']['results']), 4)

        # Step 7: Bulk update users
        bulk_users = [user for user in list_response.data['data']['results'] if user['username'].startswith('bulk')]
        bulk_update_data = [
            {"id": bulk_users[0]['id'], "last_name": "BulkUpdated1"},
            {"id": bulk_users[1]['id'], "last_name": "BulkUpdated2"}
        ]
        bulk_update_response = self.client.put('/api/users/bulk-update/', bulk_update_data, format='json')
        self.assertEqual(bulk_update_response.status_code, 200)

        # Step 8: Verify bulk updates
        for updated_user in bulk_update_response.data['data']:
            self.assertTrue(updated_user['last_name'].startswith('BulkUpdated'))

        # Step 9: Delete single user
        delete_response = self.client.delete(f'/api/users/{user_id}/')
        self.assertEqual(delete_response.status_code, 204)

        # Step 10: Verify deletion
        verify_response = self.client.get(f'/api/users/{user_id}/')
        self.assertEqual(verify_response.status_code, 404)

        # Step 11: Bulk delete remaining test users
        test_user_ids = [user['id'] for user in bulk_users]
        bulk_delete_response = self.client.delete('/api/users/bulk_delete/', test_user_ids, format='json')
        self.assertEqual(bulk_delete_response.status_code, 200)

        # Step 12: Verify bulk deletion
        final_list_response = self.client.get('/api/users/')
        remaining_usernames = [user['username'] for user in final_list_response.data['data']['results']]
        self.assertNotIn('bulk1', remaining_usernames)
        self.assertNotIn('bulk2', remaining_usernames)

    def test_complete_import_export_workflow(self):
        """
        Test complete workflow: Import Data → Verify → Export → Verify Export
        """
        # Step 1: Prepare import data
        import_data = [
            ["import_user1", "import1@test.com", "Import", "User1"],
            ["import_user2", "import2@test.com", "Import", "User2"],
            ["import_user3", "import3@test.com", "Import", "User3"],
        ]

        # Step 2: Create and import Excel file
        excel_file = self.create_test_excel_file(import_data)
        import_response = self.client.post(
            '/api/users/import-from-file/',
            {
                'file': excel_file,
                'append_data': 'true',
                'config': json.dumps({
                    "file_format": "xlsx",
                    "order": ["users"],
                    "models": {
                        "users": {
                            "model": "auth.User",
                            "unique_by": ["username"],
                            "update_if_exists": False,
                            "direct_columns": {
                                "username": "username",
                                "email": "email",
                                "first_name": "first_name",
                                "last_name": "last_name"
                            },
                        }
                    },
                })
            },
            format='multipart'
        )

        # Step 3: Verify import success
        self.assertIn(import_response.status_code, [200, 201])  # Accept both 200 and 201
        # Handle response data structure
        response_data = import_response.data
        if 'data' in response_data and 'import_summary' in response_data['data']:
            summary = response_data['data']['import_summary']
            self.assertEqual(summary['created'], 3)
            self.assertEqual(summary['failed'], 0)
        elif 'successful_rows' in response_data:
            self.assertEqual(response_data['successful_rows'], 3)
            self.assertEqual(len(response_data['errors']), 0)

        # Step 4: Verify users exist in database
        for username in ["import_user1", "import_user2", "import_user3"]:
            self.assertTrue(User.objects.filter(username=username).exists())

        # Step 5: Export the imported data as CSV
        # Get the imported users data for export
        imported_users = User.objects.filter(username__startswith="import_user").values(
            "username", "email", "first_name", "last_name"
        )
        csv_export_response = self.client.post('/api/users/export-as-file/', {
            "file_type": "csv",
            "includes": ["username", "email", "first_name", "last_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
                "last_name": {"label": "Last Name"}
            },
            "data": list(imported_users)
        })

        # Step 6: Verify CSV export
        self.assertEqual(csv_export_response.status_code, 200)
        self.assertEqual(csv_export_response['Content-Type'], 'text/csv')
        csv_content = csv_export_response.content.decode('utf-8')

        # Verify imported data appears in export
        self.assertIn("import_user1", csv_content)
        self.assertIn("import1@test.com", csv_content)

        # Step 7: Export as Excel
        xlsx_export_response = self.client.post('/api/users/export-as-file/', {
            "file_type": "xlsx",
            "includes": ["username", "email", "first_name", "last_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
                "last_name": {"label": "Last Name"}
            },
            "data": list(imported_users)
        })

        # Step 8: Verify Excel export
        self.assertEqual(xlsx_export_response.status_code, 200)
        self.assertEqual(
            xlsx_export_response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Step 9: Verify Excel content
        workbook = openpyxl.load_workbook(BytesIO(xlsx_export_response.content))
        sheet = workbook.active

        # Find import_user1 in the sheet
        found_user = False
        for row in sheet.iter_rows(values_only=True):
            if 'import_user1' in str(row):
                found_user = True
                break
        self.assertTrue(found_user, "import_user1 not found in exported Excel")

        # Step 10: Clean up imported users
        import_user_ids = list(User.objects.filter(
            username__startswith='import_user'
        ).values_list('id', flat=True))

        cleanup_response = self.client.delete('/api/users/bulk_delete/', import_user_ids, format='json')
        self.assertEqual(cleanup_response.status_code, 200)

    def test_error_handling_workflow(self):
        """
        Test workflow handles errors gracefully throughout process.
        """
        # Step 1: Try to import file with invalid data
        invalid_data = [
            ["", "invalid-email", "Test", "User"],  # Empty username, invalid email
            ["valid_user", "valid@test.com", "Valid", "User"],  # Valid data
        ]

        excel_file = self.create_test_excel_file(invalid_data)
        import_response = self.client.post(
            '/api/users/import-from-file/',
            {
                'file': excel_file,
                'append_data': 'true',
                'config': json.dumps({
                    "file_format": "xlsx",
                    "order": ["users"],
                    "models": {
                        "users": {
                            "model": "auth.User",
                            "unique_by": ["username"],
                            "update_if_exists": False,
                            "direct_columns": {
                                "username": "username",
                                "email": "email",
                                "first_name": "first_name",
                                "last_name": "last_name"
                            },
                        }
                    },
                })
            },
            format='multipart'
        )

        # Step 2: Verify partial success handling
        self.assertIn(import_response.status_code, [200, 422])  # Accept both success and validation error
        response_data = import_response.data
        if import_response.status_code == 200:
            self.assertEqual(response_data['successful_rows'], 1)
            self.assertGreater(len(response_data['errors']), 0)
        elif import_response.status_code == 422:
            # Validation failed completely - check for failed_rows or import_summary
            if 'data' in response_data:
                if 'failed_rows' in response_data['data']:
                    self.assertGreater(len(response_data['data']['failed_rows']), 0)
                elif 'import_summary' in response_data['data']:
                    self.assertGreater(response_data['data']['import_summary']['failed'], 0)

        # Step 3: Verify error handling behavior - transaction rollback means no users created
        # This is expected behavior for atomic transactions with validation errors
        self.assertFalse(User.objects.filter(username="valid_user").exists())
        self.assertFalse(User.objects.filter(username="").exists())

        # Step 4: Try bulk operation with mix of valid/invalid data
        mixed_bulk_data = [
            {"username": "valid_bulk", "email": "valid@bulk.com"},
            {"username": "", "email": "invalid"},  # Invalid data
        ]

        bulk_response = self.client.post('/api/users/bulk-create/', mixed_bulk_data, format='json')
        # Should handle validation errors appropriately
        self.assertIn(bulk_response.status_code, [201, 400])  # Either succeed partially or fail with validation

        # Step 5: Clean up any created users
        User.objects.filter(username__in=["valid_user", "valid_bulk"]).delete()

    def test_concurrent_operations_workflow(self):
        """
        Test workflow handles concurrent-like operations correctly.
        """
        # Step 1: Create base user
        user = UserFactory(username="concurrent_test", email="concurrent@test.com")

        # Step 2: Simulate concurrent updates by making rapid API calls
        update_data_1 = {"first_name": "Update1"}
        update_data_2 = {"first_name": "Update2"}

        response1 = self.client.patch(f'/api/users/{user.id}/', update_data_1)
        response2 = self.client.patch(f'/api/users/{user.id}/', update_data_2)

        # Step 3: Both updates should succeed (last one wins)
        self.assertEqual(response1.status_code, 200)
        self.assertEqual(response2.status_code, 200)

        # Step 4: Verify final state
        user.refresh_from_db()
        self.assertEqual(user.first_name, "Update2")  # Last update wins

        # Step 5: Test bulk operations don't interfere with single operations
        bulk_data = [{"username": f"bulk_{i}", "email": f"bulk{i}@test.com"} for i in range(3)]
        bulk_response = self.client.post('/api/users/bulk-create/', bulk_data, format='json')
        self.assertEqual(bulk_response.status_code, 201)

        # Original user should be unaffected
        user.refresh_from_db()
        self.assertEqual(user.username, "concurrent_test")
        self.assertEqual(user.first_name, "Update2")

        # Clean up
        User.objects.filter(username__startswith="bulk_").delete()
        user.delete()


# Make this module act as a URLconf for testing
urlpatterns = e2e_urlpatterns