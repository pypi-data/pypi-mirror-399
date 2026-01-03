"""
ViewSet integration tests with real API calls.

Tests ViewSet mixins through DRF test client without mocking.
"""

import json
from io import BytesIO

from django.contrib.auth import get_user_model
from django.urls import path, include
from django.test import override_settings

from rest_framework import serializers, viewsets
from rest_framework.routers import DefaultRouter
from rest_framework.test import APITestCase

import openpyxl

from drf_commons.common_tests.factories import UserFactory
from drf_commons.serializers.base import BaseModelSerializer
from drf_commons.views.mixins import (
    BulkCreateModelMixin,
    BulkUpdateModelMixin,
    BulkDeleteModelMixin,
    FileImportMixin,
    FileExportMixin,
)
from drf_commons.views.mixins.crud import ListModelMixin

User = get_user_model()


class UserSerializer(serializers.ModelSerializer):
    """Test serializer for User model."""

    class Meta:
        model = User
        fields = ["id", "username", "email", "first_name", "last_name"]


class StandardUserViewSet(viewsets.ModelViewSet):
    """Standard ViewSet for individual CRUD operations."""

    queryset = User.objects.all()
    serializer_class = UserSerializer


class UserBulkSerializer(BaseModelSerializer):
    """Bulk serializer for User model with optimized operations."""

    class Meta(BaseModelSerializer.Meta):
        model = User
        fields = ["id", "username", "email", "first_name", "last_name"]


class BulkUserViewSet(
    viewsets.GenericViewSet,
    ListModelMixin,
    BulkCreateModelMixin,
    BulkUpdateModelMixin,
    BulkDeleteModelMixin,
):
    """ViewSet for bulk operations only."""

    queryset = User.objects.all()
    serializer_class = UserBulkSerializer


class ImportExportUserViewSet(
    viewsets.GenericViewSet,
    ListModelMixin,
    FileImportMixin,
    FileExportMixin,
):
    """ViewSet for import/export operations."""

    queryset = User.objects.all()
    serializer_class = UserSerializer

    # FileImportMixin configuration
    import_template_name = "user_import_template.xlsx"
    import_file_config = {
        "file_format": "xlsx",
        "order": ["users"],
        "models": {
            "users": {
                "model": "auth.User",
                "unique_by": ["username"],
                "update_if_exists": True,
                "direct_columns": {
                    "username": "username",
                    "email": "email",
                    "first_name": "first_name",
                    "last_name": "last_name"
                },
            }
        },
    }


# Test URLs configuration
router = DefaultRouter()
router.register(r'standard-users', StandardUserViewSet, basename='standard-user')
router.register(r'bulk-users', BulkUserViewSet, basename='bulk-user')
router.register(r'import-export-users', ImportExportUserViewSet, basename='import-export-user')

test_urlpatterns = [
    path('api/', include(router.urls)),
]


@override_settings(ROOT_URLCONF=__name__)
class StandardCRUDTests(APITestCase):
    """Test standard CRUD operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_standard_crud_operations(self):
        """Test standard CRUD operations work through API."""
        # CREATE
        create_data = {
            "username": "api_test_user",
            "email": "api@test.com",
            "first_name": "API",
            "last_name": "Test"
        }
        create_response = self.client.post('/api/standard-users/', create_data)
        self.assertEqual(create_response.status_code, 201)
        created_user_id = create_response.data['id']

        # READ (List)
        list_response = self.client.get('/api/standard-users/')
        self.assertEqual(list_response.status_code, 200)
        self.assertGreaterEqual(len(list_response.data), 1)

        # READ (Detail)
        detail_response = self.client.get(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.data['username'], 'api_test_user')

        # UPDATE
        update_data = {"email": "updated@test.com"}
        update_response = self.client.patch(f'/api/standard-users/{created_user_id}/', update_data)
        self.assertEqual(update_response.status_code, 200)
        self.assertEqual(update_response.data['email'], 'updated@test.com')

        # DELETE
        delete_response = self.client.delete(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(delete_response.status_code, 204)

        # Verify deletion
        verify_response = self.client.get(f'/api/standard-users/{created_user_id}/')
        self.assertEqual(verify_response.status_code, 404)


@override_settings(ROOT_URLCONF=__name__)
class BulkOperationTests(APITestCase):
    """Test bulk operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_bulk_create_operation(self):
        """Test bulk create operation through API."""
        bulk_data = [
            {
                "username": "bulk_user1",
                "email": "bulk1@test.com",
                "first_name": "Bulk",
                "last_name": "One"
            },
            {
                "username": "bulk_user2",
                "email": "bulk2@test.com",
                "first_name": "Bulk",
                "last_name": "Two"
            }
        ]

        response = self.client.post('/api/bulk-users/bulk-create/', bulk_data, format='json')
        self.assertEqual(response.status_code, 201)

        # Verify users were created in database
        self.assertTrue(User.objects.filter(username="bulk_user1").exists())
        self.assertTrue(User.objects.filter(username="bulk_user2").exists())

    def test_bulk_update_operation(self):
        """Test bulk update operation through API."""
        # Create test users
        user1 = UserFactory(username="update_user1", email="old1@test.com")
        user2 = UserFactory(username="update_user2", email="old2@test.com")

        bulk_update_data = [
            {"id": user1.id, "email": "new1@test.com"},
            {"id": user2.id, "email": "new2@test.com"}
        ]

        response = self.client.put('/api/bulk-users/bulk-update/', bulk_update_data, format='json')
        self.assertEqual(response.status_code, 200)

        # Verify updates in database
        user1.refresh_from_db()
        user2.refresh_from_db()
        self.assertEqual(user1.email, "new1@test.com")
        self.assertEqual(user2.email, "new2@test.com")

    def test_bulk_delete_operation(self):
        """Test bulk delete operation through API."""
        # Create test users
        user1 = UserFactory(username="delete_user1")
        user2 = UserFactory(username="delete_user2")
        user3 = UserFactory(username="keep_user")

        delete_ids = [user1.id, user2.id]
        response = self.client.delete('/api/bulk-users/bulk_delete/', delete_ids, format='json')
        self.assertEqual(response.status_code, 200)

        # Verify deletions in database
        self.assertFalse(User.objects.filter(id=user1.id).exists())
        self.assertFalse(User.objects.filter(id=user2.id).exists())
        self.assertTrue(User.objects.filter(id=user3.id).exists())


@override_settings(ROOT_URLCONF=__name__)
class ImportExportTests(APITestCase):
    """Test import/export operations."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def create_test_excel_file(self, data_rows):
        """Create Excel file for import testing."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "users"

        # Add title rows (service expects header on row 5)
        sheet.cell(row=1, column=1, value="User Import Template")
        sheet.cell(row=2, column=1, value="Fill in the data below")
        sheet.cell(row=3, column=1, value="")
        sheet.cell(row=4, column=1, value="")

        # Headers on row 5
        headers = ["username", "email", "first_name", "last_name"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=5, column=col, value=header)

        # Data rows starting from row 6
        for row_idx, row_data in enumerate(data_rows, 6):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        file_buffer = BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)
        return file_buffer

    def test_file_import_operation(self):
        """Test file import operation through API."""
        test_data = [
            ["import_user1", "import1@test.com", "Import", "One"],
            ["import_user2", "import2@test.com", "Import", "Two"],
        ]

        excel_file = self.create_test_excel_file(test_data)

        response = self.client.post(
            '/api/import-export-users/import-from-file/',
            {
                'file': excel_file,
                'append_data': 'true',
                'config': json.dumps({
                    "file_format": "xlsx",
                    "order": ["users"],
                    "models": {
                        "users": {
                            "model": "auth.User",
                            "unique_by": ["username"],
                            "update_if_exists": False,
                            "direct_columns": {
                                "username": "username",
                                "email": "email",
                                "first_name": "first_name",
                                "last_name": "last_name"
                            },
                        }
                    },
                })
            },
            format='multipart'
        )
        self.assertIn(response.status_code, [200, 201])  # Accept both 200 and 201 for successful import

        # Check the import summary from the response
        if hasattr(response, 'data'):
            if 'data' in response.data and 'import_summary' in response.data['data']:
                summary = response.data['data']['import_summary']
                self.assertEqual(summary['created'], 2)
                self.assertEqual(summary['failed'], 0)
            elif 'successful_rows' in response.data:
                self.assertEqual(response.data['successful_rows'], 2)
                self.assertEqual(len(response.data['errors']), 0)

        # Verify users were imported
        self.assertTrue(User.objects.filter(username="import_user1").exists())
        self.assertTrue(User.objects.filter(username="import_user2").exists())

    def test_file_export_csv_operation(self):
        """Test CSV export operation through API."""
        # Create test data
        user1 = UserFactory(username="export_user1", email="export1@test.com", first_name="Export")
        user2 = UserFactory(username="export_user2", email="export2@test.com", first_name="Test")

        export_data = {
            "file_type": "csv",
            "includes": ["username", "email", "first_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"}
            },
            "data": [
                {"username": user1.username, "email": user1.email, "first_name": user1.first_name},
                {"username": user2.username, "email": user2.email, "first_name": user2.first_name}
            ]
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')

        # Verify CSV content
        content = response.content.decode('utf-8')
        self.assertIn("Username,Email,First Name", content)
        self.assertIn("export_user1,export1@test.com,Export", content)

    def test_file_export_xlsx_operation(self):
        """Test Excel export operation through API."""
        user = UserFactory(username="xlsx_user", email="xlsx@test.com", first_name="Excel")

        export_data = {
            "file_type": "xlsx",
            "includes": ["username", "email", "first_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"}
            },
            "data": [
                {"username": user.username, "email": user.email, "first_name": user.first_name}
            ]
        }

        response = self.client.post('/api/import-export-users/export-as-file/', export_data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Verify Excel content can be loaded
        workbook = openpyxl.load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertIsNotNone(sheet.cell(row=1, column=1).value)  # Has content


@override_settings(ROOT_URLCONF=__name__)
class CombinedWorkflowTests(APITestCase):
    """Test workflows combining different operation types."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        from django.conf import settings
        settings.ROOT_URLCONF = __name__

    def setUp(self):
        self.user = UserFactory()
        self.client.force_authenticate(user=self.user)

    def test_combined_operations_workflow(self):
        """Test workflow combining bulk and export operations."""
        # 1. Bulk create initial data
        initial_data = [
            {"username": "workflow1", "email": "flow1@test.com", "first_name": "Work"},
            {"username": "workflow2", "email": "flow2@test.com", "first_name": "Flow"}
        ]
        create_response = self.client.post('/api/bulk-users/bulk-create/', initial_data, format='json')
        self.assertEqual(create_response.status_code, 201)

        # 2. List to verify creation
        list_response = self.client.get('/api/bulk-users/')
        # Extract users from the response data structure
        response_data = list_response.data.get('data', {})
        results = response_data.get('results', [])
        workflow_users = [
            user for user in results
            if user['username'].startswith('workflow')
        ]
        self.assertEqual(len(workflow_users), 2)

        # 3. Bulk update the created users
        update_data = [
            {"id": workflow_users[0]['id'], "last_name": "Updated"},
            {"id": workflow_users[1]['id'], "last_name": "Updated"}
        ]
        update_response = self.client.put('/api/bulk-users/bulk-update/', update_data, format='json')
        self.assertEqual(update_response.status_code, 200)

        # 4. Export the updated data
        export_data = [
            {
                "username": workflow_users[0]['username'],
                "email": workflow_users[0]['email'],
                "first_name": workflow_users[0]['first_name'],
                "last_name": "Updated"
            },
            {
                "username": workflow_users[1]['username'],
                "email": workflow_users[1]['email'],
                "first_name": workflow_users[1]['first_name'],
                "last_name": "Updated"
            }
        ]
        export_response = self.client.post('/api/import-export-users/export-as-file/', {
            "file_type": "csv",
            "includes": ["username", "email", "first_name", "last_name"],
            "column_config": {
                "username": {"label": "Username"},
                "email": {"label": "Email"},
                "first_name": {"label": "First Name"},
                "last_name": {"label": "Last Name"}
            },
            "data": export_data
        })
        self.assertEqual(export_response.status_code, 200)

        # Verify export contains updated data
        content = export_response.content.decode('utf-8')
        self.assertIn("workflow1", content)
        self.assertIn("Updated", content)

        # 5. Clean up with bulk delete
        delete_ids = [user['id'] for user in workflow_users]
        delete_response = self.client.delete('/api/bulk-users/bulk_delete/', delete_ids, format='json')
        self.assertEqual(delete_response.status_code, 200)

        # Verify cleanup
        final_list = self.client.get('/api/bulk-users/')
        final_response_data = final_list.data.get('data', {})
        final_results = final_response_data.get('results', [])
        final_workflow_users = [
            user for user in final_results
            if user['username'].startswith('workflow')
        ]
        self.assertEqual(len(final_workflow_users), 0)


# Make this module act as a URLconf for testing
urlpatterns = test_urlpatterns