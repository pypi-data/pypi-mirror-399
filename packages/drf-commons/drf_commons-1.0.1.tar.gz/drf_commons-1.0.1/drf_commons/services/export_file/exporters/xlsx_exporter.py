"""
XLSX export implementation.
"""

from typing import Dict, List

from django.http import HttpResponse

from drf_commons.common_conf import settings

from ..utils import get_column_alignment, get_column_label, get_working_date


class XLSXExporter:
    """Handles XLSX export operations."""

    def export(
        self,
        data_rows: List[Dict],
        includes: List[str],
        column_config: Dict[str, Dict],
        filename: str,
        export_headers: List[str],
        document_titles: List[str],
    ) -> HttpResponse:
        """Export data as Excel file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as e:
            raise ImportError(
                "XLSX export requires openpyxl. "
                "Install it with: pip install drf-commons[export]"
            ) from e

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        # Style definitions from settings
        header_color = settings.EXPORTED_DOCS_DEFAULT_TABLE_HEADER_COLOR

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color=header_color, end_color=header_color, fill_type="solid"
        )

        current_row = 1

        # Write document headers (top left)
        for header_line in export_headers:
            if header_line.strip():
                cell = ws.cell(row=current_row, column=1, value=header_line)
                cell.font = Font(
                    bold=True, size=settings.EXPORTED_DOCS_HEADER_FONT_SIZE
                )
                current_row += 1

        # Add spacing after headers if we have them
        if export_headers:
            current_row += 1

        # Write document titles (centered above table)
        title_font_size = settings.EXPORTED_DOCS_TITLE_FONT_SIZE
        for title in document_titles:
            if title.strip():
                # Merge cells for title to center it across all columns
                if len(includes) > 1:
                    ws.merge_cells(
                        f"A{current_row}:{chr(64 + len(includes))}{current_row}"
                    )
                cell = ws.cell(row=current_row, column=1, value=title)
                cell.font = Font(bold=True, size=title_font_size)
                cell.alignment = Alignment(horizontal="center")
                current_row += 1

        # Add spacing after titles if we have them
        if document_titles:
            current_row += 1

        # Write column headers
        headers = [get_column_label(field, column_config) for field in includes]
        for col_idx, (header, field_name) in enumerate(zip(headers, includes), 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

            # Apply column-specific alignment for header, fallback to center for headers
            col_align = get_column_alignment(field_name, column_config)
            excel_align = {"left": "left", "center": "center", "right": "right"}.get(
                col_align, "center"
            )
            cell.alignment = Alignment(horizontal=excel_align, vertical="center")

        # Start data from next row
        data_start_row = current_row + 1

        # Write data
        for row_idx, row in enumerate(data_rows, data_start_row):
            for col_idx, field_name in enumerate(includes, 1):
                value = row.get(field_name, "")
                # Handle None values
                cell_value = value if value is not None else ""
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                # Apply column-specific alignment for data
                col_align = get_column_alignment(field_name, column_config)
                excel_align = {
                    "left": "left",
                    "center": "center",
                    "right": "right",
                }.get(col_align, "left")
                cell.alignment = Alignment(horizontal=excel_align, vertical="center")

        # Add footer with working date
        footer_row = data_start_row + len(data_rows) + 1
        footer_cell = ws.cell(
            row=footer_row, column=len(includes), value=f"Date: {get_working_date()}"
        )
        footer_cell.font = Font(italic=True)
        footer_cell.alignment = Alignment(horizontal="right")

        # Auto-adjust column widths based on settings
        if settings.EXPORTED_DOCS_AUTO_COLUMN_WIDTH:
            max_width = settings.EXPORTED_DOCS_MAX_COLUMN_WIDTH
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(
                    length + 2, max_width
                )

        # Save to response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)

        return response
