#!/usr/bin/env python3
"""
Excel to Image Converter (Round-trip verifier)
Reconstructs an image from an Excel file created by image_to_excel.py
Use this to verify pixel-fidelity of the conversion.

Features:
  - Auto-crop to solid-fill bounds (ignores stray notes/formatting)
  - Configurable background for missing cells
  - Diff heatmap generation for fidelity analysis
"""

import argparse
from pathlib import Path
import openpyxl
import numpy as np
from PIL import Image


def _cell_to_rgb(cell, background=(0, 0, 0)):
    """
    Extract RGB from an openpyxl cell fill.
    Returns (rgb_tuple, is_missing) where is_missing is True if no usable solid fill.
    """
    fill = cell.fill
    if not fill or fill.patternType != "solid":
        return background, True

    # Prefer fgColor, fall back to start_color
    color = None
    if fill.fgColor is not None:
        color = fill.fgColor
    if (color is None or not getattr(color, "rgb", None)) and fill.start_color is not None:
        color = fill.start_color

    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str):
        return background, True

    rgb = rgb.strip().upper()

    # openpyxl commonly stores ARGB (8 chars). Accept RGB (6) too.
    if len(rgb) == 8:
        rgb = rgb[2:]  # strip alpha
    if len(rgb) != 6:
        return background, True

    try:
        return (int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)), False
    except ValueError:
        return background, True


def _find_solid_bounds(ws, background=(0, 0, 0)):
    """
    Return (min_row, min_col, max_row, max_col) of cells that have valid pixel fills.
    Only considers cells with usable RGB values, not just any solid fill.
    If none found, return None.
    """
    min_r = min_c = None
    max_r = max_c = None

    for row in ws.iter_rows():
        for cell in row:
            rgb, is_missing = _cell_to_rgb(cell, background=background)
            if not is_missing:
                r, c = cell.row, cell.column
                if min_r is None:
                    min_r = max_r = r
                    min_c = max_c = c
                else:
                    if r < min_r: min_r = r
                    if c < min_c: min_c = c
                    if r > max_r: max_r = r
                    if c > max_c: max_c = c

    return None if min_r is None else (min_r, min_c, max_r, max_c)


def excel_to_image(excel_path: str, output_path: str = None, background=(0, 0, 0), autocrop=True, sheet_name=None):
    """
    Reconstruct an image from an Excel file with coloured cells.

    Args:
        excel_path: Path to Excel file created by image_to_excel.py
        output_path: Path for output image (default: same name with .png)
        background: RGB tuple for cells without a solid fill
        autocrop: If True, crop to solid-fill bounds (ignores stray formatting)
        sheet_name: Specific worksheet to use (default: 'Image' if present, else active)
    """
    wb = openpyxl.load_workbook(excel_path)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb["Image"] if "Image" in wb.sheetnames else wb.active

    if autocrop:
        bounds = _find_solid_bounds(ws, background=background)
        if bounds is None:
            raise ValueError("No solid-filled cells found to reconstruct an image.")
        min_r, min_c, max_r, max_c = bounds
    else:
        min_r, min_c, max_r, max_c = 1, 1, ws.max_row, ws.max_column

    h = max_r - min_r + 1
    w = max_c - min_c + 1
    img = np.zeros((h, w, 3), dtype=np.uint8)

    missing = 0
    for rr, r in enumerate(range(min_r, max_r + 1)):
        for cc, c in enumerate(range(min_c, max_c + 1)):
            cell = ws.cell(r, c)
            rgb, is_missing = _cell_to_rgb(cell, background=background)
            if is_missing:
                missing += 1
            img[rr, cc] = rgb

    if output_path is None:
        output_path = str(Path(excel_path).with_suffix("")) + "_reconstructed.png"

    Image.fromarray(img).save(output_path)
    print(f"✓ Reconstructed: {output_path}")
    print(f"  Dimensions: {w}x{h} pixels")
    print(f"  Missing (non-solid fill) cells: {missing}")
    if autocrop:
        print(f"  Crop bounds: rows {min_r}-{max_r}, cols {min_c}-{max_c}")

    return output_path


def compare_images(original_path: str, reconstructed_path: str, diff_output: str = None, diff_threshold: int = None):
    """
    Compare original and reconstructed images for fidelity check.
    Optionally saves a diff heatmap showing per-pixel differences.
    
    Args:
        original_path: Path to original image
        reconstructed_path: Path to reconstructed image
        diff_output: Optional path to save diff heatmap
        diff_threshold: Optional threshold for pass/fail pixel count
    """
    orig = Image.open(original_path).convert("RGB")
    recon = Image.open(reconstructed_path).convert("RGB")

    # Resize original to match reconstructed for comparison
    orig_resized = orig.resize(recon.size, Image.Resampling.LANCZOS)

    orig_arr = np.array(orig_resized, dtype=np.int16)
    recon_arr = np.array(recon, dtype=np.int16)

    diff = np.abs(orig_arr - recon_arr)
    max_diff = int(diff.max())
    mean_diff = float(diff.mean())

    print("\n📊 Fidelity Report:")
    print(f"  Max pixel difference: {max_diff}")
    print(f"  Mean pixel difference: {mean_diff:.2f}")

    if max_diff == 0:
        print("  ✅ PERFECT: Pixel-perfect reconstruction!")
    elif mean_diff < 1:
        print("  ✅ EXCELLENT: Near-perfect fidelity")
    elif mean_diff < 5:
        print("  🟡 GOOD: Minor differences (likely from resizing)")
    else:
        print("  🟠 FAIR: Noticeable differences")

    # Threshold analysis for CI pass/fail
    if diff_threshold is not None:
        # Per-pixel max channel difference
        pixel_max_diff = diff.max(axis=2)
        total_pixels = pixel_max_diff.size
        pixels_above = int((pixel_max_diff > diff_threshold).sum())
        pct_above = (pixels_above / total_pixels) * 100
        
        print(f"\n🎯 Threshold Analysis (>{diff_threshold}):")
        print(f"  Pixels above threshold: {pixels_above:,} / {total_pixels:,} ({pct_above:.2f}%)")
        
        if pixels_above == 0:
            print("  ✅ PASS: All pixels within threshold")
        else:
            print(f"  ❌ FAIL: {pixels_above:,} pixels exceed threshold")

    # Generate diff heatmap if requested
    if diff_output:
        # Sum across RGB channels and normalize to 0-255
        diff_magnitude = diff.sum(axis=2).astype(np.float32)
        if diff_magnitude.max() > 0:
            diff_normalized = (diff_magnitude / diff_magnitude.max() * 255).astype(np.uint8)
        else:
            diff_normalized = diff_magnitude.astype(np.uint8)
        
        # Apply colormap: black (no diff) -> red -> yellow -> white (max diff)
        # Cast to int16 to prevent uint8 underflow on subtraction
        dn = diff_normalized.astype(np.int16)
        heatmap = np.zeros((diff_normalized.shape[0], diff_normalized.shape[1], 3), dtype=np.uint8)
        heatmap[:, :, 0] = np.clip(dn * 2, 0, 255).astype(np.uint8)  # Red
        heatmap[:, :, 1] = np.clip((dn - 128) * 2, 0, 255).astype(np.uint8)  # Green (high diffs)
        heatmap[:, :, 2] = np.clip((dn - 200) * 4, 0, 255).astype(np.uint8)  # Blue (very high diffs)
        
        Image.fromarray(heatmap).save(diff_output)
        print(f"  Diff heatmap saved: {diff_output}")


def parse_rgb(value: str) -> tuple:
    """Parse R,G,B string into tuple."""
    try:
        parts = [int(x.strip()) for x in value.split(",")]
        if len(parts) != 3 or not all(0 <= p <= 255 for p in parts):
            raise ValueError
        return tuple(parts)
    except:
        raise argparse.ArgumentTypeError(f"Invalid RGB format: '{value}'. Use R,G,B (e.g., 255,0,128)")


def main():
    parser = argparse.ArgumentParser(
        description="Reconstruct an image from an Excel file with coloured cells.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  xlsx2image image.xlsx
  xlsx2image image.xlsx -o output.png
  xlsx2image image.xlsx --sheet Sheet1
  xlsx2image image.xlsx --compare original.png
  xlsx2image image.xlsx --compare original.png --diff-image diff.png
  xlsx2image image.xlsx --compare original.png --diff-threshold 5
  xlsx2image image.xlsx --background 255,255,255 --no-autocrop
        """
    )
    
    parser.add_argument("excel_file", help="Excel file to convert")
    parser.add_argument("-o", "--output", help="Output image path (default: <input>_reconstructed.png)")
    parser.add_argument("--sheet", help="Worksheet name to use (default: 'Image' if present, else active sheet)")
    parser.add_argument("--compare", metavar="ORIGINAL", help="Compare with original image for fidelity check")
    parser.add_argument("--diff-image", metavar="PATH", help="Save diff heatmap image (requires --compare)")
    parser.add_argument("--diff-threshold", type=int, metavar="N",
                        help="Report pixels with diff > N (requires --compare). Useful for CI pass/fail.")
    parser.add_argument("--background", type=parse_rgb, default=(0, 0, 0),
                        help="Background RGB for missing cells (default: 0,0,0)")
    parser.add_argument("--no-autocrop", action="store_true",
                        help="Disable auto-crop to solid-fill bounds")
    
    args = parser.parse_args()
    
    if args.diff_image and not args.compare:
        parser.error("--diff-image requires --compare")
    if args.diff_threshold is not None and not args.compare:
        parser.error("--diff-threshold requires --compare")
    
    reconstructed = excel_to_image(
        args.excel_file,
        args.output,
        background=args.background,
        autocrop=not args.no_autocrop,
        sheet_name=args.sheet
    )
    
    if args.compare:
        compare_images(args.compare, reconstructed, args.diff_image, args.diff_threshold)


if __name__ == "__main__":
    main()
