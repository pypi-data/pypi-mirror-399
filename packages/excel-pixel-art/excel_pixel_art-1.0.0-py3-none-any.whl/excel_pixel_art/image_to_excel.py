#!/usr/bin/env python3
"""
Image to Excel Converter
Converts an image into an Excel file where each cell represents a pixel,
with cell background colours matching the original image.
"""

import argparse
from pathlib import Path
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def rgb_to_argb(r, g, b):
    """Convert RGB values to ARGB hex string for Excel (FF = full opacity)."""
    return f"FF{r:02X}{g:02X}{b:02X}"


def image_to_excel(image_path: str, output_path: str, max_size: int = 100):
    """
    Convert an image to an Excel file with coloured cells.
    
    Args:
        image_path: Path to input image
        output_path: Path for output Excel file
        max_size: Maximum dimension (width or height) in cells. 
                  Larger = more detail but bigger file. Default 100.
    """
    img = Image.open(image_path).convert('RGB')
    
    # Calculate resize dimensions maintaining aspect ratio
    width, height = img.size
    if width > height:
        new_width = min(max_size, width)
        new_height = int(height * (new_width / width))
    else:
        new_height = min(max_size, height)
        new_width = int(width * (new_height / height))
    
    # Resize image for reasonable Excel output
    img_resized = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
    pixels = img_resized.load()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Image"
    
    # Set cell dimensions to make them square
    cell_size = 2.5  # Excel column width units
    row_height = 15  # Points
    
    # Set column widths
    for col in range(1, new_width + 1):
        ws.column_dimensions[get_column_letter(col)].width = cell_size
    
    # Cache PatternFill objects by colour for performance
    fill_cache = {}
    
    # Set row heights and fill cells with colours
    for row in range(1, new_height + 1):
        ws.row_dimensions[row].height = row_height
        for col in range(1, new_width + 1):
            r, g, b = pixels[col - 1, row - 1]
            argb = rgb_to_argb(r, g, b)
            
            # Reuse cached fill or create new one
            if argb not in fill_cache:
                fill_cache[argb] = PatternFill(start_color=argb, end_color=argb, fill_type='solid')
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_cache[argb]
    
    unique_colours = len(fill_cache)
    print(f"  Unique colours: {unique_colours} (cached for performance)")
    
    # Add metadata sheet with info
    info_sheet = wb.create_sheet("Info")
    info_sheet['A1'] = "Original Image"
    info_sheet['B1'] = Path(image_path).name
    info_sheet['A2'] = "Original Size"
    info_sheet['B2'] = f"{width} x {height}"
    info_sheet['A3'] = "Excel Size"
    info_sheet['B3'] = f"{new_width} x {new_height} cells"
    info_sheet['A4'] = "Max Size Setting"
    info_sheet['B4'] = max_size
    info_sheet.column_dimensions['A'].width = 20
    info_sheet.column_dimensions['B'].width = 30
    
    wb.save(output_path)
    print(f"✓ Created: {output_path}")
    print(f"  Original: {width}x{height} → Excel: {new_width}x{new_height} cells")


def main():
    parser = argparse.ArgumentParser(
        description="Convert an image to an Excel file with coloured cells (pixel art).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  image2xlsx photo.png
  image2xlsx photo.png -o output.xlsx
  image2xlsx photo.png --max-size 150
  image2xlsx photo.png -o art.xlsx --max-size 80
        """
    )
    
    parser.add_argument("image", help="Input image file")
    parser.add_argument("-o", "--output", help="Output Excel file (default: <input>_excel.xlsx)")
    parser.add_argument("--max-size", type=int, default=100,
                        help="Maximum dimension in cells (default: 100)")
    
    args = parser.parse_args()
    
    output_path = args.output or (Path(args.image).stem + "_excel.xlsx")
    image_to_excel(args.image, output_path, args.max_size)


if __name__ == "__main__":
    main()
