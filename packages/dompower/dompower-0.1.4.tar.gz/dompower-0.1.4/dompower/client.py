"""Main client for the dompower library."""

from __future__ import annotations

import base64
import io
import json
import logging
from datetime import date, datetime
from typing import TYPE_CHECKING, Any
from zoneinfo import ZoneInfo

from .auth import TokenManager, TokenUpdateCallback
from .const import (
    BASE_URL,
    DEFAULT_HEADERS,
    ENDPOINT_BILL_FORECAST,
    ENDPOINT_GET_BP_NUMBER,
    ENDPOINT_GET_BUSINESS_MASTER,
    ENDPOINT_USAGE_EXCEL,
)
from .exceptions import (
    ApiError,
    CannotConnectError,
    InvalidAuthError,
    RateLimitError,
)
from .models import (
    AccountInfo,
    BillForecast,
    BillPeriodData,
    CustomerInfo,
    IntervalUsageData,
    MeterDevice,
    ServiceAddress,
    ServiceType,
)

if TYPE_CHECKING:
    from aiohttp import ClientSession

_LOGGER = logging.getLogger(__name__)

# Dominion Energy operates exclusively in Eastern time zone
DOMINION_TIMEZONE = ZoneInfo("America/New_York")


class DompowerClient:
    """Async client for the Dominion Energy API.

    This client handles authentication, token refresh, and API requests
    for retrieving energy usage data from Dominion Energy.

    Example usage:
        async with aiohttp.ClientSession() as session:
            client = DompowerClient(
                session,
                access_token="...",
                refresh_token="...",
                token_update_callback=save_tokens,
            )
            usage = await client.async_get_interval_usage(
                account_number="123456",
                meter_number="789",
                start_date=date(2024, 1, 1),
                end_date=date(2024, 1, 31),
            )
    """

    def __init__(
        self,
        session: ClientSession,
        *,
        access_token: str | None = None,
        refresh_token: str | None = None,
        token_update_callback: TokenUpdateCallback | None = None,
    ) -> None:
        """Initialize the client.

        Args:
            session: aiohttp ClientSession (owned by caller, not closed by client).
            access_token: Initial access token from previous authentication.
            refresh_token: Initial refresh token from previous authentication.
            token_update_callback: Called when tokens are refreshed with signature
                (access_token: str, refresh_token: str) -> None.
                Use this to persist new tokens (e.g., in Home Assistant config).
        """
        self._session = session
        self._token_manager = TokenManager(
            session,
            access_token=access_token,
            refresh_token=refresh_token,
            token_update_callback=token_update_callback,
        )

    async def __aenter__(self) -> DompowerClient:
        """Enter async context manager."""
        return self

    async def __aexit__(self, *args: object) -> None:
        """Exit async context manager.

        Note: Does NOT close the session - caller owns the session lifecycle.
        """
        pass

    @property
    def has_tokens(self) -> bool:
        """Check if the client has tokens configured."""
        return self._token_manager.has_tokens

    async def async_set_tokens(
        self,
        access_token: str,
        refresh_token: str,
    ) -> None:
        """Set authentication tokens.

        Args:
            access_token: Access token from authentication.
            refresh_token: Refresh token from authentication.
        """
        self._token_manager.set_tokens(access_token, refresh_token)

    async def async_refresh_tokens(self) -> None:
        """Manually refresh authentication tokens.

        This is typically called automatically when needed, but can be
        called manually if desired.

        Raises:
            InvalidAuthError: If no tokens are available.
            TokenExpiredError: If refresh token is invalid/expired.
        """
        await self._token_manager.async_refresh_tokens()

    async def async_login(
        self,
        username: str,
        password: str,
        tfa_code_callback: Any = None,
        cookie_file: Any = None,
    ) -> None:
        """Authenticate with username and password.

        This method uses the Gigya authentication flow with TFA support.
        After successful login, tokens are automatically set on this client.

        Args:
            username: Dominion Energy email address.
            password: Account password.
            tfa_code_callback: Async callback to get TFA code from user.
                Signature: async (target: TFATarget) -> str
            cookie_file: Optional Path to persist cookies for TFA bypass.

        Raises:
            InvalidCredentialsError: Wrong username or password.
            TFARequiredError: TFA required but no callback provided.
            TFAVerificationError: TFA code verification failed.
        """
        from .gigya_auth import GigyaAuthenticator

        auth = GigyaAuthenticator(self._session, cookie_file=cookie_file)
        tokens = await auth.async_login(
            username=username,
            password=password,
            tfa_code_callback=tfa_code_callback,
        )

        self._token_manager.set_tokens(tokens.access_token, tokens.refresh_token)

    async def _async_request(
        self,
        method: str,
        endpoint: str,
        *,
        json_data: dict[str, Any] | None = None,
        params: dict[str, Any] | None = None,
        expect_binary: bool = False,
    ) -> dict[str, Any] | bytes:
        """Make an authenticated API request.

        Args:
            method: HTTP method (GET, POST, etc.).
            endpoint: API endpoint path.
            json_data: JSON body for the request.
            params: Query parameters.
            expect_binary: If True, return raw bytes instead of JSON.

        Returns:
            JSON response as dict, or bytes if expect_binary is True.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
            CannotConnectError: If unable to connect.
        """
        token = await self._token_manager.async_ensure_valid_token()

        url = f"{BASE_URL}{endpoint}"
        headers = {
            **DEFAULT_HEADERS,
            "Authorization": f"Bearer {token}",
        }

        _LOGGER.debug("Making %s request to %s", method, endpoint)

        try:
            async with self._session.request(
                method,
                url,
                headers=headers,
                json=json_data,
                params=params,
            ) as response:
                # Handle 401 - try refresh and retry once
                if response.status == 401:
                    _LOGGER.debug("Got 401, refreshing token and retrying")
                    new_tokens = await self._token_manager.async_refresh_tokens()
                    headers["Authorization"] = f"Bearer {new_tokens.access_token}"

                    async with self._session.request(
                        method,
                        url,
                        headers=headers,
                        json=json_data,
                        params=params,
                    ) as retry_response:
                        return await self._handle_response(
                            retry_response, expect_binary
                        )

                return await self._handle_response(response, expect_binary)

        except (InvalidAuthError, ApiError, RateLimitError):
            raise
        except Exception as err:
            raise CannotConnectError(f"Failed to connect to API: {err}") from err

    async def _handle_response(
        self,
        response: object,
        expect_binary: bool,
    ) -> dict[str, Any] | bytes:
        """Handle API response.

        Args:
            response: aiohttp response object.
            expect_binary: If True, return raw bytes.

        Returns:
            Parsed JSON or raw bytes.

        Raises:
            ApiError: If the response indicates an error.
            RateLimitError: If rate limited.
        """
        # Type hint workaround for aiohttp response
        from aiohttp import ClientResponse

        resp: ClientResponse = response  # type: ignore[assignment]

        if resp.status == 429:
            retry_after = resp.headers.get("Retry-After")
            raise RateLimitError(
                "Rate limited by API",
                status_code=429,
                retry_after=int(retry_after) if retry_after else None,
            )

        if resp.status >= 400:
            text = await resp.text()
            raise ApiError(
                f"API error: {resp.status}",
                status_code=resp.status,
                response_text=text,
            )

        if expect_binary:
            return await resp.read()

        data: dict[str, Any] = await resp.json()

        # Check for API-level errors in response body
        status = data.get("status", {})
        if isinstance(status, dict) and status.get("error"):
            raise ApiError(
                status.get("message", "Unknown API error"),
                api_code=status.get("code"),
                api_message=status.get("message"),
            )

        return data

    async def async_get_interval_usage(
        self,
        account_number: str,
        meter_number: str,
        start_date: date,
        end_date: date,
        *,
        service_type: ServiceType = ServiceType.ELECTRICITY,
    ) -> list[IntervalUsageData]:
        """Get 30-minute interval usage data.

        This fetches high-resolution usage data in 30-minute intervals,
        which is ideal for Home Assistant energy monitoring.

        Args:
            account_number: Dominion Energy account number.
            meter_number: Meter number for the account.
            start_date: Start date for the data range.
            end_date: End date for the data range.
            service_type: Type of service (electricity or gas).

        Returns:
            List of IntervalUsageData objects with 30-minute consumption data.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
        """
        payload = {
            "Format": "Csv",
            "accountNumber": account_number,
            "meterIds": [meter_number],
            "uom": "kWh",
            "periodicity": "HH",
            "serviceType": service_type.value,
            "decimalPlaces": "2",
            "isNetUsage": "false",
            "displayUnit": "kWh",
            "from": start_date.isoformat(),
            "to": end_date.isoformat(),
        }

        _LOGGER.debug(
            "Fetching interval usage for account %s from %s to %s",
            account_number,
            start_date,
            end_date,
        )

        response = await self._async_request(
            "POST",
            ENDPOINT_USAGE_EXCEL,
            json_data=payload,
            expect_binary=True,
        )

        # Parse the Excel response and filter to requested date range
        return self._parse_excel_usage(
            response,  # type: ignore[arg-type]
            start_date,
            end_date,
        )

    def _parse_excel_usage(
        self,
        excel_data: bytes,
        start_date: date,
        end_date: date,
    ) -> list[IntervalUsageData]:
        """Parse Excel usage data into IntervalUsageData objects.

        The Excel file is in wide format:
        - Row 1: Headers (Account No, Recorder ID, Date, 12:00 AM kWH, ...)
        - Row 2+: Data rows with date in column C and 48 half-hour readings

        Note: The API may return more data than requested, so we filter to
        the requested date range.

        Args:
            excel_data: Raw Excel file bytes.
            start_date: Start date for filtering results.
            end_date: End date for filtering results.

        Returns:
            List of IntervalUsageData objects sorted by timestamp,
            filtered to the requested date range.
        """
        try:
            import openpyxl
        except ImportError as err:
            raise ImportError(
                "openpyxl is required for parsing Excel usage data. "
                "Install it with: pip install openpyxl"
            ) from err

        workbook = openpyxl.load_workbook(io.BytesIO(excel_data), data_only=True)
        sheet = workbook.active

        if sheet is None:
            return []

        usage_data: list[IntervalUsageData] = []

        # Get header row to parse time slots
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # Parse time slots from headers (columns D onwards, index 3+)
        # Format: "12:00 AM kWH", "12:30 AM kWH", etc.
        time_slots: list[tuple[int, int]] = []  # (hour, minute) for each column
        for header in headers[3:]:  # Skip Account No, Recorder ID, Date
            if header is None or not isinstance(header, str):
                continue
            # Parse "HH:MM AM/PM kWH" format
            try:
                time_part = header.replace(" kWH", "").strip()
                parsed_time = datetime.strptime(time_part, "%I:%M %p")
                time_slots.append((parsed_time.hour, parsed_time.minute))
            except ValueError:
                _LOGGER.warning("Failed to parse time header: %s", header)
                continue

        # Process data rows (row 2 onwards)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue

            date_val = row[2]  # Date is in column C (index 2)

            # Parse the date
            if isinstance(date_val, datetime):
                row_date = date_val.date()
            elif isinstance(date_val, str):
                try:
                    row_date = datetime.strptime(date_val, "%m/%d/%Y").date()
                except ValueError:
                    _LOGGER.warning("Failed to parse date: %s", date_val)
                    continue
            else:
                continue

            # Process each time slot (columns D onwards, index 3+)
            for i, consumption_val in enumerate(row[3:]):
                if i >= len(time_slots):
                    break

                if consumption_val is None:
                    continue

                try:
                    consumption = float(consumption_val)
                except (ValueError, TypeError):
                    continue

                hour, minute = time_slots[i]
                timestamp = datetime(
                    row_date.year,
                    row_date.month,
                    row_date.day,
                    hour,
                    minute,
                    tzinfo=DOMINION_TIMEZONE,
                )

                usage_data.append(
                    IntervalUsageData(
                        timestamp=timestamp,
                        consumption=consumption,
                        unit="kWh",
                    )
                )

        # Filter to requested date range (API may return more data than requested)
        usage_data = [
            u for u in usage_data if start_date <= u.timestamp.date() <= end_date
        ]

        # Sort by timestamp
        usage_data.sort(key=lambda x: x.timestamp)

        _LOGGER.debug(
            "Parsed %d interval usage records (filtered to %s - %s)",
            len(usage_data),
            start_date,
            end_date,
        )

        return usage_data

    async def async_get_raw_excel(
        self,
        account_number: str,
        meter_number: str,
        start_date: date,
        end_date: date,
        *,
        service_type: ServiceType = ServiceType.ELECTRICITY,
    ) -> bytes:
        """Get raw Excel file with usage data.

        Use this if you need the original Excel file instead of parsed data.

        Args:
            account_number: Dominion Energy account number.
            meter_number: Meter number for the account.
            start_date: Start date for the data range.
            end_date: End date for the data range.
            service_type: Type of service (electricity or gas).

        Returns:
            Raw Excel file bytes.
        """
        payload = {
            "Format": "Csv",
            "accountNumber": account_number,
            "meterIds": [meter_number],
            "uom": "kWh",
            "periodicity": "HH",
            "serviceType": service_type.value,
            "decimalPlaces": "2",
            "isNetUsage": "false",
            "displayUnit": "kWh",
            "from": start_date.isoformat(),
            "to": end_date.isoformat(),
        }

        response = await self._async_request(
            "POST",
            ENDPOINT_USAGE_EXCEL,
            json_data=payload,
            expect_binary=True,
        )

        return response  # type: ignore[return-value]

    async def async_get_bill_forecast(
        self,
        account_number: str,
    ) -> BillForecast:
        """Get bill forecast including last bill and current period data.

        This fetches billing forecast data which includes the previous bill's
        charges and usage, allowing calculation of the effective rate.

        Args:
            account_number: Dominion Energy account number.

        Returns:
            BillForecast object with last bill data and current period info.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
        """
        _LOGGER.debug("Fetching bill forecast for account %s", account_number)

        response = await self._async_request(
            "GET",
            ENDPOINT_BILL_FORECAST,
            params={"accountNumber": account_number, "actionCode": "4"},
        )

        return self._parse_bill_forecast(response)  # type: ignore[arg-type]

    def _parse_bill_forecast(self, response: dict[str, Any]) -> BillForecast:
        """Parse bill forecast response into BillForecast object.

        Args:
            response: JSON response from the bill forecast API.

        Returns:
            Parsed BillForecast object.
        """
        data = response.get("data", {})

        # Parse last bill data
        last_bill_data = data.get("lastBill", {})
        last_bill = BillPeriodData(
            charges=float(last_bill_data.get("charges", 0)),
            usage=float(last_bill_data.get("usage", 0)),
            period_start=self._parse_date(last_bill_data.get("billPerdStDate")),
            period_end=self._parse_date(last_bill_data.get("billPerdEdDate")),
        )

        # Parse current period dates
        current_start = self._parse_date(data.get("billperdstdate"))
        current_end = self._parse_date(data.get("billperdeddate"))

        # Handle missing dates by using defaults
        if current_start is None:
            current_start = date.today()
        if current_end is None:
            current_end = date.today()

        return BillForecast(
            last_bill=last_bill,
            current_period_start=current_start,
            current_period_end=current_end,
            current_usage_kwh=float(data.get("currentUsageKwh", 0)),
            is_tou=data.get("tou", "N") == "Y",
        )

    def _parse_date(self, date_str: str | None) -> date | None:
        """Parse date string from API response.

        Args:
            date_str: Date string in format "MM/DD/YYYY HH:MM:SS" or None.

        Returns:
            Parsed date or None if parsing fails.
        """
        if not date_str:
            return None

        try:
            # API returns dates like "10/23/2025 00:00:00"
            dt = datetime.strptime(date_str, "%m/%d/%Y %H:%M:%S")
            return dt.date()
        except ValueError:
            _LOGGER.warning("Failed to parse date: %s", date_str)
            return None

    def _extract_uuid_from_token(self) -> str:
        """Extract UUID from the JWT access token payload.

        The Dominion API includes a UUID in the JWT claims that is required
        for certain API endpoints.

        Returns:
            The UUID from the token, or empty string if not found.

        Raises:
            InvalidAuthError: If no access token is available.
        """
        token = self._token_manager._access_token
        if not token:
            raise InvalidAuthError("No access token available")

        try:
            # JWT format: header.payload.signature
            parts = token.split(".")
            if len(parts) != 3:
                _LOGGER.warning("Invalid JWT format")
                return ""

            payload = parts[1]
            # Add padding if needed for base64 decoding
            padding = 4 - len(payload) % 4
            if padding != 4:
                payload += "=" * padding

            decoded = base64.urlsafe_b64decode(payload)
            claims = json.loads(decoded)
            uuid = str(claims.get("Uuid", ""))
            uuid_display = uuid[:8] + "..." if uuid else "none"
            _LOGGER.debug("Extracted UUID from token: %s", uuid_display)
            return uuid

        except Exception as err:
            _LOGGER.warning("Failed to extract UUID from token: %s", err)
            return ""

    async def async_get_customer_number(self) -> str:
        """Get the customer number (business partner ID) for the authenticated user.

        Returns:
            Customer number as a string.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
        """
        uuid = self._extract_uuid_from_token()
        if not uuid:
            raise ApiError("Could not extract UUID from token")

        _LOGGER.debug("Fetching customer number for UUID %s", uuid[:8] + "...")

        response = await self._async_request(
            "GET",
            ENDPOINT_GET_BP_NUMBER,
            params={"Uuid": uuid},
        )

        data = response.get("data", {})  # type: ignore[union-attr]
        customer_number = str(data.get("customerNumber", ""))

        if not customer_number:
            raise ApiError("Customer number not found in response")

        _LOGGER.debug("Got customer number: %s", customer_number)
        return customer_number

    async def async_get_accounts(
        self,
        *,
        include_inactive: bool = False,
    ) -> list[AccountInfo]:
        """Get all accounts for the authenticated user.

        Args:
            include_inactive: If True, include closed/inactive accounts.
                Defaults to False (active accounts only).

        Returns:
            List of AccountInfo objects with service addresses and meters.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
        """
        uuid = self._extract_uuid_from_token()
        if not uuid:
            raise ApiError("Could not extract UUID from token")

        customer_number = await self.async_get_customer_number()

        _LOGGER.debug("Fetching accounts for customer %s", customer_number)

        response = await self._async_request(
            "GET",
            ENDPOINT_GET_BUSINESS_MASTER,
            params={"customerNumber": customer_number, "uuid": uuid},
        )

        return self._parse_accounts_response(response, include_inactive)  # type: ignore[arg-type]

    async def async_get_customer_info(
        self,
        *,
        include_inactive: bool = False,
    ) -> CustomerInfo:
        """Get complete customer profile including all accounts.

        Convenience method that combines GetBpNumber and GetBusinessMaster
        to return a complete customer profile.

        Args:
            include_inactive: If True, include closed/inactive accounts.

        Returns:
            CustomerInfo with customer details and all accounts.

        Raises:
            InvalidAuthError: If not authenticated.
            ApiError: If the API returns an error.
        """
        uuid = self._extract_uuid_from_token()
        if not uuid:
            raise ApiError("Could not extract UUID from token")

        customer_number = await self.async_get_customer_number()

        _LOGGER.debug("Fetching full customer info for %s", customer_number)

        response = await self._async_request(
            "GET",
            ENDPOINT_GET_BUSINESS_MASTER,
            params={"customerNumber": customer_number, "uuid": uuid},
        )

        assert isinstance(response, dict)
        return self._parse_customer_info_response(
            response,
            customer_number,
            include_inactive,
        )

    def _parse_accounts_response(
        self,
        response: dict[str, Any],
        include_inactive: bool,
    ) -> list[AccountInfo]:
        """Parse GetBusinessMaster response into AccountInfo objects.

        Args:
            response: JSON response from the API.
            include_inactive: Whether to include inactive accounts.

        Returns:
            List of parsed AccountInfo objects.
        """
        accounts: list[AccountInfo] = []
        data_list = response.get("data", [])

        if not data_list:
            return accounts

        # Response is an array with customer data
        customer_data = data_list[0] if data_list else {}
        nav_data = customer_data.get("zbpMaintRegEnroll_nav", {})
        results = nav_data.get("results", [])

        for acct_data in results:
            is_active = acct_data.get("accountActive", "N") == "Y"

            # Skip inactive accounts unless requested
            if not include_inactive and not is_active:
                continue

            account = self._parse_single_account(acct_data)
            accounts.append(account)

        _LOGGER.debug("Parsed %d accounts", len(accounts))
        return accounts

    def _parse_customer_info_response(
        self,
        response: dict[str, Any],
        customer_number: str,
        include_inactive: bool,
    ) -> CustomerInfo:
        """Parse GetBusinessMaster response into CustomerInfo.

        Args:
            response: JSON response from the API.
            customer_number: Customer number from GetBpNumber.
            include_inactive: Whether to include inactive accounts.

        Returns:
            Parsed CustomerInfo object.
        """
        data_list = response.get("data", [])
        customer_data = data_list[0] if data_list else {}

        accounts = self._parse_accounts_response(response, include_inactive)

        return CustomerInfo(
            customer_number=customer_number,
            first_name=customer_data.get("firstName", ""),
            last_name=customer_data.get("lastName", ""),
            email=customer_data.get("primaryEmail", ""),
            accounts=accounts,
        )

    def _parse_single_account(self, acct_data: dict[str, Any]) -> AccountInfo:
        """Parse a single account from the API response.

        Args:
            acct_data: Account data dictionary from the API.

        Returns:
            Parsed AccountInfo object.
        """
        # Parse service address
        addr_data = acct_data.get("serviceAddress", {})
        service_address = ServiceAddress(
            street=addr_data.get("street", ""),
            house_number=addr_data.get("houseNum1", ""),
            city=addr_data.get("city", ""),
            state=addr_data.get("state", ""),
            zip_code=addr_data.get("zipCode", ""),
            country=addr_data.get("country", "US"),
        )

        # Parse meters/devices
        meters: list[MeterDevice] = []
        for dev_data in acct_data.get("conDev", []):
            net_metering_val = dev_data.get("netMetering")
            if net_metering_val is None:
                net_metering = None
            else:
                net_metering = net_metering_val == "Y"

            meter = MeterDevice(
                device_id=dev_data.get("device", ""),
                contract_id=dev_data.get("contract", ""),
                is_active=dev_data.get("isActive", "N") == "Y",
                has_ami=dev_data.get("amiFlag", False),
                net_metering=net_metering,
            )
            meters.append(meter)

        # Parse closing date if present
        closing_date = self._parse_date(acct_data.get("closingDate"))

        return AccountInfo(
            account_number=acct_data.get("account", ""),
            premise_number=acct_data.get("premiseNumber", ""),
            service_address=service_address,
            nickname=acct_data.get("nickName") or None,
            is_active=acct_data.get("accountActive", "N") == "Y",
            is_default=acct_data.get("default", False),
            meters=meters,
            ebill_enrolled=acct_data.get("ebillStatus", "N") == "Y",
            closing_date=closing_date,
        )
