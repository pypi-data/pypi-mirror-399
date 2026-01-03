from __future__ import annotations

import re
from dataclasses import dataclass

from excelminer.backends.base import AnalysisContext, BackendReport
from excelminer.model.entities import FormulaCell, Sheet
from excelminer.model.graph import WorkbookGraph


_CELL_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_ ]+)!)?(?P<cell>\$?[A-Z]{1,3}\$?\d+)",
)


def _normalize_sheet_name(s: str) -> str:
    s = (s or "").strip()
    if s.startswith("'") and s.endswith("'") and len(s) >= 2:
        s = s[1:-1]
    return s


def _sheet_key(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip())


def _extract_deps(formula: str) -> dict[str, object]:
    if not formula:
        return {}

    refs: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for m in _CELL_REF_RE.finditer(formula):
        sheet = _normalize_sheet_name(m.group("sheet") or "")
        cell = (m.group("cell") or "").replace("$", "")
        key = (sheet, cell)
        if key in seen:
            continue
        seen.add(key)
        refs.append({"sheet": sheet, "cell": cell})

    return {"refs": refs} if refs else {}


@dataclass(slots=True)
class OpenpyxlBackend:
    """Semantic backend: scans worksheets for formulas and basic cell stats.

    Notes:
    - openpyxl does not evaluate formulas; it reads the formula text.
    - This backend is best used after OOXMLZipBackend has registered sheets.
    """

    name: str = "openpyxl"

    def can_handle(self, ctx: AnalysisContext) -> bool:
        p = ctx.path
        if not p.exists() or not p.is_file():
            return False
        if not (ctx.options.include_formulas or ctx.options.include_cells):
            return False
        return p.suffix.lower() in (".xlsx", ".xlsm", ".xltx", ".xltm")

    def extract(self, ctx: AnalysisContext, graph: WorkbookGraph) -> BackendReport:
        report = BackendReport(backend=self.name)

        try:
            from openpyxl import load_workbook
        except Exception as e:  # noqa: BLE001
            report.issues.append(f"openpyxl import failed: {e}")
            return report

        try:
            wb = load_workbook(filename=str(ctx.path), read_only=True, data_only=False)
        except Exception as e:  # noqa: BLE001
            report.issues.append(f"failed to load workbook: {e}")
            return report

        formula_count = 0
        sheets_scanned = 0

        for ws in wb.worksheets:
            sheets_scanned += 1
            if ctx.options.max_sheets is not None and sheets_scanned > ctx.options.max_sheets:
                break

            # Ensure Sheet node exists (OOXML backend may have created it already).
            sheet_key = _sheet_key(ws.title)
            sheet_id = f"sheet:{sheet_key}"
            sheet_node = graph.get_by_key("sheet", sheet_key)
            if not sheet_node:
                sheet_node = graph.upsert(Sheet.make(key=sheet_key, id=sheet_id, name=ws.title))

            if not ctx.options.include_formulas:
                continue

            scanned_cells = 0
            for row in ws.iter_rows():
                for cell in row:
                    scanned_cells += 1
                    if ctx.options.max_cells_per_sheet is not None and scanned_cells > ctx.options.max_cells_per_sheet:
                        break

                    v = cell.value
                    if not isinstance(v, str) or not v.startswith("="):
                        continue

                    addr = cell.coordinate
                    key = f"{ws.title}!{addr}"
                    node = graph.upsert(
                        FormulaCell.make(
                            key=key,
                            id=f"formula:{key}",
                            sheet_name=ws.title,
                            address=addr,
                            formula=v,
                            deps=_extract_deps(v),
                        )
                    )
                    graph.add_edge(sheet_node.id, node.id, "contains")
                    formula_count += 1

                if ctx.options.max_cells_per_sheet is not None and scanned_cells > ctx.options.max_cells_per_sheet:
                    break

        report.stats = {
            "sheets_scanned": sheets_scanned,
            "formula_cells": formula_count,
            **graph.stats(),
        }
        return report
