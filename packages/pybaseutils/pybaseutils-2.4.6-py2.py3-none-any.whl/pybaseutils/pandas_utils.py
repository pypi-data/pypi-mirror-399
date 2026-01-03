# -*- coding: utf-8 -*-
"""
# --------------------------------------------------------
# @Author : Pan
# @E-mail :
# @Date   : 2022-04-29 09:13:09
# @Brief  :
# --------------------------------------------------------
"""
import os
import numpy as np
import pandas as pd
import csv


class CSVWriter(object):
    def __init__(self, filename, title=[], keys=[], mode='a'):
        """
        :param filename: *.csv file
        :param title: 表头名称(第一行的表格)
        :param keys: 表头字典(默认与title一致)
        :param mode: 打开文件模式
        """
        self.filename = filename
        self.title = title
        self.keys = keys if keys else self.title
        self.file = open(self.filename, mode=mode, newline='')
        self.writer = csv.writer(self.file)
        # 检查文件是否为空，如果是，写入表头
        if self.file.tell() == 0:
            self.writer.writerow(title)
            self.file.flush()

    def add(self, data: list or dict, flush=True):
        """
        :param data:
        :param flush:
        :return:
        """
        if isinstance(data, dict): data = [data.get(k, None) for k in self.title]
        self.writer.writerow(data)
        if flush: self.file.flush()  # 强制写入磁盘
        return data

    def close(self):  # 手动关闭文件（推荐在程序结束时调用）
        self.file.close()


def read_csv(filename, sep=";"):
    """
    :param filename:
    :param sep: 分隔符
    :return:
    """
    file = pd.read_csv(filename, sep=sep)
    df = pd.DataFrame(file)
    return df


def get_cols(df, keys, to_dict=True) -> pd.Series:
    """
    获得某一列的数据
    data =  data[["image_ids","label"]]
    data =  get_cols(df, ["image_ids", "label"]).values
    :param df:
    :param keys:  list | int
    :param to_dict: 转换为字典
    :return:
    """
    data = df[keys]  # 或者
    # data = df.loc[:, keys]
    if to_dict: data = data.to_dict()
    return data


def get_rows(df, index, to_dict=True) -> pd.Series:
    """
    获得某一行的数据
    loc：基于行标签（index）获取数据。
    iloc：基于行位置（整数索引）获取数据。
    :param df:
    :param index: list | int
    :param to_dict: 转换为字典
    :return:
    """
    if isinstance(index, list):
        data = df.loc[index]
        if to_dict: data = data.to_dict(orient='index')
    else:
        data = df.loc[index]
        if to_dict: data = data.to_dict()  # 无需index
    return data


def df2list(df):
    """pandas.values转为list"""
    list_ = df.values.tolist()
    return list_


def list2df(data: list):
    return pd.DataFrame(data)


def save_csv(filename, df: pd.DataFrame, rows=True):
    """
    :param filename:
    :param df:
    :param rows:
    :return:
    """
    if rows is None: rows = True
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    df.to_csv(filename, index=rows, sep=',', header=True)


def print_info(class_name, labels):
    """
    :param class_name:
    :param labels:
    :return:
    """
    # index =range(len(class_name))+1
    index = np.arange(0, len(class_name)) + 1
    columns = ['class_name', 'labels']
    content = np.array([class_name, labels]).T
    df = pd.DataFrame(content, index=index, columns=columns)  # 生成6行4列位置
    print(df)  # 输出6行4列的表格
    save_csv("my_test.csv", df)


def data2df(data, cols=None, rows=None, file=None) -> pd.DataFrame:
    """
    将data数据转为pd.DataFrame
    :param data: 表单数据
    :param cols: (columns)表单列名称
    :param rows: (index)表单行名称
    :param file:
    :return: pd.DataFrame
    """
    df = pd.DataFrame(data, index=rows, columns=cols)  # 生成6行4列位置
    if file: save_csv(file, df, rows=rows)
    return df


def dict2df(data: dict, cols=None, T=False, file=None) -> pd.DataFrame:
    """
    :param data: 表单数据
    :param cols: (columns)表单列名称
    :param T: 是否转置表单
    :return: pd.DataFrame
    """
    if T:
        df = pd.DataFrame.from_dict(data, columns=cols)  # 键按照列进行转换
    else:
        df = pd.DataFrame.from_dict(data, columns=cols, orient='index')  # 键按照行进行转换
    if file: save_csv(file, df, rows=True)
    return df


def df2dict(df: pd.DataFrame, orient="index") -> dict:
    """
    :param df:
    :param orient: 常用的是list,index,dict,records,
    :return:
    """
    data = df.to_dict(orient=orient)
    return data


def sort_df(df: pd.DataFrame, key, ascending=True) -> pd.DataFrame:
    """按照key进行排序,默认值为 True，表示升序排序。如果设置为 False，则表示降序排序"""
    return df.sort_values(by=key, ascending=ascending)


def concat_df(df1: pd.DataFrame, df2: pd.DataFrame, key=None) -> pd.DataFrame:
    """
    合并两个DF数据,以key作为索引删除重复的Key，避免存在相同的key的数据(优先用df2的值填充df1)
    如果df2本身有重复key,则不会删除重复key；但可以连续两次重复key,如
    df = concat_df(df1,df2,key)
    df = concat_df(df,df,key) # 去除重复的key
    :param df1:
    :param df2:
    :param key: 以key作为索引删除重复的Key，避免存在相同的key的数据(优先用df2的值填充df1)
    :return:
    """
    if df1 is None or df1.empty: return df2
    if df2 is None or df2.empty: return df1
    # 获得合并后列表
    keys = df1.keys().tolist() + df2.keys().tolist()
    keys = list(dict.fromkeys(keys))
    if key:
        # BUG: DataFrame index must be unique
        df1 = df1.drop_duplicates(subset=key, keep='last')
        df2 = df2.drop_duplicates(subset=key, keep='last')
        df1 = df1.set_index(key)
        df2 = df2.set_index(key)
        # 合并数据框
        df = df2.combine_first(df1)
        # 使用原始索引顺序进行排序
        index = df1.index.tolist() + [index for index in df2.index if index not in df1.index]
        df = df.loc[index].reset_index()
    else:
        df = pd.concat([df1, df2]).reset_index()
    df = df[keys]  # 保持keys序列一致
    return df


def combine_df(df1: pd.DataFrame, df2: pd.DataFrame, key=None, cmp=None) -> pd.DataFrame:
    """
    合并两个DF数据
    :param df1:
    :param df2:
    :param key: 以key作为索引删除重复的Key，避免存在相同的key的数据(优先用df2的值填充df1)
    :return:
    """
    if df1 is None or df1.empty: return df2
    if df2 is None or df2.empty: return df1
    if not cmp: cmp = combine_dict
    # 获得合并后列表
    keys = df1.keys().tolist() + df2.keys().tolist()
    keys = list(dict.fromkeys(keys))
    if key:
        # BUG: DataFrame index must be unique
        df1 = df1.drop_duplicates(subset=key, keep='last')
        df2 = df2.drop_duplicates(subset=key, keep='last')
        df1 = df1.set_index(key)
        df2 = df2.set_index(key)
        d1 = df2dict(df1, orient="index")
        d2 = df2dict(df2, orient="index")
        # 使用df2的数据去更新df1,如果df2为空则不更新
        d1 = cmp(d1, d2)
        df = []
        for k, v in d1.items():
            v[key] = k
            df.append(v)
        df = pd.DataFrame(df)
    else:
        df = pd.concat([df1, df2]).reset_index()
    df = df[keys]  # 保持keys序列一致
    return df


def combine_dict(d1: dict, d2: dict) -> dict:
    """
    合并两个字典，用d2更新d1的数据，若为空则不更新
    :param d1:
    :param d2:
    :return:
    """
    # 使用df2的数据去更新df1,如果df2为空则不更新
    for k2, v2 in d2.items():
        if v2 or (k2 not in d1):
            d1[k2] = {**d1.get(k2, {}), **v2}
    return d1


def combine_path(d1: dict, d2: dict) -> dict:
    """
    合并两个字典，用d2更新d1的数据，若为空则不更新
    如果path的文件名相同，则进行合并更新
    :param d1:
    :param d2:
    :return:
    """
    # 使用df2的数据去更新df1,如果df2为空则不更新
    for k2, v2 in d2.items():
        if v2 or (k2 not in d1):
            v2 = {k_: v_ for k_, v_ in v2.items() if v_}  # 剔除空数据
            d1[k2] = {**d1.get(k2, {}), **v2}
    out = {}
    while d1:
        k1 = next(iter(d1.keys()))  # 第一key
        k2 = os.path.basename(k1)  # 可能存在的后面的key
        v1 = d1.pop(k1, {})
        v2 = d1.pop(k2, {})
        k = k1 if len(k1) > len(k2) else k2
        if v2:  # 当k1!=k2,v2才不为空
            if v1: v2 = {k_: v_ for k_, v_ in v2.items() if v_}  # 剔除空数据
            out[k] = {**v1, **v2}
        else:
            if not v1: v1 = out.pop(k1, {})
            if not v2: v2 = out.pop(k2, {})
            if v2: v1 = {k_: v_ for k_, v_ in v1.items() if v_}  # 剔除空数据
            out[k] = {**v2, **v1}
    return out


def df_apply(df: pd.DataFrame, func, axis=0, *args):
    """
    对DataFrame的每一行或每一列应用一个函数
    :param df:
    :param func: 应用的函数
    :param axis: 0表示按行应用，1表示按列应用
    :param args: 应用函数func的额外参数
    :return:
    """
    return df.apply(func, axis=axis, *args)


def read_merged_tables(xlsx_file, sheet="Sheet1"):
    """
    读取合并单元格的表格数据,并填充合并单元格的值
    :param xlsx_file:
    :param sheet: 工作表名称
    :return:
    """
    from openpyxl import load_workbook

    # 使用openpyxl加载工作簿
    wb = load_workbook(xlsx_file)
    ws = wb[sheet]
    # 获取所有合并单元格
    merged_cells = ws.merged_cells.ranges
    # 创建一个字典来存储合并单元格的值
    merged_values = {}
    # 对于每个合并区域，获取左上角单元格的值
    for merged_range in merged_cells:
        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = top_left_cell.value
        # 将合并区域内的所有单元格都标记为需要填充这个值
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value
    # 读取表格数据
    df = pd.read_excel(xlsx_file, sheet_name=sheet, header=0)
    # 填充合并单元格
    for idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            cell_row = idx + 2  # +2是因为第一行是表头，pandas从0开始索引
            cell_col = col_idx + 1  # +1是因为openpyxl从1开始索引

            if (cell_row, cell_col) in merged_values and pd.isna(value):
                df.iloc[idx, col_idx] = merged_values[(cell_row, cell_col)]
    return df


if __name__ == "__main__":
    df1 = pd.DataFrame({
        "path": ["image1.jpg", "image2.jpg", "image3.jpg"],
        # "label": None,
        "label": ["A1", "A2", "A3"],
        # "score": ["A1", "A2", "A3"],
    })
    df2 = pd.DataFrame({
        # "label": ["B3","B5", "B4"],
        "label": None,
        # "score": None,
        "path": ["image3.jpg", "image5.jpg", "image4.jpg"]
    })

    print("====================")
    print(concat_df(df1, df2, key="path"))
    print("====================")
    print(combine_df(df1, df2, key="path", cmp=combine_path))
