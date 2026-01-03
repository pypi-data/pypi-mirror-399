"""
VGD File Parser

Thermo Scientific VGD files use Microsoft OLE2 Compound File format.
Data is stored as Kinetic Energy and converted to Binding Energy using:
    BE = Source_Energy - KE

Supports both single-spectrum and multi-spectrum VGD files.
"""

import os
import re
import struct
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Union

import numpy as np

try:
    import olefile
except ImportError:
    olefile = None


@dataclass
class VGDMetadata:
    """Metadata extracted from VGD file OLE properties."""
    title: str = ""
    subject: str = ""
    author: str = ""
    create_time: str = ""
    saved_time: str = ""


@dataclass
class VGDAcquisitionParams:
    """Acquisition parameters from VGD file."""
    source_energy: float = 1486.68  # Default Al K-alpha
    pass_energy: Optional[float] = None
    work_function: Optional[float] = None
    dwell_time: Optional[float] = None
    periods: Optional[int] = None
    txf_coefficients: List[float] = field(default_factory=list)


@dataclass
class VGDSpectrum:
    """A single XPS spectrum from a VGD file."""
    
    # Core data
    binding_energy: np.ndarray
    kinetic_energy: np.ndarray
    intensity: np.ndarray  # Raw counts
    corrected_intensity: np.ndarray  # Normalized by dwell*periods
    
    # Spectrum info
    core_level: str = ""
    spectrum_index: int = 0
    total_spectra: int = 1
    num_points: int = 0
    
    # Energy range
    be_start: float = 0.0
    be_end: float = 0.0
    be_step: float = 0.0
    ke_start: float = 0.0
    ke_step: float = 0.0
    
    # Acquisition parameters
    source_energy: float = 1486.68
    source_label: str = ""
    pass_energy: Optional[float] = None
    work_function: Optional[float] = None
    dwell_time: Optional[float] = None
    periods: Optional[int] = None
    
    # Transmission function
    txf_applied: bool = False
    txf_coefficients: List[float] = field(default_factory=list)
    
    # Metadata
    sample_id: str = ""
    title: str = ""
    author: str = ""
    date_created: str = ""
    time_created: str = ""
    date_saved: str = ""
    time_saved: str = ""
    technique: str = "XPS"
    
    def to_dict(self) -> Dict:
        """Convert spectrum to dictionary matching KherveFitting format."""
        return {
            'Sample ID': self.sample_id,
            'Title': self.title,
            'Author': self.author,
            'Date Created': self.date_created,
            'Time Created': self.time_created,
            'Date Saved': self.date_saved,
            'Time Saved': self.time_saved,
            'Technique': self.technique,
            'Species & Transition': self.core_level,
            'Spectrum Index': str(self.spectrum_index),
            'Total Spectra': str(self.total_spectra),
            'Source Label': self.source_label,
            'Source Energy': f"{self.source_energy:.2f}",
            'Pass Energy': f"{self.pass_energy:.2f}" if self.pass_energy else 'Unknown',
            'Work Function': f"{self.work_function:.2f}" if self.work_function else 'Unknown',
            'Dwell Time': f"{self.dwell_time:.4f}" if self.dwell_time else 'Unknown',
            'Periods': str(self.periods) if self.periods else 'Unknown',
            'Number of Points': str(self.num_points),
            'BE Start': f"{self.be_start:.2f}",
            'BE End': f"{self.be_end:.2f}",
            'BE Step': f"{abs(self.be_step):.4f}",
            'KE Start': f"{self.ke_start:.2f}",
            'KE Step': f"{abs(self.ke_step):.4f}",
            'TXF Applied': 'Yes' if self.txf_applied else 'No',
            'TXF Coefficients': ', '.join([f"{c:.6f}" for c in self.txf_coefficients]) if self.txf_coefficients else 'N/A',
            # Also include raw arrays for programmatic access
            'binding_energy': self.binding_energy.tolist(),
            'kinetic_energy': self.kinetic_energy.tolist(),
            'intensity': self.intensity.tolist(),
            'corrected_intensity': self.corrected_intensity.tolist(),
        }


@dataclass  
class VGDFile:
    """
    Represents a parsed VGD file with one or more spectra.
    
    Usage:
        vgd = VGDFile("sample.vgd")
        print(f"Found {vgd.num_spectra} spectra")
        for spectrum in vgd.spectra:
            print(spectrum.core_level, spectrum.binding_energy)
    """
    
    filepath: str
    spectra: List[VGDSpectrum] = field(default_factory=list)
    metadata: VGDMetadata = field(default_factory=VGDMetadata)
    acquisition: VGDAcquisitionParams = field(default_factory=VGDAcquisitionParams)
    
    def __post_init__(self):
        if os.path.exists(self.filepath):
            self._parse()
    
    @property
    def num_spectra(self) -> int:
        return len(self.spectra)
    
    @property
    def core_level(self) -> str:
        """Primary core level name from filename."""
        return extract_core_level_name(os.path.basename(self.filepath))
    
    # Convenience accessors for single-spectrum files
    @property
    def binding_energy(self) -> np.ndarray:
        """Binding energy array (first spectrum)."""
        return self.spectra[0].binding_energy if self.spectra else np.array([])
    
    @property
    def intensity(self) -> np.ndarray:
        """Raw intensity array (first spectrum)."""
        return self.spectra[0].intensity if self.spectra else np.array([])
    
    @property
    def corrected_intensity(self) -> np.ndarray:
        """Corrected intensity array (first spectrum)."""
        return self.spectra[0].corrected_intensity if self.spectra else np.array([])
    
    def _parse(self):
        """Parse the VGD file and populate spectra."""
        parsed = parse_vgd_file(self.filepath)
        
        # Store metadata
        meta = parsed['metadata']
        self.metadata = VGDMetadata(
            title=meta['title'],
            subject=meta['subject'],
            author=meta['author'],
            create_time=meta['create_time'],
            saved_time=meta['saved_time'],
        )
        
        # Store acquisition params
        self.acquisition = VGDAcquisitionParams(
            source_energy=parsed['source_energy'],
            pass_energy=parsed['pass_energy'],
            work_function=parsed['work_fn'],
            dwell_time=parsed['dwell_time'],
            periods=parsed['periods'],
            txf_coefficients=parsed['txf_coeffs'] or [],
        )
        
        # Determine source label
        source_energy = parsed['source_energy']
        if abs(source_energy - 1486.68) < 0.1:
            source_label = "Al K-alpha Monochromated"
        elif abs(source_energy - 1253.6) < 0.1:
            source_label = "Mg K-alpha"
        else:
            source_label = f"X-ray {source_energy:.2f} eV"
        
        # Parse date/time components
        create_time = meta['create_time']
        saved_time = meta['saved_time']
        date_created = create_time.split()[0] if create_time else ''
        time_created = create_time.split()[1] if create_time and len(create_time.split()) > 1 else ''
        date_saved = saved_time.split()[0] if saved_time else ''
        time_saved = saved_time.split()[1] if saved_time and len(saved_time.split()) > 1 else ''
        
        # Create spectrum objects
        core_level = self.core_level
        num_spectra = parsed.get('num_spectra', 1)
        base_name = os.path.splitext(os.path.basename(self.filepath))[0]
        
        for idx in range(num_spectra):
            calc = calculate_spectrum_data(parsed, idx)
            
            spectrum = VGDSpectrum(
                binding_energy=np.array(calc['be_values']),
                kinetic_energy=np.array(calc['ke_values']),
                intensity=np.array(calc['intensities']),
                corrected_intensity=np.array(calc['corrected_data']),
                core_level=core_level,
                spectrum_index=idx,
                total_spectra=num_spectra,
                num_points=parsed['num_points'],
                be_start=calc['be_start'],
                be_end=calc['be_end'],
                be_step=calc['be_step'],
                ke_start=parsed['ke_start'] or 0.0,
                ke_step=parsed['ke_step'] or 0.0,
                source_energy=source_energy,
                source_label=source_label,
                pass_energy=parsed['pass_energy'],
                work_function=parsed['work_fn'],
                dwell_time=parsed['dwell_time'],
                periods=parsed['periods'],
                txf_applied=calc['txf_valid'],
                txf_coefficients=parsed['txf_coeffs'] or [],
                sample_id=meta['subject'] or base_name,
                title=meta['title'],
                author=meta['author'],
                date_created=date_created,
                time_created=time_created,
                date_saved=date_saved,
                time_saved=time_saved,
                technique='XPS',
            )
            self.spectra.append(spectrum)
    
    def to_dataframe(self):
        """
        Convert to pandas DataFrame (requires pandas).
        
        Returns DataFrame with columns: BE, Intensity, Corrected_Intensity, Spectrum_Index
        """
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("pandas is required for to_dataframe(). Install with: pip install pandas")
        
        dfs = []
        for spectrum in self.spectra:
            df = pd.DataFrame({
                'BE': spectrum.binding_energy,
                'KE': spectrum.kinetic_energy,
                'Intensity': spectrum.intensity,
                'Corrected_Intensity': spectrum.corrected_intensity,
                'Spectrum_Index': spectrum.spectrum_index,
                'Core_Level': spectrum.core_level,
            })
            dfs.append(df)
        
        return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
    
    def to_excel(self, output_path: str, include_metadata: bool = True):
        """
        Export to Excel file matching KherveFitting format.
        
        Args:
            output_path: Path for output .xlsx file
            include_metadata: If True, include experimental info in column 50
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for to_excel(). Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for spectrum in self.spectra:
            # Create sheet name
            if spectrum.spectrum_index == 0:
                sheet_name = spectrum.core_level
            else:
                sheet_name = f"{spectrum.core_level}{spectrum.spectrum_index}"
            
            ws = wb.create_sheet(sheet_name)
            
            # Headers
            ws.cell(row=1, column=1, value='BE')
            ws.cell(row=1, column=2, value='Corrected Data')
            ws.cell(row=1, column=3, value='Raw Data')
            
            # Data
            for i, (be, corr, raw) in enumerate(zip(
                    spectrum.binding_energy,
                    spectrum.corrected_intensity,
                    spectrum.intensity), start=2):
                ws.cell(row=i, column=1, value=round(float(be), 2))
                ws.cell(row=i, column=2, value=round(float(corr), 2))
                ws.cell(row=i, column=3, value=round(float(raw), 2))
            
            # Metadata in column 50 (matching KherveFitting)
            if include_metadata:
                exp_col = 50
                ws.cell(row=1, column=exp_col, value="Experimental Description")
                
                metadata_items = [
                    ('Sample ID', spectrum.sample_id),
                    ('Title', spectrum.title),
                    ('Author', spectrum.author),
                    ('Date Created', spectrum.date_created),
                    ('Time Created', spectrum.time_created),
                    ('Date Saved', spectrum.date_saved),
                    ('Time Saved', spectrum.time_saved),
                    ('Technique', spectrum.technique),
                    ('Species & Transition', spectrum.core_level),
                    ('Spectrum Index', str(spectrum.spectrum_index)),
                    ('Total Spectra', str(spectrum.total_spectra)),
                    ('Source Label', spectrum.source_label),
                    ('Source Energy', f"{spectrum.source_energy:.2f}"),
                    ('Pass Energy', f"{spectrum.pass_energy:.2f}" if spectrum.pass_energy else 'Unknown'),
                    ('Work Function', f"{spectrum.work_function:.2f}" if spectrum.work_function else 'Unknown'),
                    ('Dwell Time', f"{spectrum.dwell_time:.4f}" if spectrum.dwell_time else 'Unknown'),
                    ('Periods', str(spectrum.periods) if spectrum.periods else 'Unknown'),
                    ('Number of Points', str(spectrum.num_points)),
                    ('BE Start', f"{spectrum.be_start:.2f}"),
                    ('BE End', f"{spectrum.be_end:.2f}"),
                    ('BE Step', f"{abs(spectrum.be_step):.4f}"),
                    ('KE Start', f"{spectrum.ke_start:.2f}"),
                    ('KE Step', f"{abs(spectrum.ke_step):.4f}"),
                    ('TXF Applied', 'Yes' if spectrum.txf_applied else 'No'),
                    ('TXF Coefficients', ', '.join([f"{c:.6f}" for c in spectrum.txf_coefficients]) if spectrum.txf_coefficients else 'N/A'),
                ]
                
                for row, (key, value) in enumerate(metadata_items, start=2):
                    ws.cell(row=row, column=exp_col, value=key)
                    ws.cell(row=row, column=exp_col + 1, value=str(value))
                
                ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col)].width = 25
                ws.column_dimensions[openpyxl.utils.get_column_letter(exp_col + 1)].width = 40
        
        wb.save(output_path)


def read_vgd(filepath: str) -> VGDFile:
    """
    Read a VGD file and return a VGDFile object.
    
    This is the main entry point for parsing VGD files.
    
    Args:
        filepath: Path to the .vgd file
        
    Returns:
        VGDFile object with parsed spectra
        
    Example:
        data = read_vgd("O1s_Scan.vgd")
        print(data.binding_energy)
        print(data.intensity)
    """
    if olefile is None:
        raise ImportError(
            "The 'olefile' library is required to read VGD files.\n"
            "Install it with: pip install olefile"
        )
    
    return VGDFile(filepath)


def extract_core_level_name(filename: str) -> str:
    """
    Extract clean core level name from filename.

    Examples:
        'O1s_Scan.VGD' -> 'O1s'
        'O1s Scan.VGD' -> 'O1s'
        'Sr3d Scan.VGD' -> 'Sr3d'
        'Ni2p core level.VGD' -> 'Ni2p'
        'C1s_Region.VGD' -> 'C1s'
    """
    base_name = os.path.splitext(filename)[0]

    # Remove common suffixes (case insensitive)
    suffixes_to_remove = [
        '_Scan', '_scan', '_SCAN',
        ' Scan', ' scan', ' SCAN',
        '_Region', '_region', '_REGION',
        ' Region', ' region', ' REGION',
        '_core level', '_Core Level', '_Core level',
        ' core level', ' Core Level', ' Core level',
        '_spectrum', '_Spectrum', ' spectrum', ' Spectrum'
    ]

    result = base_name
    for suffix in suffixes_to_remove:
        if result.endswith(suffix):
            result = result[:-len(suffix)]
            break

    # Also try regex to catch variations like "O1s scan" or "Ni2p_scan"
    match = re.match(r'^([A-Z][a-z]?\d+[spdfgh]\d*)[\s_]', result + ' ')
    if match:
        result = match.group(1)

    return result.strip()


def parse_vgd_file(file_path: str) -> Dict:
    """
    Parse a VGD file and return extracted data as a dictionary.
    
    Handles both single-spectrum and multi-spectrum VGD files.

    Args:
        file_path: Path to the .vgd file

    Returns:
        dict with keys: 
            - intensities: List or list of lists (for multi-spectrum)
            - ke_start: Kinetic energy start value
            - ke_step: Kinetic energy step
            - num_points: Points per spectrum
            - num_spectra: Number of spectra in file
            - source_energy: X-ray source energy (eV)
            - txf_coeffs: Transmission function coefficients
            - pass_energy: Analyzer pass energy
            - work_fn: Work function
            - dwell_time: Dwell time per point
            - periods: Number of scan periods
            - metadata: dict with title, subject, author, times
    """
    if olefile is None:
        raise ImportError(
            "The 'olefile' library is required to read VGD files.\n"
            "Install it with: pip install olefile"
        )

    ole = olefile.OleFileIO(file_path)

    # Get metadata from OLE properties
    metadata = ole.get_metadata()
    title = metadata.title.split('\x00')[0] if metadata.title else "Unknown"
    subject = metadata.subject.split('\x00')[0] if metadata.subject else ""
    author = metadata.author.split('\x00')[0] if metadata.author else ""
    create_time = str(metadata.create_time) if metadata.create_time else ""
    saved_time = str(metadata.last_saved_time) if metadata.last_saved_time else ""

    # Read intensity data from VGData stream
    vgdata = ole.openstream('VGData').read()
    total_points = len(vgdata) // 8

    # Check VGDataAxes for multi-spectrum info
    data_axes = ole.openstream('VGDataAxes').read()

    num_spectra = 1
    points_per_spectrum = total_points

    if len(data_axes) >= 32:
        dim1 = struct.unpack('<i', data_axes[12:16])[0] + 1  # Points per spectrum
        dim2 = struct.unpack('<i', data_axes[28:32])[0] + 1  # Number of spectra

        # Validate dimensions
        if dim1 * dim2 == total_points and dim2 > 1:
            num_spectra = dim2
            points_per_spectrum = dim1

    # Parse all intensities
    all_intensities = []
    for i in range(total_points):
        val = struct.unpack('<d', vgdata[i * 8:i * 8 + 8])[0]
        all_intensities.append(val)

    # Reshape into list of spectra if multi-spectrum
    if num_spectra > 1:
        intensities_2d = []
        for s in range(num_spectra):
            start_idx = s * points_per_spectrum
            end_idx = start_idx + points_per_spectrum
            intensities_2d.append(all_intensities[start_idx:end_idx])
        intensities = intensities_2d
    else:
        intensities = all_intensities

    # Extract KE range from VGSpaceAxes
    space_axes = ole.openstream('VGSpaceAxes').read()
    ke_start = struct.unpack('<d', space_axes[30:38])[0] if len(space_axes) >= 46 else None
    ke_step = struct.unpack('<d', space_axes[38:46])[0] if len(space_axes) >= 46 else None

    # Extract acquisition parameters from property stream
    prop_stream_name = '\x05Q5nw4m3lIjudbfwyAayojlptCa'
    prop_data = ole.openstream(prop_stream_name).read()

    # Extract X-ray source energy
    source_energy = None
    for i in range(0, len(prop_data) - 4, 4):
        val = struct.unpack('<f', prop_data[i:i + 4])[0]
        if 1480 < val < 1490:
            source_energy = val
            break
    if source_energy is None:
        source_energy = 1486.68

    # Extract Pass Energy and use its offset as reference
    pass_energy = None
    pe_offset = None
    for i in range(0, len(prop_data) - 4, 4):
        val = struct.unpack('<f', prop_data[i:i + 4])[0]
        if val in [10.0, 20.0, 35.0, 50.0, 100.0, 160.0, 200.0]:
            pass_energy = val
            pe_offset = i
            break

    # Dwell Time is at PE offset + 8
    dwell_time = None
    if pe_offset is not None and pe_offset + 12 <= len(prop_data):
        dwell_time = struct.unpack('<f', prop_data[pe_offset + 8:pe_offset + 12])[0]
        if not (0.001 < dwell_time < 10.0):
            dwell_time = None

    # Periods is at PE offset - 32
    periods = None
    if pe_offset is not None and pe_offset >= 32:
        periods = struct.unpack('<i', prop_data[pe_offset - 32:pe_offset - 28])[0]
        if not (1 <= periods <= 1000):
            periods = None

    # Work Function is at PE offset + 16
    work_fn = None
    if pe_offset is not None and pe_offset + 20 <= len(prop_data):
        work_fn = struct.unpack('<f', prop_data[pe_offset + 16:pe_offset + 20])[0]
        if not (3.0 < work_fn < 6.0):
            work_fn = None

    # Extract Transmission Function coefficients
    txf_coeffs = []
    for start in range(3100, min(3800, len(prop_data) - 32), 4):
        val = struct.unpack('<f', prop_data[start:start + 4])[0]
        if 4.0 < val < 4.5:
            vals = []
            valid = True
            for j in range(4):
                off = start + j * 8
                if off + 4 <= len(prop_data):
                    v = struct.unpack('<f', prop_data[off:off + 4])[0]
                    vals.append(v)
                else:
                    valid = False
                    break
            if valid and len(vals) == 4 and 0.5 < vals[1] < 1.0:
                txf_coeffs = vals
                break

    ole.close()

    return {
        'intensities': intensities,
        'ke_start': ke_start,
        'ke_step': ke_step,
        'num_points': points_per_spectrum,
        'total_points': total_points,
        'num_spectra': num_spectra,
        'source_energy': source_energy,
        'txf_coeffs': txf_coeffs,
        'pass_energy': pass_energy,
        'work_fn': work_fn,
        'dwell_time': dwell_time,
        'periods': periods,
        'metadata': {
            'title': title,
            'subject': subject,
            'author': author,
            'create_time': create_time,
            'saved_time': saved_time
        }
    }


def calculate_spectrum_data(parsed_data: Dict, spectrum_index: int = 0) -> Dict:
    """
    Calculate BE values and corrected data from parsed VGD data.

    Args:
        parsed_data: Output from parse_vgd_file
        spectrum_index: For multi-spectrum files, which spectrum to process (0-indexed)

    Returns:
        dict with keys: be_values, ke_values, corrected_data, intensities,
                       be_start, be_end, be_step, txf_valid
    """
    num_spectra = parsed_data.get('num_spectra', 1)

    # Get intensities for the specified spectrum
    if num_spectra > 1:
        intensities = parsed_data['intensities'][spectrum_index]
    else:
        intensities = parsed_data['intensities']

    ke_start = parsed_data['ke_start']
    ke_step = parsed_data['ke_step']
    num_points = parsed_data['num_points']
    source_energy = parsed_data['source_energy']

    # Calculate KE values (increasing)
    ke_values = [ke_start + i * ke_step for i in range(num_points)]

    # Convert to BE (decreasing - high to low, correct for XPS)
    be_values = [source_energy - ke for ke in ke_values]

    be_start = be_values[0]
    be_end = be_values[-1]
    be_step = (be_end - be_start) / (num_points - 1) if num_points > 1 else 0

    # Calculate corrected data: Raw Data / (Periods * Dwell Time)
    dwell_time = parsed_data.get('dwell_time')
    periods = parsed_data.get('periods')
    txf_valid = True

    if dwell_time and periods and dwell_time > 0 and periods > 0:
        correction_factor = periods * dwell_time
        corrected_data = [intensity / correction_factor for intensity in intensities]
    else:
        corrected_data = list(intensities)
        txf_valid = False

    return {
        'be_values': be_values,
        'ke_values': ke_values,
        'corrected_data': corrected_data,
        'intensities': list(intensities),
        'be_start': be_start,
        'be_end': be_end,
        'be_step': be_step,
        'txf_valid': txf_valid
    }
