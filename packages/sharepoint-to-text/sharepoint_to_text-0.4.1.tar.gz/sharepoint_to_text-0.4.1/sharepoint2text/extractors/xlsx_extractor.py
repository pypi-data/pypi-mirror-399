"""
XLSX content extractor using only the openpyxl library.
"""

import datetime
import io
import logging
from typing import Any, Dict, Generator, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from sharepoint2text.extractors.data_types import XlsxContent, XlsxMetadata, XlsxSheet

logger = logging.getLogger(__name__)


def _read_metadata(file_like: io.BytesIO) -> XlsxMetadata:
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)
    props = wb.properties

    metadata = XlsxMetadata(
        title=props.title or "",
        description=props.description or "",
        creator=props.creator or "",
        last_modified_by=props.lastModifiedBy or "",
        created=(
            props.created.isoformat()
            if isinstance(props.created, datetime.datetime)
            else ""
        ),
        modified=(
            props.modified.isoformat()
            if isinstance(props.modified, datetime.datetime)
            else ""
        ),
        keywords=props.keywords or "",
        language=props.language or "",
        revision=props.revision,
    )
    wb.close()
    return metadata


def _get_cell_value(cell_value: Any) -> Any:
    """Convert cell value to appropriate Python type."""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime.datetime):
        return cell_value.isoformat()
    if isinstance(cell_value, datetime.date):
        return cell_value.isoformat()
    if isinstance(cell_value, datetime.time):
        return cell_value.isoformat()
    return cell_value


def _format_value_for_display(value: Any) -> str:
    """Format a value for text display."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value)


def _find_last_data_column(rows: List[tuple]) -> int:
    """Find the index of the last column that contains any data."""
    if not rows:
        return 0

    max_col = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            val = row[i]
            if val is not None and (not isinstance(val, str) or val.strip() != ""):
                max_col = max(max_col, i + 1)
                break
    return max_col


def _find_last_data_row(rows: List[tuple]) -> int:
    """Find the index of the last row that contains any data."""
    if not rows:
        return 0

    for i in range(len(rows) - 1, -1, -1):
        row = rows[i]
        for val in row:
            if val is not None and (not isinstance(val, str) or val.strip() != ""):
                return i + 1
    return 0


def _read_sheet_data(ws: Worksheet) -> tuple[List[Dict[str, Any]], List[List[Any]]]:
    """
    Read sheet data and return both records format and raw rows.

    Returns:
        Tuple of (records as list of dicts, raw rows including header)
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Trim trailing empty rows and columns
    last_row = _find_last_data_row(rows)
    rows = rows[:last_row]
    if not rows:
        return [], []

    last_col = _find_last_data_column(rows)
    rows = [row[:last_col] for row in rows]

    # First row is the header
    header_row = rows[0]

    # Generate column names, using "Unnamed: N" for empty headers
    headers = []
    for i, val in enumerate(header_row):
        if val is None or (isinstance(val, str) and val.strip() == ""):
            headers.append(f"Unnamed: {i}")
        else:
            headers.append(str(val))

    # Convert remaining rows to records format
    records = []
    all_rows = [headers]

    for row in rows[1:]:
        record = {}
        row_values = []
        for i, cell_value in enumerate(row):
            if i < len(headers):
                record[headers[i]] = _get_cell_value(cell_value)
            row_values.append(_get_cell_value(cell_value))
        records.append(record)
        all_rows.append(row_values)

    return records, all_rows


def _format_sheet_as_text(all_rows: List[List[Any]]) -> str:
    """
    Format sheet data as aligned text table (similar to pandas to_string).

    Args:
        all_rows: List of rows including header row
    """
    if not all_rows:
        return ""

    # Calculate column widths
    num_cols = max(len(row) for row in all_rows) if all_rows else 0
    col_widths = [0] * num_cols

    formatted_rows = []
    for row in all_rows:
        formatted_row = []
        for i in range(num_cols):
            val = row[i] if i < len(row) else None
            formatted_val = _format_value_for_display(val)
            formatted_row.append(formatted_val)
            col_widths[i] = max(col_widths[i], len(formatted_val))
        formatted_rows.append(formatted_row)

    # Build text output with right-aligned columns
    lines = []
    for formatted_row in formatted_rows:
        cells = []
        for i, val in enumerate(formatted_row):
            cells.append(val.rjust(col_widths[i]))
        lines.append(" ".join(cells))

    return "\n".join(lines)


def _read_content(file_like: io.BytesIO) -> List[XlsxSheet]:
    logger.debug("Reading content")
    file_like.seek(0)
    wb = load_workbook(file_like, read_only=True, data_only=True)

    sheets = []
    for sheet_name in wb.sheetnames:
        logger.debug(f"Reading sheet: [{sheet_name}]")
        ws = wb[sheet_name]
        records, all_rows = _read_sheet_data(ws)
        text = _format_sheet_as_text(all_rows)
        sheets.append(
            XlsxSheet(
                name=str(sheet_name),
                data=records,
                text=text,
            )
        )

    wb.close()
    return sheets


def read_xlsx(
    file_like: io.BytesIO, path: str | None = None
) -> Generator[XlsxContent, Any, None]:
    """
    Extract all relevant content from an XLSX file.

    Args:
        file_like: A BytesIO object containing the XLSX file data.
        path: Optional file path to populate file metadata fields.

    Yields:
        MicrosoftXlsxContent dataclass with all extracted content.
    """
    sheets = _read_content(file_like)
    metadata = _read_metadata(file_like)
    metadata.populate_from_path(path)

    yield XlsxContent(metadata=metadata, sheets=sheets)
